# -*- coding: utf-8 -*-
import os, glob
import openpyxl

data_dir = r"D:\supermarket-system\data"

# List all xlsx files
files = glob.glob(os.path.join(data_dir, "*.xlsx"))
print(f"Found {len(files)} xlsx files\n")

# Read the largest one (likely the main file)
files.sort(key=lambda f: os.path.getsize(f), reverse=True)

for f in files[:3]:
    fname = os.path.basename(f)
    size = os.path.getsize(f) // 1024
    print(f"=== {fname} ({size} KB) ===")
    try:
        wb = openpyxl.load_workbook(f, read_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"  Sheet: '{sheet_name}' rows={ws.max_row} cols={ws.max_column}")
            # Print header row
            headers = []
            for cell in ws[1]:
                headers.append(str(cell.value) if cell.value else '')
            print(f"  Headers: {headers[:15]}")
            # Print first 3 data rows
            for i, row in enumerate(ws.iter_rows(min_row=2, max_row=4)):
                vals = [str(cell.value)[:30] if cell.value else '' for cell in row[:15]]
                print(f"  Row {i+2}: {vals}")
        wb.close()
    except Exception as e:
        print(f"  Error: {e}")
    print()
