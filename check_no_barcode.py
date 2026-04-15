# -*- coding: utf-8 -*-
import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('D:/supermarket-system/web/new_items.xlsx', data_only=True)
ws = wb.active

# تحليل: أصناف بدون أي باركود
items = {}
for row in range(2, ws.max_row + 1):
    code = str(ws.cell(row=row, column=1).value or '').strip()
    name = str(ws.cell(row=row, column=2).value or '').strip()
    unit = str(ws.cell(row=row, column=3).value or '').strip()
    barcode = str(ws.cell(row=row, column=4).value or '').strip()
    pack = str(ws.cell(row=row, column=5).value or '').strip()
    
    if not code:
        continue
    
    if code not in items:
        items[code] = {'name': name, 'units': [], 'has_barcode': False}
    
    items[code]['units'].append({'unit': unit, 'barcode': barcode, 'pack': pack})
    if barcode and barcode != 'None' and barcode != '':
        items[code]['has_barcode'] = True

# إحصائيات
with_bc = sum(1 for v in items.values() if v['has_barcode'])
without_bc = sum(1 for v in items.values() if not v['has_barcode'])

print(f"=== إحصائيات الباركود ===")
print(f"  إجمالي الأصناف: {len(items)}")
print(f"  أصناف لها باركود (ولو وحدة واحدة): {with_bc}")
print(f"  أصناف بدون أي باركود: {without_bc}")

if without_bc > 0:
    print(f"\n=== الأصناف بدون باركود ({without_bc}) ===")
    count = 0
    for code in sorted(items.keys()):
        v = items[code]
        if not v['has_barcode']:
            units_str = ' + '.join([f"{u['unit']}(عبوة:{u['pack']})" for u in v['units']])
            print(f"  {code} | {v['name']} | {units_str}")
            count += 1

# تحليل الوحدات بدون باركود
print(f"\n=== وحدات بدون باركود (لأصناف لها باركود على الأقل) ===")
units_no_bc = 0
units_with_bc = 0
for v in items.values():
    for u in v['units']:
        if u['barcode'] and u['barcode'] != 'None':
            units_with_bc += 1
        else:
            units_no_bc += 1

print(f"  وحدات لها باركود: {units_with_bc}")
print(f"  وحدات بدون باركود: {units_no_bc}")
print(f"  النسبة بدون باركود: {units_no_bc * 100 // (units_with_bc + units_no_bc)}%")
