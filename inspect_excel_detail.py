# -*- coding: utf-8 -*-
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

path = r"C:\Users\qbas\Desktop\مجلد جديد\نسخة من اصناف_مواد_غذائية(1).xlsx"
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

print(f"الشيتات: {wb.sheetnames}")
print()

for sname in wb.sheetnames:
    ws = wb[sname]
    print(f"{'='*70}")
    print(f"شيت: {sname} | rows={ws.max_row} | cols={ws.max_column}")
    print(f"{'='*70}")
    
    # Headers
    row1 = list(ws.iter_rows(min_row=1, max_row=1))[0]
    headers = [(i, str(c.value)) for i, c in enumerate(row1) if c.value]
    print("الأعمدة:")
    for i, h in headers:
        print(f"  {chr(65+i) if i<26 else str(i+1):>3} ({i+1}): {h}")
    
    # First 5 data rows - full details
    print("\nأول 5 صفوف:")
    for ri, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True)):
        if not any(row[:5]):
            continue
        vals = {}
        for i, h in headers:
            v = row[i] if i < len(row) else None
            if v is not None:
                vals[h] = v
        print(f"  صف {ri+2}: {vals}")
    
    # Count with prices
    total = 0
    with_sell = 0
    with_cost = 0
    with_barcode = 0
    units = {}
    groups = set()
    
    # Determine column indices based on sheet
    for row in ws.iter_rows(min_row=2, values_only=True):
        if sname == 'Sheet1':
            name = row[4] if len(row) > 4 else None
            unit = row[5] if len(row) > 5 else None
            barcode = row[7] if len(row) > 7 else None
            cost = row[11] if len(row) > 11 else None
            sell = row[12] if len(row) > 12 else None
            group = row[0] if len(row) > 0 else None
        else:
            name = row[2] if len(row) > 2 else None
            unit = row[3] if len(row) > 3 else None
            barcode = row[4] if len(row) > 4 else None
            cost = row[8] if len(row) > 8 else None
            sell = row[9] if len(row) > 9 else None
            group = row[0] if len(row) > 0 else None
        
        if not name:
            continue
        total += 1
        if group:
            groups.add(str(group))
        if sell and isinstance(sell, (int, float)) and sell > 0:
            with_sell += 1
        if cost and isinstance(cost, (int, float)) and cost > 0:
            with_cost += 1
        if barcode and str(barcode).strip() not in ('', 'None', 'بدون'):
            with_barcode += 1
        u = str(unit).strip() if unit else 'حبه'
        units[u] = units.get(u, 0) + 1
    
    print(f"\nإحصائيات:")
    print(f"  إجمالي الأصناف: {total}")
    print(f"  بسعر بيع: {with_sell}")
    print(f"  بسعر تكلفة: {with_cost}")
    print(f"  بباركود: {with_barcode}")
    print(f"  أقسام: {sorted(groups)}")
    print(f"  الوحدات: {dict(sorted(units.items(), key=lambda x: x[1], reverse=True))}")
    print()

wb.close()
