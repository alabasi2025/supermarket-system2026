# -*- coding: utf-8 -*-
"""
🏪 نظام إدارة مشتريات السوبرماركت
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Backend: Flask + PostgreSQL
Frontend: Tailwind CSS + Alpine.js
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""

import os
import json
import secrets
import uuid
import time
import psycopg2
import psycopg2.extras
from decimal import Decimal
from datetime import datetime, timedelta
from functools import wraps
from flask import (
    Flask, render_template, request, redirect, url_for,
    flash, session, jsonify, g, send_file
)
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

# ═══════════════════════════════════════════════════════════════
# إعدادات التطبيق
# ═══════════════════════════════════════════════════════════════

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY') or 'supermarket-alabbasi-secret-key-2026'
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0  # No cache during development

# PostgreSQL Configuration
app.config['PG_HOST'] = os.environ.get('PG_HOST', 'localhost')
app.config['PG_PORT'] = os.environ.get('PG_PORT', '5432')
app.config['PG_DATABASE'] = os.environ.get('PG_DATABASE', 'supermarket')
app.config['PG_USER'] = os.environ.get('PG_USER', 'postgres')
app.config['PG_PASSWORD'] = os.environ.get('PG_PASSWORD', '774424555')

app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('SESSION_COOKIE_SECURE', '0') == '1'
app.config['UPLOAD_FOLDER'] = os.path.join(app.root_path, 'static', 'uploads')
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
app.config['DEBUG'] = os.environ.get('FLASK_DEBUG', '0') == '1'

# ═══════════════════════════════════════════════════════════════
# قاعدة البيانات PostgreSQL
# ═══════════════════════════════════════════════════════════════

class PostgreSQLWrapper:
    """Wrapper to make PostgreSQL work like SQLite's db.execute() pattern"""
    
    def __init__(self, conn):
        self.conn = conn
        self._cursor = None
    
    def execute(self, query, params=None):
        """Execute a query and return cursor-like object"""
        self._cursor = self.conn.cursor()
        if params:
            self._cursor.execute(query, params)
        else:
            self._cursor.execute(query)
        return CursorWrapper(self._cursor, self.conn)
    
    def commit(self):
        self.conn.commit()
    
    def rollback(self):
        self.conn.rollback()
    
    def close(self):
        self.conn.close()
    
    @property
    def lastrowid(self):
        return self._cursor.fetchone()['id'] if self._cursor else None

class CursorWrapper:
    """Wrapper for cursor to support fetchone/fetchall with dict-like access"""
    
    def __init__(self, cursor, conn):
        self.cursor = cursor
        self.conn = conn
        self._lastrowid = None
    
    def fetchone(self):
        row = self.cursor.fetchone()
        if row is None:
            return None
        return dict(row)
    
    def fetchall(self):
        rows = self.cursor.fetchall()
        return [dict(row) for row in rows]
    
    @property
    def lastrowid(self):
        return self._lastrowid
    
    @property
    def rowcount(self):
        return self.cursor.rowcount

def get_db():
    """الحصول على اتصال قاعدة البيانات PostgreSQL"""
    if 'db' not in g:
        conn = psycopg2.connect(
            host=app.config['PG_HOST'],
            port=app.config['PG_PORT'],
            database=app.config['PG_DATABASE'],
            user=app.config['PG_USER'],
            password=app.config['PG_PASSWORD'],
            cursor_factory=psycopg2.extras.RealDictCursor
        )
        g.db = PostgreSQLWrapper(conn)
    return g.db

def get_cursor():
    """الحصول على cursor"""
    db = get_db()
    return db.conn.cursor()

@app.teardown_appcontext
def close_db(error):
    """إغلاق الاتصال"""
    db = g.pop('db', None)
    if db is not None:
        if error is None:
            db.commit()
        db.close()

def table_exists(db, table_name):
    """التحقق من وجود جدول في قاعدة البيانات"""
    result = db.execute("""
        SELECT EXISTS (
            SELECT FROM information_schema.tables 
            WHERE table_schema = 'public' AND table_name = %s
        ) as exists
    """, (table_name,))
    return result.fetchone()['exists']

def init_db():
    """إنشاء جداول قاعدة البيانات"""
    db = get_db()
    
    # ─── المستخدمين ───
    db.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id SERIAL PRIMARY KEY ,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            display_name TEXT NOT NULL,
            role TEXT NOT NULL CHECK(role IN ('manager', 'agent', 'warehouse')),
            is_active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # ─── الموظفين (مع مواقع السكن) ───
    db.execute('''
        CREATE TABLE IF NOT EXISTS employees (
            id SERIAL PRIMARY KEY ,
            user_id INTEGER,
            name TEXT NOT NULL,
            job_title TEXT,
            phone TEXT,
            email TEXT,
            address TEXT,
            latitude REAL,
            longitude REAL,
            department TEXT,
            hire_date TEXT,
            salary REAL,
            notes TEXT,
            is_active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    ''')
    
    # ─── الموردين ───
    db.execute('''
        CREATE TABLE IF NOT EXISTS suppliers (
            id SERIAL PRIMARY KEY ,
            name TEXT NOT NULL,
            phone TEXT,
            address TEXT,
            latitude REAL,
            longitude REAL,
            notes TEXT,
            balance REAL DEFAULT 0,
            is_active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # ─── أرقام جوال المورد ───
    db.execute('''
        CREATE TABLE IF NOT EXISTS supplier_phones (
            id SERIAL PRIMARY KEY ,
            supplier_id INTEGER NOT NULL,
            phone TEXT NOT NULL,
            label TEXT DEFAULT 'جوال',
            is_primary INTEGER DEFAULT 0,
            FOREIGN KEY (supplier_id) REFERENCES suppliers(id) ON DELETE CASCADE
        )
    ''')

    # ─── حساباتك المسجلة عند المورد ───
    db.execute('''
        CREATE TABLE IF NOT EXISTS supplier_accounts (
            id SERIAL PRIMARY KEY ,
            supplier_id INTEGER NOT NULL,
            account_name TEXT NOT NULL,
            account_number TEXT,
            notes TEXT,
            is_active INTEGER DEFAULT 1,
            FOREIGN KEY (supplier_id) REFERENCES suppliers(id) ON DELETE CASCADE
        )
    ''')

    # ─── المنافسين ───
    db.execute('''
        CREATE TABLE IF NOT EXISTS competitors (
            id SERIAL PRIMARY KEY ,
            name TEXT NOT NULL,
            type TEXT DEFAULT 'سوبرماركت',
            phone TEXT,
            address TEXT,
            latitude REAL,
            longitude REAL,
            notes TEXT,
            is_active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # ─── المخازن ───
    db.execute('''
        CREATE TABLE IF NOT EXISTS warehouses (
            id SERIAL PRIMARY KEY ,
            name TEXT NOT NULL,
            address TEXT,
            latitude REAL,
            longitude REAL,
            manager_id INTEGER,
            capacity TEXT,
            notes TEXT,
            is_active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (manager_id) REFERENCES employees(id)
        )
    ''')
    
    # ─── الفروع/المتاجر ───
    db.execute('''
        CREATE TABLE IF NOT EXISTS stores (
            id SERIAL PRIMARY KEY ,
            name TEXT NOT NULL,
            address TEXT,
            phone TEXT,
            latitude REAL,
            longitude REAL,
            manager_id INTEGER,
            working_hours TEXT,
            notes TEXT,
            is_active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (manager_id) REFERENCES employees(id)
        )
    ''')
    
    # ─── الإشعارات ───
    db.execute('''
        CREATE TABLE IF NOT EXISTS notifications (
            id SERIAL PRIMARY KEY ,
            user_id INTEGER,
            title TEXT NOT NULL,
            message TEXT,
            type TEXT DEFAULT 'info' CHECK(type IN ('info', 'warning', 'error', 'success')),
            is_read INTEGER DEFAULT 0,
            link TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    ''')
    
    # ─── سجل التدقيق ───
    db.execute('''
        CREATE TABLE IF NOT EXISTS audit_log (
            id SERIAL PRIMARY KEY ,
            user_id INTEGER,
            action TEXT NOT NULL,
            table_name TEXT,
            record_id INTEGER,
            old_data TEXT,
            new_data TEXT,
            ip_address TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    ''')
    
    # ─── الوحدات ───
    db.execute('''
        CREATE TABLE IF NOT EXISTS units (
            id SERIAL PRIMARY KEY ,
            name TEXT NOT NULL UNIQUE,
            symbol TEXT,
            is_active INTEGER DEFAULT 1
        )
    ''')
    
    # ─── الأقسام ───
    db.execute('''
        CREATE TABLE IF NOT EXISTS categories (
            id SERIAL PRIMARY KEY ,
            name TEXT NOT NULL,
            parent_id INTEGER,
            sort_order INTEGER DEFAULT 0,
            FOREIGN KEY (parent_id) REFERENCES categories(id)
        )
    ''')
    
    # ─── الأصناف الداخلية (أسماؤك) ───
    db.execute('''
        CREATE TABLE IF NOT EXISTS products (
            id SERIAL PRIMARY KEY ,
            name TEXT NOT NULL,
            name_en TEXT,
            receipt_name TEXT,
            barcode TEXT,
            product_code TEXT,
            brand TEXT,
            category_id INTEGER,
            unit TEXT DEFAULT 'قطعة',
            pack_size INTEGER DEFAULT 1,
            cost_price REAL DEFAULT 0,
            sell_price REAL DEFAULT 0,
            min_stock REAL DEFAULT 0,
            current_stock REAL DEFAULT 0,
            notes TEXT,
            is_active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (category_id) REFERENCES categories(id)
        )
    ''')
    
    # ─── وحدات الصنف المتعددة ───
    db.execute('''
        CREATE TABLE IF NOT EXISTS product_units (
            id SERIAL PRIMARY KEY ,
            product_id INTEGER NOT NULL,
            unit_id INTEGER NOT NULL,
            conversion_factor REAL DEFAULT 1,
            barcode TEXT,
            price REAL DEFAULT 0,
            is_default INTEGER DEFAULT 0,
            FOREIGN KEY (product_id) REFERENCES products(id) ON DELETE CASCADE,
            FOREIGN KEY (unit_id) REFERENCES units(id),
            UNIQUE(product_id, unit_id)
        )
    ''')
    
    # ─── ربط أسماء المورد بأسمائك ───
    db.execute('''
        CREATE TABLE IF NOT EXISTS supplier_products (
            id SERIAL PRIMARY KEY ,
            supplier_id INTEGER NOT NULL,
            supplier_product_name TEXT NOT NULL,
            supplier_product_code TEXT,
            supplier_barcode TEXT,
            supplier_unit TEXT DEFAULT 'كرتون',
            product_id INTEGER,
            pack_size INTEGER DEFAULT 1,
            purchase_price REAL DEFAULT 0,
            min_order_qty INTEGER DEFAULT 1,
            is_active INTEGER DEFAULT 1,
            notes TEXT,
            FOREIGN KEY (supplier_id) REFERENCES suppliers(id),
            FOREIGN KEY (product_id) REFERENCES products(id),
            UNIQUE(supplier_id, supplier_product_name)
        )
    ''')
    # migration: جعل product_id اختيارياً في قواعد البيانات القديمة
    try:
        cols = [r[1] for r in db.execute('PRAGMA table_info(supplier_products)').fetchall()]
        if 'product_id' in cols:
            db.execute('''
                CREATE TABLE IF NOT EXISTS supplier_products_new (
                    id SERIAL PRIMARY KEY ,
                    supplier_id INTEGER NOT NULL,
                    supplier_product_name TEXT NOT NULL,
                    supplier_product_code TEXT,
                    supplier_barcode TEXT,
                    supplier_unit TEXT DEFAULT 'كرتون',
                    product_id INTEGER,
                    pack_size INTEGER DEFAULT 1,
                    purchase_price REAL DEFAULT 0,
                    min_order_qty INTEGER DEFAULT 1,
                    is_active INTEGER DEFAULT 1,
                    notes TEXT,
                    FOREIGN KEY (supplier_id) REFERENCES suppliers(id),
                    FOREIGN KEY (product_id) REFERENCES products(id),
                    UNIQUE(supplier_id, supplier_product_name)
                )
            ''')
            existing = db.execute('SELECT name FROM sqlite_master WHERE type="table" AND name="supplier_products_new"').fetchone()
            if existing:
                db.execute('INSERT OR IGNORE INTO supplier_products_new SELECT * FROM supplier_products')
                db.execute('DROP TABLE supplier_products')
                db.execute('ALTER TABLE supplier_products_new RENAME TO supplier_products')
                db.commit()
    except Exception:
        pass
    
    # ─── إعدادات الفرز (صنف المورد يتفرز لأصناف محددة فقط) ───
    db.execute('''
        CREATE TABLE IF NOT EXISTS sorting_rules (
            id SERIAL PRIMARY KEY ,
            supplier_product_id INTEGER NOT NULL,
            allowed_product_id INTEGER NOT NULL,
            FOREIGN KEY (supplier_product_id) REFERENCES supplier_products(id) ON DELETE CASCADE,
            FOREIGN KEY (allowed_product_id) REFERENCES products(id),
            UNIQUE(supplier_product_id, allowed_product_id)
        )
    ''')
    
    # ─── أسعار الموردين ───
    db.execute('''
        CREATE TABLE IF NOT EXISTS supplier_prices (
            id SERIAL PRIMARY KEY ,
            supplier_id INTEGER NOT NULL,
            product_id INTEGER NOT NULL,
            unit_price REAL NOT NULL,
            pack_price REAL,
            pack_size INTEGER DEFAULT 1,
            effective_date TEXT DEFAULT CURRENT_TIMESTAMP,
            notes TEXT,
            FOREIGN KEY (supplier_id) REFERENCES suppliers(id),
            FOREIGN KEY (product_id) REFERENCES products(id)
        )
    ''')
    
    # ─── المنافسين ───
    db.execute('''
        CREATE TABLE IF NOT EXISTS competitors (
            id SERIAL PRIMARY KEY ,
            name TEXT NOT NULL,
            address TEXT,
            notes TEXT,
            is_active INTEGER DEFAULT 1
        )
    ''')
    
    # ─── أسعار المنافسين ───
    db.execute('''
        CREATE TABLE IF NOT EXISTS competitor_prices (
            id SERIAL PRIMARY KEY ,
            competitor_id INTEGER NOT NULL,
            product_id INTEGER NOT NULL,
            price REAL NOT NULL,
            recorded_date TEXT DEFAULT CURRENT_TIMESTAMP,
            notes TEXT,
            FOREIGN KEY (competitor_id) REFERENCES competitors(id),
            FOREIGN KEY (product_id) REFERENCES products(id)
        )
    ''')
    
    # ─── تسعيرات البيع (عدة تسعيرات لكل صنف) ───
    db.execute('''
        CREATE TABLE IF NOT EXISTS price_lists (
            id SERIAL PRIMARY KEY ,
            name TEXT NOT NULL,
            description TEXT,
            is_default INTEGER DEFAULT 0
        )
    ''')
    
    db.execute('''
        CREATE TABLE IF NOT EXISTS product_prices (
            id SERIAL PRIMARY KEY ,
            product_id INTEGER NOT NULL,
            price_list_id INTEGER NOT NULL,
            price REAL NOT NULL,
            FOREIGN KEY (product_id) REFERENCES products(id),
            FOREIGN KEY (price_list_id) REFERENCES price_lists(id),
            UNIQUE(product_id, price_list_id)
        )
    ''')
    
    # ─── فواتير المورد (لا تتعدل) ───
    db.execute('''
        CREATE TABLE IF NOT EXISTS supplier_invoices (
            id SERIAL PRIMARY KEY ,
            invoice_number TEXT UNIQUE,
            supplier_id INTEGER NOT NULL,
            invoice_date TEXT NOT NULL,
            total_amount REAL DEFAULT 0,
            paid_amount REAL DEFAULT 0,
            status TEXT DEFAULT 'pending' CHECK(status IN ('pending', 'sorting', 'delivered', 'received')),
            notes TEXT,
            created_by INTEGER NOT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (supplier_id) REFERENCES suppliers(id),
            FOREIGN KEY (created_by) REFERENCES users(id)
        )
    ''')
    
    # ─── بنود فاتورة المورد ───
    db.execute('''
        CREATE TABLE IF NOT EXISTS supplier_invoice_items (
            id SERIAL PRIMARY KEY ,
            invoice_id INTEGER NOT NULL,
            supplier_product_id INTEGER NOT NULL,
            quantity REAL NOT NULL,
            unit_price REAL NOT NULL,
            total_price REAL NOT NULL,
            sorted_quantity REAL DEFAULT 0,
            FOREIGN KEY (invoice_id) REFERENCES supplier_invoices(id) ON DELETE CASCADE,
            FOREIGN KEY (supplier_product_id) REFERENCES supplier_products(id)
        )
    ''')
    
    # ─── فاتورة المندوب (الفرز) ───
    db.execute('''
        CREATE TABLE IF NOT EXISTS agent_invoices (
            id SERIAL PRIMARY KEY ,
            supplier_invoice_id INTEGER NOT NULL,
            agent_id INTEGER NOT NULL,
            status TEXT DEFAULT 'draft' CHECK(status IN ('draft', 'submitted', 'received')),
            submitted_at TEXT,
            received_at TEXT,
            received_by INTEGER,
            notes TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (supplier_invoice_id) REFERENCES supplier_invoices(id),
            FOREIGN KEY (agent_id) REFERENCES users(id),
            FOREIGN KEY (received_by) REFERENCES users(id)
        )
    ''')
    
    # ─── بنود فاتورة المندوب (التفريز) ───
    db.execute('''
        CREATE TABLE IF NOT EXISTS agent_invoice_items (
            id SERIAL PRIMARY KEY ,
            agent_invoice_id INTEGER NOT NULL,
            supplier_invoice_item_id INTEGER NOT NULL,
            product_id INTEGER NOT NULL,
            quantity REAL NOT NULL,
            notes TEXT,
            FOREIGN KEY (agent_invoice_id) REFERENCES agent_invoices(id) ON DELETE CASCADE,
            FOREIGN KEY (supplier_invoice_item_id) REFERENCES supplier_invoice_items(id),
            FOREIGN KEY (product_id) REFERENCES products(id)
        )
    ''')
    
    # ─── حركات المخزون ───
    db.execute('''
        CREATE TABLE IF NOT EXISTS inventory_movements (
            id SERIAL PRIMARY KEY ,
            product_id INTEGER NOT NULL,
            batch_id INTEGER,
            movement_type TEXT NOT NULL CHECK(movement_type IN ('in', 'out', 'adjust')),
            quantity REAL NOT NULL,
            reference_type TEXT,
            reference_id INTEGER,
            notes TEXT,
            created_by INTEGER,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (product_id) REFERENCES products(id),
            FOREIGN KEY (batch_id) REFERENCES product_batches(id),
            FOREIGN KEY (created_by) REFERENCES users(id)
        )
    ''')
    
    # ─── دفعات المنتجات (تتبع تواريخ الإنتاج والانتهاء) ───
    db.execute('''
        CREATE TABLE IF NOT EXISTS product_batches (
            id SERIAL PRIMARY KEY ,
            product_id INTEGER NOT NULL,
            batch_number TEXT,
            production_date TEXT,
            expiry_date TEXT,
            quantity_received REAL NOT NULL DEFAULT 0,
            quantity_remaining REAL NOT NULL DEFAULT 0,
            purchase_price REAL,
            supplier_invoice_id INTEGER,
            supplier_invoice_item_id INTEGER,
            location TEXT,
            status TEXT DEFAULT 'active' CHECK(status IN ('active', 'expired', 'consumed', 'returned')),
            notes TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (product_id) REFERENCES products(id),
            FOREIGN KEY (supplier_invoice_id) REFERENCES supplier_invoices(id),
            FOREIGN KEY (supplier_invoice_item_id) REFERENCES supplier_invoice_items(id)
        )
    ''')
    
    # ─── تنبيهات انتهاء الصلاحية ───
    db.execute('''
        CREATE TABLE IF NOT EXISTS expiry_alerts (
            id SERIAL PRIMARY KEY ,
            batch_id INTEGER NOT NULL,
            alert_type TEXT NOT NULL CHECK(alert_type IN ('approaching', 'expired', 'low_stock')),
            alert_date TEXT NOT NULL,
            is_handled INTEGER DEFAULT 0,
            handled_by INTEGER,
            handled_at TEXT,
            action_taken TEXT,
            notes TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (batch_id) REFERENCES product_batches(id),
            FOREIGN KEY (handled_by) REFERENCES users(id)
        )
    ''')
    
    # ─── فواتير نقاط البيع ───
    db.execute('''
        CREATE TABLE IF NOT EXISTS pos_invoices (
            id SERIAL PRIMARY KEY ,
            invoice_number TEXT UNIQUE NOT NULL,
            total_amount REAL NOT NULL DEFAULT 0,
            discount REAL DEFAULT 0,
            tax REAL DEFAULT 0,
            payment_method TEXT DEFAULT 'cash' CHECK(payment_method IN ('cash', 'card', 'bank', 'credit', 'mixed')),
            paid_amount REAL DEFAULT 0,
            change_amount REAL DEFAULT 0,
            customer_name TEXT,
            customer_phone TEXT,
            customer_id INTEGER,
            shift_id INTEGER,
            notes TEXT,
            status TEXT DEFAULT 'completed' CHECK(status IN ('completed', 'cancelled', 'refunded')),
            created_by INTEGER,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (customer_id) REFERENCES customers(id),
            FOREIGN KEY (created_by) REFERENCES users(id)
        )
    ''')
    
    # ─── بنود فواتير نقاط البيع ───
    db.execute('''
        CREATE TABLE IF NOT EXISTS pos_invoice_items (
            id SERIAL PRIMARY KEY ,
            invoice_id INTEGER NOT NULL,
            product_id INTEGER,
            product_name TEXT NOT NULL,
            quantity REAL NOT NULL DEFAULT 1,
            price REAL NOT NULL DEFAULT 0,
            discount REAL DEFAULT 0,
            total REAL NOT NULL DEFAULT 0,
            batch_id INTEGER,
            FOREIGN KEY (invoice_id) REFERENCES pos_invoices(id),
            FOREIGN KEY (product_id) REFERENCES products(id),
            FOREIGN KEY (batch_id) REFERENCES product_batches(id)
        )
    ''')
    
    # ─── ورديات الكاشير ───
    db.execute('''
        CREATE TABLE IF NOT EXISTS pos_shifts (
            id SERIAL PRIMARY KEY ,
            user_id INTEGER NOT NULL,
            start_time TEXT NOT NULL,
            end_time TEXT,
            opening_balance REAL DEFAULT 0,
            closing_balance REAL,
            total_cash REAL DEFAULT 0,
            total_card REAL DEFAULT 0,
            total_sales REAL DEFAULT 0,
            total_refunds REAL DEFAULT 0,
            status TEXT DEFAULT 'open' CHECK(status IN ('open', 'closed')),
            notes TEXT,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    ''')
    
    # ─── مرتجعات نقاط البيع ───
    db.execute('''
        CREATE TABLE IF NOT EXISTS pos_returns (
            id SERIAL PRIMARY KEY ,
            return_number TEXT UNIQUE NOT NULL,
            total_amount REAL NOT NULL DEFAULT 0,
            refund_method TEXT DEFAULT 'cash' CHECK(refund_method IN ('cash', 'credit')),
            created_by INTEGER,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (created_by) REFERENCES users(id)
        )
    ''')
    
    # ─── بنود المرتجعات ───
    db.execute('''
        CREATE TABLE IF NOT EXISTS pos_return_items (
            id SERIAL PRIMARY KEY ,
            return_id INTEGER NOT NULL,
            invoice_id INTEGER,
            product_id INTEGER,
            product_name TEXT NOT NULL,
            quantity REAL NOT NULL DEFAULT 1,
            price REAL NOT NULL DEFAULT 0,
            reason TEXT,
            FOREIGN KEY (return_id) REFERENCES pos_returns(id),
            FOREIGN KEY (invoice_id) REFERENCES pos_invoices(id),
            FOREIGN KEY (product_id) REFERENCES products(id)
        )
    ''')
    
    # ─── إعدادات النظام ───
    db.execute('''
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT,
            description TEXT
        )
    ''')
    
    # ─── إنشاء مستخدم افتراضي ───
    cursor = db.execute('SELECT COUNT(*) as count FROM users')
    if cursor.fetchone()['count'] == 0:
        db.execute('''
            INSERT INTO users (username, password_hash, display_name, role)
            VALUES (%s, %s, %s, %s)
        ''', ('admin', generate_password_hash('admin'), 'مدير النظام', 'manager'))
    
    # ─── إعدادات افتراضية ───
    default_settings = [
        ('store_name', 'سوبرماركت', 'اسم المتجر'),
        ('store_phone', '', 'رقم الهاتف'),
        ('store_address', '', 'العنوان'),
        ('currency', 'ريال', 'العملة'),
        ('tax_rate', '0', 'نسبة الضريبة %'),
    ]
    for key, value, desc in default_settings:
        db.execute('INSERT OR IGNORE INTO settings (key, value, description) VALUES (%s, %s, %s)', (key, value, desc))
    
    # ─── إنشاء قوائم أسعار افتراضية ───
    cursor = db.execute('SELECT COUNT(*) as count FROM price_lists')
    if cursor.fetchone()['count'] == 0:
        db.execute("INSERT INTO price_lists (name, description, is_default) VALUES ('سعر المفرق', 'السعر العادي للعملاء', 1)")
        db.execute("INSERT INTO price_lists (name, description) VALUES ('سعر الجملة', 'سعر خاص للكميات الكبيرة')")
        db.execute("INSERT INTO price_lists (name, description) VALUES ('سعر خاص', 'سعر خاص لعملاء محددين')")
    
    # ─── إنشاء وحدات افتراضية ───
    cursor = db.execute('SELECT COUNT(*) as count FROM units')
    if cursor.fetchone()['count'] == 0:
        default_units = [
            ('قطعة', 'قطعة'),
            ('كرتون', 'كرتون'),
            ('كيلو', 'كجم'),
            ('جرام', 'جم'),
            ('لتر', 'لتر'),
            ('مل', 'مل'),
            ('باكت', 'باكت'),
            ('علبة', 'علبة'),
            ('كيس', 'كيس'),
            ('شدة', 'شدة'),
            ('صندوق', 'صندوق'),
            ('شوال', 'شوال'),
            ('درزن', 'درزن'),
            ('متر', 'م'),
        ]
        for unit_name, symbol in default_units:
            db.execute("INSERT INTO units (name, symbol) VALUES (%s, %s)", (unit_name, symbol))
    
    db.commit()

# ═══════════════════════════════════════════════════════════════
# التحقق من تسجيل الدخول
# ═══════════════════════════════════════════════════════════════

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('يرجى تسجيل الدخول أولاً', 'warning')
            return redirect(url_for('login', next=request.path))
        if session.get('must_change_password'):
            return redirect(url_for('force_change_password'))
        return f(*args, **kwargs)
    return decorated_function

def get_user_permissions(user_id):
    """جلب صلاحيات المستخدم"""
    try:
        db = get_db()
        try:
            db.rollback()
        except:
            pass
        perms = db.execute('SELECT module_code, can_view, can_add, can_edit, can_delete FROM user_permissions WHERE user_id = %s', (user_id,)).fetchall()
        result = {}
        for p in perms:
            result[p['module_code']] = {
                'view': p['can_view'], 'add': p['can_add'],
                'edit': p['can_edit'], 'delete': p['can_delete']
            }
        return result
    except:
        return {}

def permission_required(module_code, action='view'):
    """التحقق من صلاحية المستخدم لصفحة/إجراء معين"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                return redirect(url_for('login'))
            if session.get('must_change_password'):
                return redirect(url_for('force_change_password'))
            
            perms = get_user_permissions(session['user_id'])
            module_perm = perms.get(module_code, {})
            
            if not module_perm.get(action, False):
                msg = 'ليس لديك صلاحية للوصول لهذه الصفحة'
                if is_api_request():
                    return jsonify({'success': False, 'message': msg}), 403
                flash(msg, 'error')
                return redirect(url_for('dashboard'))
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def is_api_request():
    return request.path.startswith('/api/') or request.is_json

def forbidden_response(message):
    if is_api_request():
        return jsonify({'success': False, 'message': message}), 403
    flash(message, 'error')
    return redirect(url_for('dashboard'))

def manager_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if session.get('must_change_password'):
            return redirect(url_for('force_change_password'))
        if session.get('role') != 'manager':
            return forbidden_response('هذه الصفحة للمدير فقط')
        return f(*args, **kwargs)
    return decorated_function

def role_required(*allowed_roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                return redirect(url_for('login'))
            
            if session.get('must_change_password'):
                return redirect(url_for('force_change_password'))

            if session.get('role') not in allowed_roles:
                return forbidden_response('ليس لديك صلاحية للوصول إلى هذه الصفحة')

            return f(*args, **kwargs)
        return decorated_function
    return decorator

def get_csrf_token():
    token = session.get('_csrf_token')
    if not token:
        token = secrets.token_urlsafe(32)
        session['_csrf_token'] = token
    return token

@app.context_processor
def inject_csrf_token():
    user_perms = {}
    if 'user_id' in session:
        user_perms = get_user_permissions(session['user_id'])
    return {'csrf_token': get_csrf_token, 'user_perms': user_perms}

def has_perm(module_code, action='view'):
    """helper للقوالب"""
    perms = get_user_permissions(session.get('user_id', 0))
    return perms.get(module_code, {}).get(action, False)

AUDIT_REDACT_TOKENS = (
    'password', 'pass', 'pin', 'token', 'secret', 'csrf',
    'authorization', 'cookie', 'session'
)

def _audit_key_is_sensitive(key):
    k = (str(key or '')).lower()
    return any(token in k for token in AUDIT_REDACT_TOKENS)

def _sanitize_for_audit(value, depth=0):
    if depth > 4:
        return '...'
    if value is None:
        return None
    if isinstance(value, (int, float, bool)):
        return value
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, str):
        return value if len(value) <= 500 else (value[:500] + '...')
    if isinstance(value, dict):
        out = {}
        count = 0
        for k, v in value.items():
            if count >= 50:
                out['__truncated__'] = True
                break
            if _audit_key_is_sensitive(k):
                out[str(k)] = '***'
            else:
                out[str(k)] = _sanitize_for_audit(v, depth + 1)
            count += 1
        return out
    if isinstance(value, (list, tuple, set)):
        arr = list(value)
        cleaned = [_sanitize_for_audit(v, depth + 1) for v in arr[:50]]
        if len(arr) > 50:
            cleaned.append('...')
        return cleaned
    return str(value)

def _request_form_as_dict():
    out = {}
    for key in request.form.keys():
        vals = request.form.getlist(key)
        out[key] = vals if len(vals) > 1 else request.form.get(key)
    return out

def _request_files_meta():
    if not request.files:
        return None
    files_meta = {}
    for key in request.files.keys():
        files = request.files.getlist(key)
        files_meta[key] = [f.filename for f in files if getattr(f, 'filename', None)]
    return files_meta

def _extract_request_body_for_audit():
    if request.method in ('GET', 'HEAD', 'OPTIONS'):
        return None
    body = None
    if request.is_json:
        body = request.get_json(silent=True)
    elif request.form:
        body = _request_form_as_dict()
    files = _request_files_meta()
    if files:
        if body is None:
            body = {}
        if not isinstance(body, dict):
            body = {'payload': body}
        body['files'] = files
    return _sanitize_for_audit(body)

def _extract_record_id_from_request():
    view_args = request.view_args or {}
    for k, v in view_args.items():
        if 'id' in str(k).lower():
            try:
                return int(v)
            except Exception:
                return None
    return None

def _classify_audit_action(status_code):
    path = request.path or ''
    method = (request.method or '').upper()
    if path == '/login' and method == 'POST':
        return 'LOGIN_SUCCESS' if status_code < 400 else 'LOGIN_FAILED'
    if path == '/logout':
        return 'LOGOUT'
    if method == 'POST':
        return 'ADD'
    if method in ('PUT', 'PATCH'):
        return 'EDIT'
    if method == 'DELETE':
        return 'DELETE'
    return 'VIEW'

def _should_skip_audit():
    path = request.path or ''
    if path.startswith('/static/') or path == '/favicon.ico':
        return True
    if request.method == 'OPTIONS':
        return True
    return False

@app.before_request
def track_request_start_time():
    g.audit_started_at = time.perf_counter()

@app.after_request
def add_no_cache_headers(response):
    if 'text/html' in response.content_type:
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response

@app.after_request
def audit_after_request(response):
    try:
        if _should_skip_audit():
            return response

        started = getattr(g, 'audit_started_at', None)
        duration_ms = None
        if started is not None:
            duration_ms = int((time.perf_counter() - started) * 1000)

        user_id = session.get('user_id')
        if getattr(g, 'logout_user_id', None):
            user_id = g.logout_user_id

        payload = {
            'method': request.method,
            'path': request.path,
            'endpoint': request.endpoint,
            'status_code': response.status_code,
            'query': _sanitize_for_audit(dict(request.args or {})),
            'body': _extract_request_body_for_audit(),
            'view_args': _sanitize_for_audit(request.view_args or {}),
            'duration_ms': duration_ms,
            'user_agent': _sanitize_for_audit(request.headers.get('User-Agent')),
            'username': session.get('username'),
            'display_name': session.get('display_name'),
            'role': session.get('role')
        }

        log_action(
            action=_classify_audit_action(response.status_code),
            table_name=request.endpoint or request.path,
            record_id=_extract_record_id_from_request(),
            old_data=None,
            new_data=payload,
            user_id_override=user_id,
            use_direct=True
        )
    except Exception:
        pass
    return response

@app.before_request
def csrf_protect():
    if request.method not in ('POST', 'PUT', 'PATCH', 'DELETE'):
        return None
    
    # استثناء بعض الصفحات من CSRF
    csrf_exempt = ['/login', '/api/pos/verify-pin', '/api/pos/invoice', '/api/pos/return', '/force-change-password']
    if any(request.path.startswith(path) for path in csrf_exempt):
        return None

    session_token = session.get('_csrf_token')
    request_token = request.headers.get('X-CSRF-Token') or request.form.get('csrf_token')

    if session_token and request_token and secrets.compare_digest(session_token, request_token):
        return None

    if is_api_request():
        return jsonify({'success': False, 'message': 'رمز الحماية غير صالح، أعد تحميل الصفحة وحاول مرة أخرى'}), 400

    flash('الطلب غير صالح، أعد تحميل الصفحة وحاول مرة أخرى', 'error')
    return redirect(url_for('login'))

# ═══════════════════════════════════════════════════════════════
# CSRF Token Refresh API
# ═══════════════════════════════════════════════════════════════

@app.route('/api/csrf-token')
def api_csrf_token():
    return jsonify({'token': get_csrf_token()})

# ═══════════════════════════════════════════════════════════════
# الصفحات الرئيسية
# ═══════════════════════════════════════════════════════════════

@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        
        db = get_db()
        # البحث بالرقم أولاً، ثم بالاسم
        user = None
        if username.isdigit():
            user = db.execute(
                'SELECT * FROM users WHERE id = %s AND is_active = TRUE',
                (int(username),)
            ).fetchone()
        if not user:
            user = db.execute(
                'SELECT * FROM users WHERE username = %s AND is_active = TRUE',
                (username,)
            ).fetchone()
        
        if user and check_password_hash(user['password_hash'], password):
            session.clear()
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['display_name'] = user['display_name']
            session['role'] = user['role']
            get_csrf_token()
            
            # Check if must change password
            if user['must_change_password']:
                session['must_change_password'] = True
                return redirect(url_for('force_change_password'))
            
            flash(f'مرحباً {user["display_name"]}', 'success')
            # إرجاع المستخدم للصفحة التي كان فيها
            next_page = request.form.get('next') or request.args.get('next') or url_for('dashboard')
            # التأكد أن الصفحة آمنة (لا تبدأ بـ http)
            if not next_page.startswith('/'):
                next_page = url_for('dashboard')
            return redirect(next_page)
        else:
            flash('اسم المستخدم أو كلمة المرور غير صحيحة', 'error')
    
    return render_template('login.html')

@app.route('/force-change-password', methods=['GET', 'POST'])
def force_change_password():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        new_password = request.form.get('new_password', '')
        confirm_password = request.form.get('confirm_password', '')
        
        if len(new_password) < 4:
            flash('كلمة المرور يجب أن تكون 4 أحرف على الأقل', 'error')
        elif new_password != confirm_password:
            flash('كلمة المرور غير متطابقة', 'error')
        else:
            db = get_db()
            user_id = session['user_id']
            db.execute('UPDATE users SET password=%s, must_change_password=0 WHERE id=%s',
                      (generate_password_hash(new_password), user_id))
            db.commit()
            
            # تحديث الجلسة
            session.pop('must_change_password', None)
            session.modified = True
            
            flash('تم تغيير كلمة المرور بنجاح', 'success')
            return redirect(url_for('dashboard'))
    
    return render_template('force_change_password.html')

@app.route('/logout')
def logout():
    g.logout_user_id = session.get('user_id')
    session.clear()
    flash('تم تسجيل الخروج', 'info')
    return redirect(url_for('login'))

# ═══════════════════════════════════════════════════════════════
# إدارة المستخدمين
# ═══════════════════════════════════════════════════════════════

@app.route('/users')
@manager_required
def users():
    db = get_db()
    users_list = db.execute('''
        SELECT * FROM users ORDER BY role, display_name
    ''').fetchall()
    return render_template('users.html', users=users_list)

@app.route('/api/users', methods=['POST'])
@manager_required
def api_users_create():
    db = get_db()
    data = request.json
    
    # التحقق من عدم تكرار اسم المستخدم
    existing = db.execute('SELECT id FROM users WHERE username = %s', (data['username'],)).fetchone()
    if existing:
        return jsonify({'success': False, 'message': 'اسم المستخدم موجود مسبقاً'})
    
    db.execute('''
        INSERT INTO users (username, password_hash, display_name, role)
        VALUES (%s, %s, %s, %s)
    ''', (
        data['username'],
        generate_password_hash(data['password_hash']),
        data['display_name'],
        data['role']
    ))
    db.commit()
    return jsonify({'success': True, 'message': 'تم إضافة المستخدم'})

@app.route('/api/users/<int:id>', methods=['PUT', 'DELETE'])
@manager_required
def api_user(id):
    db = get_db()
    
    if request.method == 'PUT':
        data = request.json
        
        # التحقق من عدم تكرار اسم المستخدم
        existing = db.execute('SELECT id FROM users WHERE username = %s AND id != %s', (data['username'], id)).fetchone()
        if existing:
            return jsonify({'success': False, 'message': 'اسم المستخدم موجود مسبقاً'})
        
        db.execute('''
            UPDATE users SET username=%s, display_name=%s, role=%s
            WHERE id=%s
        ''', (data['username'], data['display_name'], data['role'], id))
        db.commit()
        return jsonify({'success': True, 'message': 'تم تحديث المستخدم'})
    
    elif request.method == 'DELETE':
        # لا يمكن حذف المستخدم الحالي
        if id == session.get('user_id'):
            return jsonify({'success': False, 'message': 'لا يمكنك حذف نفسك'})
        
        db.execute('UPDATE users SET is_active=FALSE WHERE id=%s', (id,))
        db.commit()
        return jsonify({'success': True, 'message': 'تم تعطيل المستخدم'})

@app.route('/api/users/<int:id>/reset-password', methods=['POST'])
@manager_required
def api_user_reset_password(id):
    db = get_db()
    data = request.json
    
    db.execute('UPDATE users SET password=%s WHERE id=%s', 
               (generate_password_hash(data['password_hash']), id))
    db.commit()
    return jsonify({'success': True, 'message': 'تم تغيير كلمة المرور'})

@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST':
        data = request.json
        db = get_db()
        
        # التحقق من كلمة المرور الحالية
        user = db.execute('SELECT password FROM users WHERE id = %s', (session['user_id'],)).fetchone()
        if not check_password_hash(user['password_hash'], data['current_password']):
            return jsonify({'success': False, 'message': 'كلمة المرور الحالية غير صحيحة'})
        
        # تحديث كلمة المرور
        db.execute('UPDATE users SET password=%s WHERE id=%s',
                   (generate_password_hash(data['new_password']), session['user_id']))
        db.commit()
        return jsonify({'success': True, 'message': 'تم تغيير كلمة المرور بنجاح'})
    
    return render_template('change_password.html')

@app.route('/dashboard')
@login_required
def dashboard():
    db = get_db()
    
    # إحصائيات سريعة
    stats = {
        'suppliers': db.execute('SELECT COUNT(*) as count FROM suppliers WHERE is_active=TRUE').fetchone()['count'],
        'products': db.execute('SELECT COUNT(*) as count FROM products WHERE is_active=TRUE').fetchone()['count'],
        'low_stock': db.execute('SELECT COUNT(*) as count FROM products WHERE current_stock <= min_stock AND is_active=TRUE').fetchone()['count'],
        'pending_invoices': db.execute("SELECT COUNT(*) as count FROM supplier_invoices WHERE status IN ('pending', 'sorting')").fetchone()['count'],
        'pending_receive': db.execute("SELECT COUNT(*) as count FROM agent_invoices WHERE status = 'submitted'").fetchone()['count'],
    }
    
    # آخر الفواتير
    recent_invoices = db.execute('''
        SELECT si.*, s.name as supplier_name, u.display_name as created_by_name
        FROM supplier_invoices si
        JOIN suppliers s ON s.id = si.supplier_id
        JOIN users u ON u.id = si.created_by
        ORDER BY si.created_at DESC LIMIT 5
    ''').fetchall()
    
    # أصناف تحت الحد الأدنى
    low_stock_items = db.execute('''
        SELECT p.*, c.name as category_name
        FROM products p
        LEFT JOIN categories c ON c.id = p.category_id
        WHERE p.current_stock <= p.min_stock AND p.is_active = TRUE
        ORDER BY (p.current_stock - p.min_stock) ASC
        LIMIT 10
    ''').fetchall()
    
    return render_template('dashboard.html', 
                         stats=stats, 
                         recent_invoices=recent_invoices,
                         low_stock_items=low_stock_items)

# ═══════════════════════════════════════════════════════════════
# إدارة الموردين
# ═══════════════════════════════════════════════════════════════

@app.route('/suppliers')
@role_required('manager')
def suppliers():
    db = get_db()
    try:
        suppliers_list = db.execute('''
            SELECT s.*, 
                   COUNT(DISTINCT sp.id) as products_count,
                   COUNT(DISTINCT si.id) as invoices_count
            FROM suppliers s
            LEFT JOIN supplier_products sp ON sp.supplier_id = s.id
            LEFT JOIN supplier_invoices si ON si.supplier_id = s.id
            WHERE s.is_active = TRUE
            GROUP BY s.id
            ORDER BY s.name
        ''').fetchall()
    except Exception as e:
        import traceback
        traceback.print_exc()
        return f"Error: {e}", 500
    return render_template('suppliers.html', suppliers=suppliers_list)

@app.route('/maps')
@role_required('manager')
def maps():
    db = get_db()
    # الموردين
    suppliers_list = db.execute('''
        SELECT * FROM suppliers WHERE is_active = TRUE ORDER BY name
    ''').fetchall()
    
    # المنافسين
    competitors_list = db.execute('''
        SELECT * FROM competitors WHERE is_active = TRUE ORDER BY name
    ''').fetchall() if table_exists(db, 'competitors') else []
    
    # المخازن
    warehouses_list = db.execute('''
        SELECT * FROM warehouses WHERE is_active = TRUE ORDER BY name
    ''').fetchall() if table_exists(db, 'warehouses') else []
    
    # الفروع/المتاجر
    stores_list = db.execute('''
        SELECT * FROM stores WHERE is_active = TRUE ORDER BY name
    ''').fetchall() if table_exists(db, 'stores') else []
    
    # الموظفين
    employees_list = db.execute('''
        SELECT * FROM employees WHERE is_active = TRUE ORDER BY name
    ''').fetchall() if table_exists(db, 'employees') else []
    
    return render_template('suppliers_map.html', 
                         suppliers=suppliers_list,
                         competitors=competitors_list,
                         warehouses=warehouses_list,
                         stores=stores_list,
                         employees=employees_list)

@app.route('/api/suppliers', methods=['GET', 'POST'])
@role_required('manager')
def api_suppliers():
    db = get_db()
    
    if request.method == 'POST':
        data = request.json
        try:
            row = db.execute('''
                INSERT INTO suppliers (name, phone, address, notes, latitude, longitude)
                VALUES (%s, %s, %s, %s, %s, %s)
                RETURNING id
            ''', (data['name'], data.get('phone'), data.get('address'), data.get('notes'),
                  data.get('latitude'), data.get('longitude'))).fetchone()
            supplier_id = row['id']
            # حفظ الأرقام الإضافية
            for ph in data.get('phones', []):
                if ph.get('phone'):
                    db.execute('INSERT INTO supplier_phones (supplier_id, phone, label, is_primary) VALUES (%s,%s,%s,%s)',
                               (supplier_id, ph['phone'], ph.get('label','جوال'), ph.get('is_primary',0)))
            # حفظ الحسابات
            for acc in data.get('accounts', []):
                if acc.get('account_name'):
                    db.execute('INSERT INTO supplier_accounts (supplier_id, account_name, account_number, notes) VALUES (%s,%s,%s,%s)',
                               (supplier_id, acc['account_name'], acc.get('account_number'), acc.get('notes')))
            db.commit()
            return jsonify({'success': True, 'id': supplier_id, 'message': 'تم إضافة المورد'})
        except Exception as e:
            db.rollback()
            import traceback
            traceback.print_exc()
            return jsonify({'success': False, 'message': str(e)}), 500
    
    suppliers_list = db.execute('SELECT * FROM suppliers WHERE is_active=TRUE ORDER BY name').fetchall()
    return jsonify([dict(s) for s in suppliers_list])

@app.route('/api/suppliers/<int:id>', methods=['GET', 'PUT', 'DELETE'])
@role_required('manager')
def api_supplier(id):
    db = get_db()

    if request.method == 'GET':
        supplier = db.execute('SELECT * FROM suppliers WHERE id=%s', (id,)).fetchone()
        if not supplier:
            return jsonify({}), 404
        result = dict(supplier)
        result['phones'] = [dict(p) for p in db.execute(
            'SELECT * FROM supplier_phones WHERE supplier_id=%s ORDER BY is_primary DESC', (id,)).fetchall()]
        result['accounts'] = [dict(a) for a in db.execute(
            'SELECT * FROM supplier_accounts WHERE supplier_id=%s AND is_active=TRUE', (id,)).fetchall()]
        return jsonify(result)
    
    if request.method == 'PUT':
        data = request.json
        db.execute('''
            UPDATE suppliers SET name=%s, phone=%s, address=%s, notes=%s, latitude=%s, longitude=%s
            WHERE id=%s
        ''', (data['name'], data.get('phone'), data.get('address'), data.get('notes'),
              data.get('latitude'), data.get('longitude'), id))
        # تحديث الأرقام
        db.execute('DELETE FROM supplier_phones WHERE supplier_id=%s', (id,))
        for ph in data.get('phones', []):
            if ph.get('phone'):
                db.execute('INSERT INTO supplier_phones (supplier_id, phone, label, is_primary) VALUES (%s,%s,%s,%s)',
                           (id, ph['phone'], ph.get('label','جوال'), ph.get('is_primary',0)))
        # تحديث الحسابات
        db.execute('DELETE FROM supplier_accounts WHERE supplier_id=%s', (id,))
        for acc in data.get('accounts', []):
            if acc.get('account_name'):
                db.execute('INSERT INTO supplier_accounts (supplier_id, account_name, account_number, notes) VALUES (%s,%s,%s,%s)',
                           (id, acc['account_name'], acc.get('account_number'), acc.get('notes')))
        db.commit()
        return jsonify({'success': True, 'message': 'تم تحديث المورد'})
    
    elif request.method == 'DELETE':
        db.execute('UPDATE suppliers SET is_active=FALSE WHERE id=%s', (id,))
        db.commit()
        return jsonify({'success': True, 'message': 'تم حذف المورد'})

# ═══════════════════════════════════════════════════════════════
# إدارة المنافسين
# ═══════════════════════════════════════════════════════════════

@app.route('/api/competitors', methods=['GET', 'POST'])
@role_required('manager')
def api_competitors():
    db = get_db()
    
    if request.method == 'POST':
        data = request.json
        db.execute('''
            INSERT INTO competitors (name, type, phone, address, latitude, longitude, notes)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        ''', (data['name'], data.get('type', 'سوبرماركت'), data.get('phone'), 
              data.get('address'), data.get('latitude'), data.get('longitude'), data.get('notes')))
        db.commit()
        return jsonify({'success': True, 'message': 'تم إضافة المنافس'})
    
    competitors_list = db.execute('SELECT * FROM competitors WHERE is_active=TRUE ORDER BY name').fetchall()
    return jsonify([dict(c) for c in competitors_list])

@app.route('/api/competitors/<int:id>', methods=['PUT', 'DELETE'])
@role_required('manager')
def api_competitor(id):
    db = get_db()
    
    if request.method == 'PUT':
        data = request.json
        db.execute('''
            UPDATE competitors SET name=%s, type=%s, phone=%s, address=%s, latitude=%s, longitude=%s, notes=%s
            WHERE id=%s
        ''', (data['name'], data.get('type', 'سوبرماركت'), data.get('phone'), 
              data.get('address'), data.get('latitude'), data.get('longitude'), data.get('notes'), id))
        db.commit()
        return jsonify({'success': True, 'message': 'تم تحديث المنافس'})
    
    elif request.method == 'DELETE':
        db.execute('UPDATE competitors SET is_active=FALSE WHERE id=%s', (id,))
        db.commit()
        return jsonify({'success': True, 'message': 'تم حذف المنافس'})

# ═══════════════════════════════════════════════════════════════
# إدارة المخازن API
# ═══════════════════════════════════════════════════════════════

@app.route('/api/warehouses', methods=['GET', 'POST'])
@role_required('manager')
def api_warehouses():
    db = get_db()
    
    if request.method == 'POST':
        data = request.get_json()
        db.execute('''
            INSERT INTO warehouses (name, address, latitude, longitude)
            VALUES (%s, %s, %s, %s)
        ''', (data['name'], data.get('address'), data.get('latitude'), data.get('longitude')))
        db.commit()
        return jsonify({'success': True, 'message': 'تم إضافة المخزن'})
    
    warehouses_list = db.execute('SELECT * FROM warehouses WHERE is_active=TRUE ORDER BY name').fetchall()
    return jsonify([dict(w) for w in warehouses_list])

@app.route('/api/warehouses/<int:id>', methods=['PUT', 'DELETE'])
@role_required('manager')
def api_warehouse(id):
    db = get_db()
    
    if request.method == 'PUT':
        data = request.get_json()
        db.execute('''
            UPDATE warehouses SET name=%s, address=%s, latitude=%s, longitude=%s
            WHERE id=%s
        ''', (data['name'], data.get('address'), data.get('latitude'), data.get('longitude'), id))
        db.commit()
        return jsonify({'success': True, 'message': 'تم تحديث المخزن'})
    
    elif request.method == 'DELETE':
        db.execute('UPDATE warehouses SET is_active=FALSE WHERE id=%s', (id,))
        db.commit()
        return jsonify({'success': True, 'message': 'تم حذف المخزن'})

# ═══════════════════════════════════════════════════════════════
# إدارة الفروع API
# ═══════════════════════════════════════════════════════════════

@app.route('/api/stores', methods=['GET', 'POST'])
@role_required('manager')
def api_stores():
    db = get_db()
    
    if request.method == 'POST':
        data = request.get_json()
        db.execute('''
            INSERT INTO stores (name, address, phone, latitude, longitude)
            VALUES (%s, %s, %s, %s, %s)
        ''', (data['name'], data.get('address'), data.get('phone'), data.get('latitude'), data.get('longitude')))
        db.commit()
        return jsonify({'success': True, 'message': 'تم إضافة الفرع'})
    
    stores_list = db.execute('SELECT * FROM stores WHERE is_active=TRUE ORDER BY name').fetchall()
    return jsonify([dict(s) for s in stores_list])

@app.route('/api/stores/<int:id>', methods=['PUT', 'DELETE'])
@role_required('manager')
def api_store(id):
    db = get_db()
    
    if request.method == 'PUT':
        data = request.get_json()
        db.execute('''
            UPDATE stores SET name=%s, address=%s, phone=%s, latitude=%s, longitude=%s
            WHERE id=%s
        ''', (data['name'], data.get('address'), data.get('phone'), data.get('latitude'), data.get('longitude'), id))
        db.commit()
        return jsonify({'success': True, 'message': 'تم تحديث الفرع'})
    
    elif request.method == 'DELETE':
        db.execute('UPDATE stores SET is_active=FALSE WHERE id=%s', (id,))
        db.commit()
        return jsonify({'success': True, 'message': 'تم حذف الفرع'})

# ═══════════════════════════════════════════════════════════════
# إدارة الموظفين
# ═══════════════════════════════════════════════════════════════

@app.route('/employees')
@role_required('manager')
def employees():
    db = get_db()
    employees_list = db.execute('''
        SELECT * FROM employees WHERE is_active = TRUE ORDER BY name
    ''').fetchall()
    return render_template('employees.html', employees=employees_list)

@app.route('/api/employees', methods=['GET', 'POST'])
@role_required('manager')
def api_employees():
    db = get_db()
    
    if request.method == 'POST':
        data = request.get_json()
        db.execute('''
            INSERT INTO employees (name, job_title, phone, address, latitude, longitude, department, notes)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        ''', (data['name'], data.get('job_title'), data.get('phone'), data.get('address'), 
              data.get('latitude'), data.get('longitude'), data.get('department'), data.get('notes')))
        db.commit()
        return jsonify({'success': True, 'message': 'تم إضافة الموظف'})
    
    employees_list = db.execute('SELECT * FROM employees WHERE is_active=TRUE ORDER BY name').fetchall()
    return jsonify([dict(e) for e in employees_list])

@app.route('/api/employees/<int:id>', methods=['PUT', 'DELETE'])
@role_required('manager')
def api_employee(id):
    db = get_db()
    
    if request.method == 'PUT':
        data = request.get_json()
        db.execute('''
            UPDATE employees SET name=%s, job_title=%s, phone=%s, address=%s, latitude=%s, longitude=%s, department=%s, notes=%s
            WHERE id=%s
        ''', (data['name'], data.get('job_title'), data.get('phone'), data.get('address'),
              data.get('latitude'), data.get('longitude'), data.get('department'), data.get('notes'), id))
        db.commit()
        return jsonify({'success': True, 'message': 'تم تحديث الموظف'})
    
    elif request.method == 'DELETE':
        db.execute('UPDATE employees SET is_active=FALSE WHERE id=%s', (id,))
        db.commit()
        return jsonify({'success': True, 'message': 'تم حذف الموظف'})

# ═══════════════════════════════════════════════════════════════
# نظام الإشعارات
# ═══════════════════════════════════════════════════════════════

def create_notification(user_id, title, message, notif_type='info', link=None):
    """إنشاء إشعار جديد"""
    db = get_db()
    db.execute('''
        INSERT INTO notifications (user_id, title, message, type, link)
        VALUES (%s, %s, %s, %s, %s)
    ''', (user_id, title, message, notif_type, link))
    db.commit()

def notify_all_managers(title, message, notif_type='info', link=None):
    """إرسال إشعار لجميع المديرين"""
    db = get_db()
    managers = db.execute("SELECT id FROM users WHERE role='manager' AND is_active=TRUE").fetchall()
    for m in managers:
        create_notification(m['id'], title, message, notif_type, link)

@app.route('/api/notifications')
@login_required
def api_notifications():
    db = get_db()
    notifications = db.execute('''
        SELECT * FROM notifications 
        WHERE user_id = %s 
        ORDER BY created_at DESC 
        LIMIT 20
    ''', (session['user_id'],)).fetchall()
    return jsonify([dict(n) for n in notifications])

@app.route('/api/notifications/unread-count')
@login_required
def api_notifications_count():
    db = get_db()
    count = db.execute('''
        SELECT COUNT(*) as count FROM notifications 
        WHERE user_id = %s AND is_read = 0
    ''', (session['user_id'],)).fetchone()['count']
    return jsonify({'count': count})

@app.route('/api/notifications/<int:id>/read', methods=['POST'])
@login_required
def api_notification_read(id):
    db = get_db()
    db.execute('UPDATE notifications SET is_read = 1 WHERE id = %s AND user_id = %s', (id, session['user_id']))
    db.commit()
    return jsonify({'success': True})

@app.route('/api/notifications/read-all', methods=['POST'])
@login_required
def api_notifications_read_all():
    db = get_db()
    db.execute('UPDATE notifications SET is_read = 1 WHERE user_id = %s', (session['user_id'],))
    db.commit()
    return jsonify({'success': True})

# ═══════════════════════════════════════════════════════════════
# سجل التدقيق
# ═══════════════════════════════════════════════════════════════

def _insert_audit_row_direct(user_id, action, table_name, record_id, old_data_json, new_data_json, ip):
    conn = None
    cur = None
    try:
        conn = psycopg2.connect(
            host=app.config['PG_HOST'],
            port=app.config['PG_PORT'],
            database=app.config['PG_DATABASE'],
            user=app.config['PG_USER'],
            password=app.config['PG_PASSWORD']
        )
        cur = conn.cursor()
        cur.execute('''
            CREATE TABLE IF NOT EXISTS audit_log (
                id SERIAL PRIMARY KEY,
                user_id INTEGER,
                action TEXT NOT NULL,
                table_name TEXT,
                record_id INTEGER,
                old_data TEXT,
                new_data TEXT,
                ip_address TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        cur.execute('''
            INSERT INTO audit_log (user_id, action, table_name, record_id, old_data, new_data, ip_address)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        ''', (user_id, action, table_name, record_id, old_data_json, new_data_json, ip))
        conn.commit()
    except Exception:
        try:
            if conn:
                conn.rollback()
        except Exception:
            pass
    finally:
        try:
            if cur:
                cur.close()
        except Exception:
            pass
        try:
            if conn:
                conn.close()
        except Exception:
            pass

def log_action(action, table_name=None, record_id=None, old_data=None, new_data=None, user_id_override=None, use_direct=True):
    """تسجيل إجراء في سجل التدقيق"""
    try:
        user_id = user_id_override if user_id_override is not None else session.get('user_id')
    except Exception:
        user_id = user_id_override

    ip = request.remote_addr if request else None
    old_data_json = json.dumps(_sanitize_for_audit(old_data), ensure_ascii=False) if old_data is not None else None
    new_data_json = json.dumps(_sanitize_for_audit(new_data), ensure_ascii=False) if new_data is not None else None

    if use_direct:
        _insert_audit_row_direct(user_id, action, table_name, record_id, old_data_json, new_data_json, ip)
        return

    db = get_db()
    db.execute('''
        CREATE TABLE IF NOT EXISTS audit_log (
            id SERIAL PRIMARY KEY,
            user_id INTEGER,
            action TEXT NOT NULL,
            table_name TEXT,
            record_id INTEGER,
            old_data TEXT,
            new_data TEXT,
            ip_address TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    db.execute('''
        INSERT INTO audit_log (user_id, action, table_name, record_id, old_data, new_data, ip_address)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
    ''', (user_id, action, table_name, record_id, old_data_json, new_data_json, ip))
    db.commit()

@app.route('/api/audit-log')
@role_required('manager')
def api_audit_log():
    db = get_db()
    db.execute('''
        CREATE TABLE IF NOT EXISTS audit_log (
            id SERIAL PRIMARY KEY,
            user_id INTEGER,
            action TEXT NOT NULL,
            table_name TEXT,
            record_id INTEGER,
            old_data TEXT,
            new_data TEXT,
            ip_address TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    logs = db.execute('''
        SELECT a.*, u.display_name as user_name
        FROM audit_log a
        LEFT JOIN users u ON u.id = a.user_id
        ORDER BY a.created_at DESC
        LIMIT 100
    ''').fetchall()
    return jsonify([dict(l) for l in logs])

# ═══════════════════════════════════════════════════════════════
# إدارة الدفعات وتواريخ الصلاحية
# ═══════════════════════════════════════════════════════════════

@app.route('/batches')
@login_required
def batches():
    db = get_db()
    
    # الدفعات مع معلومات المنتج
    batches_list = db.execute('''
        SELECT b.*, p.name as product_name, p.unit,
               s.name as supplier_name,
               CASE 
                   WHEN b.expiry_date < CURRENT_DATE THEN 'expired'
                   WHEN b.expiry_date <= (CURRENT_DATE + INTERVAL '7 days') THEN 'critical'
                   WHEN b.expiry_date <= (CURRENT_DATE + INTERVAL '30 days') THEN 'warning'
                   ELSE 'ok'
               END as expiry_status,
               (b.expiry_date - CURRENT_DATE) as days_until_expiry
        FROM product_batches b
        JOIN products p ON p.id = b.product_id
        LEFT JOIN supplier_invoices si ON si.id = b.supplier_invoice_id
        LEFT JOIN suppliers s ON s.id = si.supplier_id
        WHERE b.status = 'active' AND b.quantity_remaining > 0
        ORDER BY b.expiry_date ASC
    ''').fetchall()
    
    # إحصائيات
    stats = {
        'total': len(batches_list),
        'expired': len([b for b in batches_list if b['expiry_status'] == 'expired']),
        'critical': len([b for b in batches_list if b['expiry_status'] == 'critical']),
        'warning': len([b for b in batches_list if b['expiry_status'] == 'warning']),
    }
    
    products = db.execute('SELECT id, name FROM products WHERE is_active = TRUE ORDER BY name').fetchall()
    
    return render_template('batches.html', batches=batches_list, stats=stats, products=products)

@app.route('/api/batches', methods=['GET', 'POST'])
@login_required
def api_batches():
    db = get_db()
    
    if request.method == 'POST':
        data = request.get_json()
        
        # إنشاء رقم دفعة تلقائي إذا لم يُحدد
        batch_number = data.get('batch_number')
        if not batch_number:
            batch_number = f"LOT-{datetime.now().strftime('%Y%m%d%H%M%S')}"
        
        cursor = db.execute('''
            INSERT INTO product_batches 
            (product_id, batch_number, production_date, expiry_date, 
             quantity_received, quantity_remaining, purchase_price, location, notes)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        ''', (
            data['product_id'],
            batch_number,
            data.get('production_date'),
            data.get('expiry_date'),
            data.get('quantity', 0),
            data.get('quantity', 0),
            data.get('purchase_price'),
            data.get('location'),
            data.get('notes')
        ))
        db.commit()
        
        # تحديث مخزون المنتج
        db.execute('''
            UPDATE products SET current_stock = current_stock + %s WHERE id = %s
        ''', (data.get('quantity', 0), data['product_id']))
        db.commit()
        
        return jsonify({'success': True, 'id': cursor.lastrowid, 'batch_number': batch_number})
    
    # GET - جلب الدفعات
    batches_list = db.execute('''
        SELECT b.*, p.name as product_name
        FROM product_batches b
        JOIN products p ON p.id = b.product_id
        WHERE b.status = 'active'
        ORDER BY b.expiry_date ASC
    ''').fetchall()
    
    return jsonify([dict(b) for b in batches_list])

@app.route('/api/batches/<int:id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def api_batch(id):
    db = get_db()
    
    if request.method == 'GET':
        batch = db.execute('''
            SELECT b.*, p.name as product_name
            FROM product_batches b
            JOIN products p ON p.id = b.product_id
            WHERE b.id = %s
        ''', (id,)).fetchone()
        return jsonify(dict(batch) if batch else {})
    
    elif request.method == 'PUT':
        data = request.get_json()
        db.execute('''
            UPDATE product_batches 
            SET batch_number=%s, production_date=%s, expiry_date=%s, 
                quantity_remaining=%s, location=%s, notes=%s, status=%s
            WHERE id=%s
        ''', (
            data.get('batch_number'),
            data.get('production_date'),
            data.get('expiry_date'),
            data.get('quantity_remaining'),
            data.get('location'),
            data.get('notes'),
            data.get('status', 'active'),
            id
        ))
        db.commit()
        return jsonify({'success': True})
    
    elif request.method == 'DELETE':
        db.execute('UPDATE product_batches SET status = "consumed" WHERE id = %s', (id,))
        db.commit()
        return jsonify({'success': True})

@app.route('/api/batches/expiring')
@login_required
def api_expiring_batches():
    """الدفعات التي ستنتهي صلاحيتها قريباً"""
    db = get_db()
    days = request.args.get('days', 30, type=int)
    
    batches = db.execute('''
        SELECT b.*, p.name as product_name, p.unit,
               (b.expiry_date - CURRENT_DATE) as days_remaining
        FROM product_batches b
        JOIN products p ON p.id = b.product_id
        WHERE b.status = 'active' 
          AND b.quantity_remaining > 0
          AND b.expiry_date <= (CURRENT_DATE + (%s || ' days')::INTERVAL)
          AND b.expiry_date >= CURRENT_DATE
        ORDER BY b.expiry_date ASC
    ''', (days,)).fetchall()
    
    return jsonify([dict(b) for b in batches])

@app.route('/api/batches/expired')
@login_required
def api_expired_batches():
    """الدفعات المنتهية الصلاحية"""
    db = get_db()
    
    batches = db.execute('''
        SELECT b.*, p.name as product_name, p.unit,
               (CURRENT_DATE - b.expiry_date) as days_expired
        FROM product_batches b
        JOIN products p ON p.id = b.product_id
        WHERE b.status = 'active' 
          AND b.quantity_remaining > 0
          AND b.expiry_date < CURRENT_DATE
        ORDER BY b.expiry_date ASC
    ''').fetchall()
    
    return jsonify([dict(b) for b in batches])

@app.route('/api/products/<int:product_id>/batches')
@login_required
def api_product_batches(product_id):
    """دفعات منتج معين (FEFO - الأقرب انتهاءً أولاً)"""
    db = get_db()
    
    batches = db.execute('''
        SELECT * FROM product_batches
        WHERE product_id = %s AND status = 'active' AND quantity_remaining > 0
        ORDER BY expiry_date ASC
    ''', (product_id,)).fetchall()
    
    return jsonify([dict(b) for b in batches])

def check_expiry_alerts():
    """فحص الدفعات وإنشاء تنبيهات انتهاء الصلاحية"""
    db = get_db()
    
    # الدفعات التي ستنتهي خلال 7 أيام
    approaching = db.execute('''
        SELECT b.id, b.batch_number, p.name as product_name, b.expiry_date, b.quantity_remaining
        FROM product_batches b
        JOIN products p ON p.id = b.product_id
        WHERE b.status = 'active' 
          AND b.quantity_remaining > 0
          AND b.expiry_date <= (CURRENT_DATE + INTERVAL '7 days')
          AND b.expiry_date > CURRENT_DATE
          AND b.id NOT IN (
              SELECT batch_id FROM expiry_alerts 
              WHERE alert_type = 'approaching' AND alert_date = CURRENT_DATE
          )
    ''').fetchall()
    
    for batch in approaching:
        # إنشاء تنبيه
        db.execute('''
            INSERT INTO expiry_alerts (batch_id, alert_type, alert_date)
            VALUES (%s, 'approaching', CURRENT_DATE)
        ''', (batch['id'],))
        
        # إرسال إشعار للمديرين
        notify_all_managers(
            f"⚠️ صلاحية قريبة: {batch['product_name']}",
            f"الدفعة {batch['batch_number']} ستنتهي في {batch['expiry_date']} - الكمية: {batch['quantity_remaining']}",
            'warning',
            '/batches'
        )
    
    db.commit()

# ═══════════════════════════════════════════════════════════════
# نقاط البيع (POS)
# ═══════════════════════════════════════════════════════════════

@app.route('/pos')
@login_required
def pos():
    db = get_db()
    
    # جلب المنتجات مع الأسعار
    products = db.execute('''
        SELECT p.id, p.name, p.barcode, p.category_id, p.unit, p.current_stock,
               COALESCE(
                   (SELECT pp.price FROM product_prices pp 
                    JOIN price_lists pl ON pl.id = pp.price_list_id 
                    WHERE pp.product_id = p.id AND pl.is_default = TRUE),
                   0
               ) as price
        FROM products p
        WHERE p.is_active = TRUE
        ORDER BY p.name
    ''').fetchall()
    
    categories = db.execute('SELECT id, name FROM categories ORDER BY name').fetchall()
    
    return render_template('pos.html', 
                          products=[dict(p) for p in products],
                          categories=categories)

@app.route('/api/pos/invoice', methods=['POST'])
@login_required
def api_pos_invoice():
    """حفظ فاتورة POS"""
    db = get_db()
    data = request.get_json()
    
    # إنشاء رقم فاتورة
    invoice_number = f"POS-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    
    # إنشاء الفاتورة
    cursor = db.execute('''
        INSERT INTO pos_invoices 
        (invoice_number, total_amount, discount, payment_method, paid_amount, 
         change_amount, customer_name, customer_phone, created_by, created_at)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    ''', (
        invoice_number,
        data.get('total', 0),
        data.get('discount', 0),
        data.get('paymentMethod', 'cash'),
        data.get('paidAmount', 0),
        data.get('change', 0),
        data.get('customer', {}).get('name') if data.get('customer') else None,
        data.get('customer', {}).get('phone') if data.get('customer') else None,
        session.get('user_id'),
        datetime.now().isoformat()
    ))
    invoice_id = cursor.lastrowid
    
    # إضافة بنود الفاتورة وتحديث المخزون
    for item in data.get('items', []):
        db.execute('''
            INSERT INTO pos_invoice_items (invoice_id, product_id, product_name, quantity, price, total)
            VALUES (%s, %s, %s, %s, %s, %s)
        ''', (invoice_id, item['id'], item['name'], item['qty'], item['price'], item['price'] * item['qty']))
        
        # تحديث المخزون
        db.execute('UPDATE products SET current_stock = current_stock - %s WHERE id = %s', (item['qty'], item['id']))
    
    db.commit()
    
    return jsonify({'success': True, 'invoice_number': invoice_number, 'invoice_id': invoice_id})

@app.route('/api/pos/sync', methods=['POST'])
@login_required  
def api_pos_sync():
    """مزامنة فواتير POS من الأوفلاين"""
    db = get_db()
    data = request.get_json()
    invoices = data.get('invoices', [])
    
    synced = 0
    for invoice_data in invoices:
        if invoice_data.get('synced'):
            continue
            
        # نفس منطق حفظ الفاتورة
        cursor = db.execute('''
            INSERT INTO pos_invoices 
            (invoice_number, total_amount, discount, payment_method, paid_amount, 
             change_amount, customer_name, customer_phone, created_by, created_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ''', (
            f"POS-{invoice_data.get('number', 0)}",
            invoice_data.get('total', 0),
            invoice_data.get('discount', 0),
            invoice_data.get('paymentMethod', 'cash'),
            invoice_data.get('paidAmount', 0),
            invoice_data.get('change', 0),
            invoice_data.get('customer', {}).get('name') if invoice_data.get('customer') else None,
            invoice_data.get('customer', {}).get('phone') if invoice_data.get('customer') else None,
            session.get('user_id'),
            invoice_data.get('timestamp', datetime.now().isoformat())
        ))
        
        invoice_id = cursor.lastrowid
        for item in invoice_data.get('items', []):
            db.execute('''
                INSERT INTO pos_invoice_items (invoice_id, product_id, product_name, quantity, price, total)
                VALUES (%s, %s, %s, %s, %s, %s)
            ''', (invoice_id, item['id'], item['name'], item['qty'], item['price'], item['price'] * item['qty']))
        
        synced += 1
    
    db.commit()
    return jsonify({'success': True, 'synced': synced})

@app.route('/api/pos/reports/daily')
@login_required
def api_pos_daily_report():
    """تقرير المبيعات اليومية"""
    db = get_db()
    
    today = datetime.now().strftime('%Y-%m-%d')
    
    data = db.execute('''
        SELECT 
            COUNT(*) as total_invoices,
            SUM(total_amount) as total_sales,
            SUM(discount) as total_discount,
            SUM(CASE WHEN payment_method = 'cash' THEN total_amount ELSE 0 END) as cash_sales,
            SUM(CASE WHEN payment_method = 'card' THEN total_amount ELSE 0 END) as card_sales,
            SUM(CASE WHEN payment_method = 'credit' THEN total_amount ELSE 0 END) as credit_sales
        FROM pos_invoices
        WHERE created_at::DATE = %s
    ''', (today,)).fetchone()
    
    return jsonify(dict(data) if data else {})

@app.route('/pos/login')
def pos_login():
    """صفحة تسجيل دخول الكاشير"""
    db = get_db()
    cashiers = db.execute("SELECT id, display_name FROM users WHERE role IN ('cashier', 'manager')").fetchall()
    return render_template('pos_login.html', cashiers=cashiers)

@app.route('/api/pos/verify-pin', methods=['POST'])
def api_pos_verify_pin():
    """التحقق من PIN الكاشير"""
    data = request.get_json()
    user_id = data.get('user_id')
    pin = str(data.get('pin', ''))
    
    if not user_id:
        return jsonify({'success': False, 'message': 'يرجى اختيار المستخدم'})
    
    db = get_db()
    user = db.execute('SELECT * FROM users WHERE id = %s', (int(user_id),)).fetchone()
    
    if user:
        # للتبسيط: PIN = 1234 افتراضياً لجميع المستخدمين
        # في الإنتاج يجب إضافة عمود pin_code في جدول users
        if pin == '1234':
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['display_name'] = user['display_name']
            session['role'] = user['role']
            return jsonify({'success': True})
    
    return jsonify({'success': False, 'message': 'رمز PIN غير صحيح'})

@app.route('/pos/returns')
@login_required
def pos_returns():
    """صفحة المرتجعات"""
    return render_template('pos_returns.html')

@app.route('/api/pos/invoice/<invoice_number>')
@login_required
def api_pos_get_invoice(invoice_number):
    """جلب تفاصيل فاتورة"""
    db = get_db()
    
    invoice = db.execute('''
        SELECT * FROM pos_invoices WHERE invoice_number = %s OR id = %s
    ''', (invoice_number, invoice_number)).fetchone()
    
    if not invoice:
        return jsonify({'found': False})
    
    items = db.execute('''
        SELECT * FROM pos_invoice_items WHERE invoice_id = %s
    ''', (invoice['id'],)).fetchall()
    
    return jsonify({
        'found': True,
        'invoice': dict(invoice),
        'items': [dict(i) for i in items]
    })

@app.route('/api/pos/return', methods=['POST'])
@login_required
def api_pos_return():
    """معالجة المرتجعات"""
    db = get_db()
    data = request.get_json()
    
    return_number = f"RET-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    
    # إنشاء سجل المرتجع
    cursor = db.execute('''
        INSERT INTO pos_returns (return_number, total_amount, refund_method, created_by, created_at)
        VALUES (%s, %s, %s, %s, %s)
    ''', (return_number, data.get('total', 0), data.get('refund_method', 'cash'), 
          session.get('user_id'), datetime.now().isoformat()))
    
    return_id = cursor.lastrowid
    
    # إضافة بنود المرتجع وتحديث المخزون
    for item in data.get('items', []):
        db.execute('''
            INSERT INTO pos_return_items 
            (return_id, invoice_id, product_id, product_name, quantity, price, reason)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        ''', (return_id, item['invoice_id'], item['product_id'], item['product_name'],
              item['return_qty'], item['price'], item.get('reason', '')))
        
        # إرجاع الكمية للمخزون
        db.execute('UPDATE products SET current_stock = current_stock + %s WHERE id = %s', (item['return_qty'], item['product_id']))
    
    db.commit()
    return jsonify({'success': True, 'return_number': return_number})

@app.route('/pos/shift')
@login_required
def pos_shift():
    """صفحة إدارة الورديات"""
    return render_template('pos_shift.html')

@app.route('/api/pos/shift/current')
@login_required
def api_pos_current_shift():
    """جلب الوردية الحالية"""
    db = get_db()
    
    shift = db.execute('''
        SELECT * FROM pos_shifts WHERE user_id = %s AND status = 'open'
        ORDER BY id DESC LIMIT 1
    ''', (session.get('user_id'),)).fetchone()
    
    if not shift:
        return jsonify({'shift': None, 'stats': {}, 'recent_invoices': []})
    
    # إحصائيات الوردية
    stats = db.execute('''
        SELECT 
            COUNT(*) as total_invoices,
            COALESCE(SUM(total_amount), 0) as total_sales,
            COALESCE(SUM(CASE WHEN payment_method = 'cash' THEN total_amount ELSE 0 END), 0) as cash_sales,
            COALESCE(SUM(CASE WHEN payment_method = 'card' THEN total_amount ELSE 0 END), 0) as card_sales
        FROM pos_invoices
        WHERE shift_id = %s
    ''', (shift['id'],)).fetchone()
    
    # آخر الفواتير
    invoices = db.execute('''
        SELECT id, invoice_number, payment_method, total_amount,
               strftime('%H:%M', created_at) as time
        FROM pos_invoices
        WHERE shift_id = %s
        ORDER BY id DESC LIMIT 10
    ''', (shift['id'],)).fetchall()
    
    return jsonify({
        'shift': dict(shift),
        'stats': dict(stats) if stats else {},
        'recent_invoices': [dict(i) for i in invoices]
    })

@app.route('/api/pos/shift/open', methods=['POST'])
@login_required
def api_pos_open_shift():
    """فتح وردية جديدة"""
    db = get_db()
    data = request.get_json()
    
    # التحقق من عدم وجود وردية مفتوحة
    existing = db.execute('''
        SELECT id FROM pos_shifts WHERE user_id = %s AND status = 'open'
    ''', (session.get('user_id'),)).fetchone()
    
    if existing:
        return jsonify({'success': False, 'message': 'يوجد وردية مفتوحة بالفعل'})
    
    db.execute('''
        INSERT INTO pos_shifts (user_id, start_time, opening_balance, status)
        VALUES (%s, %s, %s, 'open')
    ''', (session.get('user_id'), datetime.now().isoformat(), data.get('opening_balance', 0)))
    
    db.commit()
    return jsonify({'success': True})

@app.route('/api/pos/shift/close', methods=['POST'])
@login_required
def api_pos_close_shift():
    """إغلاق الوردية"""
    db = get_db()
    data = request.get_json()
    
    shift = db.execute('''
        SELECT * FROM pos_shifts WHERE user_id = %s AND status = 'open'
    ''', (session.get('user_id'),)).fetchone()
    
    if not shift:
        return jsonify({'success': False, 'message': 'لا توجد وردية مفتوحة'})
    
    # حساب الإجماليات
    stats = db.execute('''
        SELECT 
            COALESCE(SUM(total_amount), 0) as total_sales,
            COALESCE(SUM(CASE WHEN payment_method = 'cash' THEN total_amount ELSE 0 END), 0) as cash_sales,
            COALESCE(SUM(CASE WHEN payment_method = 'card' THEN total_amount ELSE 0 END), 0) as card_sales
        FROM pos_invoices WHERE shift_id = %s
    ''', (shift['id'],)).fetchone()
    
    db.execute('''
        UPDATE pos_shifts SET 
            end_time = %s, closing_balance = %s, total_cash = %s, total_card = %s,
            total_sales = %s, status = 'closed', notes = %s
        WHERE id = %s
    ''', (datetime.now().isoformat(), data.get('closing_balance', 0),
          stats['cash_sales'], stats['card_sales'], stats['total_sales'],
          data.get('notes', ''), shift['id']))
    
    db.commit()
    return jsonify({'success': True})

@app.route('/pos/reports')
@login_required
def pos_reports():
    """صفحة تقارير POS"""
    return render_template('pos_reports.html')

@app.route('/api/pos/reports')
@login_required
def api_pos_reports():
    """API تقارير POS"""
    db = get_db()
    date_from = request.args.get('from', datetime.now().strftime('%Y-%m-%d'))
    date_to = request.args.get('to', datetime.now().strftime('%Y-%m-%d'))
    
    # الملخص العام
    summary = db.execute('''
        SELECT 
            COUNT(*) as total_invoices,
            COALESCE(SUM(total_amount), 0) as total_sales,
            COALESCE(AVG(total_amount), 0) as avg_invoice,
            COALESCE(SUM(CASE WHEN payment_method = 'cash' THEN total_amount ELSE 0 END), 0) as cash_sales,
            COALESCE(SUM(CASE WHEN payment_method = 'card' THEN total_amount ELSE 0 END), 0) as card_sales,
            COALESCE(SUM(CASE WHEN payment_method = 'bank' THEN total_amount ELSE 0 END), 0) as bank_sales,
            COALESCE(SUM(CASE WHEN payment_method = 'credit' THEN total_amount ELSE 0 END), 0) as credit_sales
        FROM pos_invoices
        WHERE created_at::DATE BETWEEN %s AND %s
    ''', (date_from, date_to)).fetchone()
    
    # عدد الأصناف المباعة
    items_count = db.execute('''
        SELECT COALESCE(SUM(pii.quantity), 0) as total_items
        FROM pos_invoice_items pii
        JOIN pos_invoices pi ON pi.id = pii.invoice_id
        WHERE pi.created_at::DATE BETWEEN %s AND %s
    ''', (date_from, date_to)).fetchone()
    
    # أكثر المنتجات مبيعاً
    top_products = db.execute('''
        SELECT pii.product_name as name, 
               SUM(pii.quantity) as quantity,
               SUM(pii.total) as total
        FROM pos_invoice_items pii
        JOIN pos_invoices pi ON pi.id = pii.invoice_id
        WHERE pi.created_at::DATE BETWEEN %s AND %s
        GROUP BY pii.product_id, pii.product_name
        ORDER BY total DESC
        LIMIT 10
    ''', (date_from, date_to)).fetchall()
    
    # أداء الكاشير
    cashier_stats = db.execute('''
        SELECT u.display_name as name,
               COUNT(*) as invoices,
               COALESCE(SUM(pi.total_amount), 0) as total
        FROM pos_invoices pi
        JOIN users u ON u.id = pi.created_by
        WHERE pi.created_at::DATE BETWEEN %s AND %s
        GROUP BY pi.created_by
        ORDER BY total DESC
    ''', (date_from, date_to)).fetchall()
    
    # المبيعات حسب الساعة
    hourly_sales = db.execute('''
        SELECT strftime('%H', created_at) as hour,
               COALESCE(SUM(total_amount), 0) as total
        FROM pos_invoices
        WHERE created_at::DATE BETWEEN %s AND %s
        GROUP BY hour
        ORDER BY hour
    ''', (date_from, date_to)).fetchall()
    
    # آخر الفواتير
    recent_invoices = db.execute('''
        SELECT pi.*, u.display_name as cashier,
               strftime('%Y-%m-%d %H:%M', pi.created_at) as date,
               (SELECT COUNT(*) FROM pos_invoice_items WHERE invoice_id = pi.id) as items_count
        FROM pos_invoices pi
        LEFT JOIN users u ON u.id = pi.created_by
        WHERE pi.created_at::DATE BETWEEN %s AND %s
        ORDER BY pi.id DESC
        LIMIT 50
    ''', (date_from, date_to)).fetchall()
    
    return jsonify({
        'summary': {
            **dict(summary),
            'total_items': items_count['total_items'] if items_count else 0
        },
        'payment_breakdown': {
            'cash': summary['cash_sales'],
            'card': summary['card_sales'],
            'bank': summary['bank_sales'],
            'credit': summary['credit_sales']
        },
        'top_products': [dict(p) for p in top_products],
        'cashier_stats': [dict(c) for c in cashier_stats],
        'hourly_sales': [dict(h) for h in hourly_sales],
        'recent_invoices': [dict(i) for i in recent_invoices]
    })

# ═══════════════════════════════════════════════════════════════
# إدارة الأقسام
# ═══════════════════════════════════════════════════════════════

@app.route('/categories')
@role_required('manager')
def categories():
    db = get_db()
    categories_list = db.execute('''
        SELECT c.*, 
               (SELECT COUNT(*) FROM products WHERE category_id = c.id) as products_count,
               pc.name as parent_name
        FROM categories c
        LEFT JOIN categories pc ON pc.id = c.parent_id
        ORDER BY c.parent_id NULLS FIRST, c.sort_order, c.name
    ''').fetchall()
    return render_template('categories.html', categories=categories_list)

@app.route('/api/categories', methods=['GET', 'POST'])
@role_required('manager')
def api_categories():
    db = get_db()
    
    if request.method == 'POST':
        data = request.json
        parent_id = data.get('parent_id')
        if parent_id == '' or parent_id == 0:
            parent_id = None
        db.execute('''
            INSERT INTO categories (name, parent_id, sort_order, description, is_active)
            VALUES (%s, %s, %s, %s, %s)
        ''', (data['name'], parent_id, data.get('sort_order', 0), data.get('description', ''), data.get('is_active', True)))
        db.commit()
        return jsonify({'success': True, 'message': 'تم إضافة القسم'})
    
    categories_list = db.execute('SELECT * FROM categories ORDER BY parent_id NULLS FIRST, sort_order, name').fetchall()
    return jsonify([dict(c) for c in categories_list])

@app.route('/api/categories/<int:id>', methods=['PUT', 'DELETE'])
@role_required('manager')
def api_category(id):
    db = get_db()
    
    if request.method == 'PUT':
        data = request.json
        parent_id = data.get('parent_id')
        if parent_id == '' or parent_id == 0:
            parent_id = None
        db.execute('''
            UPDATE categories SET name=%s, parent_id=%s, sort_order=%s, description=%s, is_active=%s
            WHERE id=%s
        ''', (data['name'], parent_id, data.get('sort_order', 0), data.get('description', ''), data.get('is_active', True), id))
        db.commit()
        return jsonify({'success': True, 'message': 'تم تحديث القسم'})
    
    elif request.method == 'DELETE':
        # التحقق من وجود أصناف أو أقسام فرعية
        has_products = db.execute('SELECT COUNT(*) as count FROM products WHERE category_id = %s', (id,)).fetchone()['count']
        has_children = db.execute('SELECT COUNT(*) as count FROM categories WHERE parent_id = %s', (id,)).fetchone()['count']
        
        if has_products > 0:
            return jsonify({'success': False, 'message': f'لا يمكن حذف القسم - يحتوي على {has_products} صنف'})
        if has_children > 0:
            return jsonify({'success': False, 'message': f'لا يمكن حذف القسم - يحتوي على {has_children} قسم فرعي'})
        
        db.execute('DELETE FROM categories WHERE id=%s', (id,))
        db.commit()
        return jsonify({'success': True, 'message': 'تم حذف القسم'})

# ═══════════════════════════════════════════════════════════════
# إدارة الوحدات
# ═══════════════════════════════════════════════════════════════

@app.route('/units')
@role_required('manager')
def units():
    db = get_db()
    try:
        _sync_units_catalog_from_products(db)
    except Exception:
        db.rollback()
    units_list = db.execute('SELECT * FROM units WHERE is_active = TRUE ORDER BY name').fetchall()
    return render_template('units.html', units=units_list)

@app.route('/api/units', methods=['GET', 'POST'])
@role_required('manager')
def api_units():
    db = get_db()
    
    if request.method == 'POST':
        data = request.json or {}
        name = (data.get('name') or '').strip()
        symbol = (data.get('symbol') or '').strip()
        if not name:
            return jsonify({'success': False, 'message': 'اسم الوحدة مطلوب'}), 400
        # التحقق من عدم التكرار
        existing = db.execute('SELECT id FROM units WHERE LOWER(name) = LOWER(%s)', (name,)).fetchone()
        if existing:
            return jsonify({'success': False, 'message': 'الوحدة موجودة مسبقاً'}), 400
        
        db.execute('INSERT INTO units (name, symbol) VALUES (%s, %s)',
                   (name, symbol))
        db.commit()
        return jsonify({'success': True, 'message': 'تم إضافة الوحدة'})

    try:
        _sync_units_catalog_from_products(db)
    except Exception:
        db.rollback()
    units_list = db.execute('SELECT * FROM units WHERE is_active = TRUE ORDER BY name').fetchall()
    return jsonify([dict(u) for u in units_list])

@app.route('/api/units/<int:id>', methods=['PUT', 'DELETE'])
@role_required('manager')
def api_unit(id):
    db = get_db()
    
    if request.method == 'PUT':
        data = request.json or {}
        name = (data.get('name') or '').strip()
        symbol = (data.get('symbol') or '').strip()
        if not name:
            return jsonify({'success': False, 'message': 'اسم الوحدة مطلوب'}), 400

        duplicate = db.execute(
            'SELECT id FROM units WHERE LOWER(name) = LOWER(%s) AND id <> %s',
            (name, id)
        ).fetchone()
        if duplicate:
            return jsonify({'success': False, 'message': 'اسم الوحدة مستخدم مسبقاً'}), 400

        db.execute('UPDATE units SET name=%s, symbol=%s WHERE id=%s',
                   (name, symbol, id))
        db.commit()
        return jsonify({'success': True, 'message': 'تم تحديث الوحدة'})
    
    elif request.method == 'DELETE':
        # التحقق من عدم استخدام الوحدة
        used = db.execute('SELECT COUNT(*) as count FROM product_units WHERE unit_id = %s', (id,)).fetchone()['count']
        if used > 0:
            return jsonify({'success': False, 'message': f'لا يمكن حذف الوحدة - مستخدمة في {used} صنف'})
        
        db.execute('UPDATE units SET is_active = FALSE WHERE id = %s', (id,))
        db.commit()
        return jsonify({'success': True, 'message': 'تم حذف الوحدة'})

# ═══════════════════════════════════════════════════════════════
# ماسح الباركود
# ═══════════════════════════════════════════════════════════════

@app.route('/barcode-scanner')
@login_required
def barcode_scanner():
    db = get_db()
    categories = db.execute('SELECT * FROM categories ORDER BY name').fetchall()
    units = db.execute('SELECT * FROM units WHERE is_active = TRUE ORDER BY name').fetchall()
    return render_template('barcode_scanner.html', categories=categories, units=units)

@app.route('/permissions')
@login_required
@permission_required('users', 'edit')
def permissions_page():
    db = get_db()
    users = db.execute('SELECT id, username, display_name, role FROM users WHERE is_active = TRUE ORDER BY id').fetchall()
    modules = db.execute('SELECT * FROM modules ORDER BY sort_order').fetchall()
    
    # جلب كل الصلاحيات
    all_perms = db.execute('SELECT * FROM user_permissions').fetchall()
    perms_map = {}
    for p in all_perms:
        key = f"{p['user_id']}_{p['module_code']}"
        perms_map[key] = {'view': p['can_view'], 'add': p['can_add'], 'edit': p['can_edit'], 'delete': p['can_delete']}
    
    return render_template('permissions.html', users=users, modules=modules, perms_map=perms_map)

@app.route('/api/permissions/save', methods=['POST'])
@login_required
@permission_required('users', 'edit')
def api_save_permissions():
    db = get_db()
    data = request.get_json() or {}
    user_id = data.get('user_id')
    permissions = data.get('permissions', {})
    
    if not user_id:
        return jsonify({'success': False, 'message': 'المستخدم مطلوب'}), 400
    
    for module_code, perm in permissions.items():
        db.execute('''
            INSERT INTO user_permissions (user_id, module_code, can_view, can_add, can_edit, can_delete)
            VALUES (%s, %s, %s, %s, %s, %s)
            ON CONFLICT (user_id, module_code) DO UPDATE SET
                can_view = EXCLUDED.can_view, can_add = EXCLUDED.can_add,
                can_edit = EXCLUDED.can_edit, can_delete = EXCLUDED.can_delete
        ''', (user_id, module_code, perm.get('view', False), perm.get('add', False), perm.get('edit', False), perm.get('delete', False)))
    
    db.commit()
    return jsonify({'success': True, 'message': 'تم حفظ الصلاحيات'})

@app.route('/requests')
@login_required
def general_requests_page():
    db = get_db()
    status_filter = request.args.get('status', 'all')
    type_filter = request.args.get('type', 'all')
    
    where = []
    params = []
    if status_filter != 'all':
        where.append('r.status = %s')
        params.append(status_filter)
    if type_filter != 'all':
        where.append('r.request_type = %s')
        params.append(type_filter)
    
    where_sql = ' AND '.join(where) if where else '1=1'
    
    requests_list = db.execute(f'''
        SELECT r.*, u.display_name as requested_by_name, u2.display_name as reviewed_by_name,
               p.name as product_name_ref, s.name as supplier_name_ref
        FROM general_requests r
        LEFT JOIN users u ON u.id = r.requested_by
        LEFT JOIN users u2 ON u2.id = r.reviewed_by
        LEFT JOIN products p ON p.id = r.product_id
        LEFT JOIN suppliers s ON s.id = r.supplier_id
        WHERE {where_sql}
        ORDER BY r.id DESC
    ''', params).fetchall()
    
    counts = {}
    for s in ['all', 'pending', 'approved', 'rejected', 'done']:
        if s == 'all':
            counts[s] = db.execute('SELECT COUNT(*) as c FROM general_requests').fetchone()['c']
        else:
            counts[s] = db.execute('SELECT COUNT(*) as c FROM general_requests WHERE status = %s', (s,)).fetchone()['c']
    
    categories = db.execute('SELECT * FROM categories ORDER BY name').fetchall()
    suppliers = db.execute('SELECT id, name FROM suppliers ORDER BY name').fetchall()
    
    import json as json_mod
    parsed_requests = []
    for r in requests_list:
        rd = dict(r)
        rd['proposed'] = None
        details = rd.get('details') or ''
        try:
            rd['proposed'] = json_mod.loads(details)
            rd['details_text'] = ''
        except:
            rd['details_text'] = details
        parsed_requests.append(rd)
    
    return render_template('general_requests.html', requests=parsed_requests, counts=counts, 
                         status_filter=status_filter, type_filter=type_filter,
                         categories=categories, suppliers=suppliers,
                         units=db.execute('SELECT * FROM units WHERE is_active = TRUE ORDER BY name').fetchall())

@app.route('/api/requests', methods=['POST'])
@login_required
def api_create_request():
    db = get_db()
    data = request.get_json() or {}
    
    request_type = data.get('request_type')
    title = (data.get('title') or '').strip()
    details = (data.get('details') or '').strip()
    
    if not request_type or not title:
        return jsonify({'success': False, 'message': 'نوع الطلب والعنوان مطلوبين'}), 400
    
    db.execute('''
        INSERT INTO general_requests (request_type, title, details, product_id, supplier_id, barcode, priority, requested_by)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
    ''', (
        request_type, title, details,
        data.get('product_id') or None, data.get('supplier_id') or None,
        data.get('barcode') or None, data.get('priority', 'normal'),
        session['user_id']
    ))
    db.commit()
    return jsonify({'success': True, 'message': 'تم رفع الطلب بنجاح'})

@app.route('/api/requests/<int:req_id>/<action>', methods=['POST'])
@login_required
def api_request_action(req_id, action):
    db = get_db()
    if action not in ('approve', 'reject', 'done'):
        return jsonify({'success': False, 'message': 'إجراء غير صالح'}), 400
    
    status_map = {'approve': 'approved', 'reject': 'rejected', 'done': 'done'}
    data = request.get_json() or {}
    
    db.execute('UPDATE general_requests SET status = %s, reviewed_by = %s, reviewed_at = NOW(), notes = COALESCE(notes, %s) WHERE id = %s',
               (status_map[action], session['user_id'], data.get('notes', ''), req_id))
    db.commit()
    
    messages = {'approve': 'تم اعتماد الطلب', 'reject': 'تم رفض الطلب', 'done': 'تم إنجاز الطلب'}
    return jsonify({'success': True, 'message': messages[action]})

@app.route('/api/ping')
def api_ping():
    return jsonify({'ok': True, 'time': datetime.now().isoformat()})

@app.route('/api/events')
@login_required
def api_events():
    """SSE - Server-Sent Events للتحديث الفوري"""
    import time
    def generate():
        last_stocktake_count = 0
        last_request_count = 0
        last_chat_count = 0
        
        while True:
            try:
                db = get_db()
                
                # عدد عناصر الجرد
                st = db.execute('SELECT COUNT(*) as c FROM stocktake_items').fetchone()['c']
                # عدد الطلبات
                req = db.execute("SELECT COUNT(*) as c FROM stocktake_product_requests WHERE status='pending'").fetchone()['c']
                # رسائل الدردشة الجديدة
                chat = db.execute('SELECT COUNT(*) as c FROM chat_messages WHERE receiver_id = %s AND is_read = FALSE', (session['user_id'],)).fetchone()['c']
                
                data = {
                    'stocktake_items': st,
                    'pending_requests': req,
                    'unread_chat': chat,
                    'time': datetime.now().strftime('%H:%M:%S')
                }
                
                changed = (st != last_stocktake_count or req != last_request_count or chat != last_chat_count)
                
                if changed or True:  # أرسل دائماً كل 30 ثانية
                    last_stocktake_count = st
                    last_request_count = req
                    last_chat_count = chat
                    yield f"data: {json.dumps(data)}\n\n"
                
            except Exception as e:
                yield f"data: {json.dumps({'error': str(e)})}\n\n"
            
            time.sleep(30)
    
    response = app.response_class(
        generate(),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'X-Accel-Buffering': 'no'
        }
    )
    return response

@app.route('/ca.pem')
def download_ca():
    ca_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'ca.pem')
    if os.path.exists(ca_path):
        return send_file(ca_path, mimetype='application/x-x509-ca-cert', as_attachment=True, download_name='supermarket-ca.pem')
    return 'CA not found', 404

# ==================== الدردشة ====================
@app.route('/chat')
@login_required
def chat_page():
    db = get_db()
    users = db.execute('SELECT id, username, display_name FROM users ORDER BY id').fetchall()
    return render_template('chat.html', users=users)

@app.route('/api/chat/messages')
@login_required
def api_chat_messages():
    db = get_db()
    room = request.args.get('room', 'general')
    after_id = int(request.args.get('after', 0))
    
    if room == 'general':
        messages = db.execute('''
            SELECT m.*, COALESCE(u.display_name, u.username) as sender_name
            FROM chat_messages m
            JOIN users u ON u.id = m.sender_id
            WHERE m.room = %s AND m.id > %s
            ORDER BY m.created_at ASC
            LIMIT 100
        ''', (room, after_id)).fetchall()
    else:
        other_id = int(room.replace('dm_', ''))
        messages = db.execute('''
            SELECT m.*, COALESCE(u.display_name, u.username) as sender_name
            FROM chat_messages m
            JOIN users u ON u.id = m.sender_id
            WHERE m.id > %s AND (
                (m.sender_id = %s AND m.receiver_id = %s)
                OR (m.sender_id = %s AND m.receiver_id = %s)
            )
            ORDER BY m.created_at ASC
            LIMIT 100
        ''', (after_id, session['user_id'], other_id, other_id, session['user_id'])).fetchall()
        
        db.execute('UPDATE chat_messages SET is_read = TRUE WHERE receiver_id = %s AND sender_id = %s AND is_read = FALSE',
                    (session['user_id'], other_id))
        db.commit()
    
    result = []
    for m in messages:
        result.append({
            'id': m['id'],
            'sender_id': m['sender_id'],
            'sender_name': m['sender_name'],
            'message': m['message'],
            'message_type': m['message_type'],
            'file_path': m['file_path'],
            'created_at': m['created_at'].strftime('%H:%M') if m['created_at'] else '',
            'is_mine': m['sender_id'] == session['user_id']
        })
    return jsonify({'success': True, 'messages': result})

@app.route('/api/chat/send', methods=['POST'])
@login_required
def api_chat_send():
    db = get_db()
    room = request.form.get('room', 'general')
    message = (request.form.get('message') or '').strip()
    message_type = 'text'
    file_path = None
    receiver_id = None
    
    # handle file/voice
    file = request.files.get('file')
    if file and file.filename:
        ext = os.path.splitext(secure_filename(file.filename))[1] or '.bin'
        fname = f"chat_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}{ext}"
        rel = os.path.join('static', 'uploads', fname).replace('\\', '/')
        full = os.path.join(app.config['UPLOAD_FOLDER'], fname)
        file.save(full)
        file_path = rel
        if ext.lower() in ['.webm', '.mp3', '.ogg', '.wav', '.m4a']:
            message_type = 'voice'
        elif ext.lower() in ['.jpg', '.jpeg', '.png', '.gif', '.webp']:
            message_type = 'image'
        else:
            message_type = 'file'
        if not message:
            message = file.filename
    
    if not message and not file_path:
        return jsonify({'success': False, 'message': 'الرسالة فارغة'}), 400
    
    if room != 'general' and room.startswith('dm_'):
        receiver_id = int(room.replace('dm_', ''))
    
    db.execute('''
        INSERT INTO chat_messages (sender_id, receiver_id, room, message, message_type, file_path)
        VALUES (%s, %s, %s, %s, %s, %s)
    ''', (session['user_id'], receiver_id, room, message, message_type, file_path))
    db.commit()
    return jsonify({'success': True})

@app.route('/api/chat/unread')
@login_required
def api_chat_unread():
    db = get_db()
    # عدد غير المقروء لكل مرسل (رسائل خاصة)
    per_sender = db.execute('''
        SELECT sender_id, COUNT(*) as cnt 
        FROM chat_messages 
        WHERE receiver_id = %s AND is_read = FALSE
        GROUP BY sender_id
    ''', (session['user_id'],)).fetchall()
    
    # عدد غير المقروء في الغرفة العامة
    general_count = db.execute('''
        SELECT COUNT(*) as cnt FROM chat_messages 
        WHERE room = 'general' AND sender_id != %s AND is_read = FALSE
        AND id NOT IN (SELECT COALESCE(MAX(id),0) FROM chat_messages WHERE room='general')
    ''', (session['user_id'],)).fetchone()['cnt']
    
    total = sum(r['cnt'] for r in per_sender)
    senders = {str(r['sender_id']): r['cnt'] for r in per_sender}
    
    # آخر رسالة لكل محادثة خاصة
    last_msgs = db.execute('''
        SELECT DISTINCT ON (other_id) other_id, message, message_type, created_at, sender_id
        FROM (
            SELECT 
                CASE WHEN sender_id = %s THEN receiver_id ELSE sender_id END as other_id,
                message, message_type, created_at, sender_id
            FROM chat_messages 
            WHERE (sender_id = %s OR receiver_id = %s) AND room != 'general'
        ) sub
        ORDER BY other_id, created_at DESC
    ''', (session['user_id'], session['user_id'], session['user_id'])).fetchall()
    
    last_messages = {}
    for m in last_msgs:
        preview = m['message'] or ''
        if m['message_type'] == 'voice':
            preview = '🎤 رسالة صوتية'
        elif m['message_type'] == 'image':
            preview = '📷 صورة'
        elif m['message_type'] == 'file':
            preview = '📎 ملف'
        if len(preview) > 30:
            preview = preview[:30] + '...'
        is_mine = m['sender_id'] == session['user_id']
        last_messages[str(m['other_id'])] = {
            'preview': ('أنت: ' if is_mine else '') + preview,
            'time': m['created_at'].strftime('%H:%M') if m['created_at'] else ''
        }
    
    return jsonify({'success': True, 'count': total, 'senders': senders, 'general': general_count, 'last_messages': last_messages})

@app.route('/stocktake')
@login_required
def stocktake_page():
    db = get_db()
    categories = db.execute('SELECT * FROM categories ORDER BY name').fetchall()
    units = db.execute('SELECT * FROM units WHERE is_active = TRUE ORDER BY name').fetchall()

    # أي جلسة مفتوحة (الجرد مشترك بين الكل)
    open_session = db.execute('''
        SELECT s.*, u.display_name as creator_name FROM stocktake_sessions s
        LEFT JOIN users u ON u.id = s.created_by
        WHERE s.status = 'open'
        ORDER BY s.id DESC LIMIT 1
    ''').fetchone()

    recent_items = []
    recent_requests = []
    if open_session:
        recent_items = db.execute('''
            SELECT * FROM stocktake_items
            WHERE session_id = %s
            ORDER BY id DESC LIMIT 20
        ''', (open_session['id'],)).fetchall()
        recent_requests = db.execute('''
            SELECT r.*, c.name as category_name
            FROM stocktake_product_requests r
            LEFT JOIN categories c ON c.id = r.category_id
            WHERE r.session_id = %s
            ORDER BY r.id DESC LIMIT 20
        ''', (open_session['id'],)).fetchall()

    return render_template('stocktake.html', categories=categories, units=units, open_session=open_session, recent_items=recent_items, recent_requests=recent_requests)

@app.route('/api/stocktake/recent')
@login_required
def api_stocktake_recent():
    db = get_db()
    session_id = request.args.get('session_id', type=int)
    if not session_id:
        return jsonify({'items': []})
    items = db.execute('''
        SELECT product_name, barcode, selected_unit, counted_stock
        FROM stocktake_items WHERE session_id = %s
        ORDER BY id DESC LIMIT 20
    ''', (session_id,)).fetchall()
    return jsonify({'items': [dict(i) for i in items]})

@app.route('/api/stocktake/recent-requests')
@login_required
def api_stocktake_recent_requests():
    db = get_db()
    session_id = request.args.get('session_id', type=int)
    if not session_id:
        return jsonify({'requests': []})
    requests = db.execute('''
        SELECT product_name, barcode, quantity_counted
        FROM stocktake_product_requests WHERE session_id = %s
        ORDER BY id DESC LIMIT 20
    ''', (session_id,)).fetchall()
    return jsonify({'requests': [dict(r) for r in requests]})

@app.route('/api/stocktake/session', methods=['POST'])
@login_required
def api_create_stocktake_session():
    db = get_db()
    
    # فقط مدير النظام (المستخدم 1) يقدر يفتح جلسة جرد
    if session['user_id'] != 1:
        return jsonify({'success': False, 'message': 'فقط مدير النظام يستطيع فتح جلسة جرد'}), 403
    
    data = request.get_json() or {}
    title = (data.get('title') or '').strip() or f"جرد {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    notes = (data.get('notes') or '').strip()

    existing = db.execute('''
        SELECT * FROM stocktake_sessions
        WHERE status = 'open'
        ORDER BY id DESC LIMIT 1
    ''').fetchone()
    if existing:
        return jsonify({'success': True, 'session_id': existing['id'], 'session': dict(existing), 'message': 'توجد جلسة جرد مفتوحة بالفعل'})

    row = db.execute('''
        INSERT INTO stocktake_sessions (title, notes, status, created_by)
        VALUES (%s, %s, 'open', %s)
        RETURNING *
    ''', (title, notes, session['user_id'])).fetchone()
    db.commit()
    return jsonify({'success': True, 'session_id': row['id'], 'session': dict(row)})

@app.route('/api/stocktake/session/current')
@login_required
def api_current_stocktake_session():
    db = get_db()
    # الجلسة مشتركة — أي جلسة مفتوحة
    open_session = db.execute('''
        SELECT * FROM stocktake_sessions
        WHERE status = 'open'
        ORDER BY id DESC LIMIT 1
    ''').fetchone()
    return jsonify({'success': True, 'session': dict(open_session) if open_session else None})

@app.route('/api/stocktake/session/close', methods=['POST'])
@login_required
def api_close_stocktake_session():
    db = get_db()
    
    # فقط مدير النظام يقدر يقفل الجلسة
    if session['user_id'] != 1:
        return jsonify({'success': False, 'message': 'فقط مدير النظام يستطيع إغلاق جلسة الجرد'}), 403
    
    data = request.get_json() or {}
    session_id = data.get('session_id')
    
    if not session_id:
        return jsonify({'success': False, 'message': 'رقم الجلسة مطلوب'}), 400
    
    stocktake_session = db.execute('SELECT * FROM stocktake_sessions WHERE id = %s', (session_id,)).fetchone()
    if not stocktake_session:
        return jsonify({'success': False, 'message': 'جلسة الجرد غير موجودة'}), 404
    
    if stocktake_session['status'] == 'closed':
        return jsonify({'success': False, 'message': 'الجلسة مغلقة بالفعل'}), 400
    
    db.execute('UPDATE stocktake_sessions SET status = %s, closed_at = NOW() WHERE id = %s', ('closed', session_id))
    db.commit()
    return jsonify({'success': True, 'message': 'تم إغلاق جلسة الجرد'})

@app.route('/stocktake/review')
@login_required
def stocktake_review_page():
    db = get_db()
    
    # جلسات الجرد
    sessions = db.execute('''
        SELECT s.*, u.username as created_by_name,
               (SELECT COUNT(*) FROM stocktake_items WHERE session_id = s.id) as items_count,
               (SELECT COUNT(*) FROM stocktake_product_requests WHERE session_id = s.id) as requests_count
        FROM stocktake_sessions s
        LEFT JOIN users u ON u.id = s.created_by
        ORDER BY s.id DESC
        LIMIT 50
    ''').fetchall()
    
    return render_template('stocktake_review.html', sessions=sessions)

@app.route('/stocktake/requests')
@login_required
def stocktake_requests_page():
    db = get_db()
    status_filter = request.args.get('status', 'all')
    
    if status_filter == 'all':
        requests_list = db.execute('''
            SELECT r.*, c.name as category_name, 
                   u.display_name as requested_by_name,
                   u2.display_name as reviewed_by_name,
                   s.title as session_title
            FROM stocktake_product_requests r
            LEFT JOIN categories c ON c.id = r.category_id
            LEFT JOIN users u ON u.id = r.requested_by
            LEFT JOIN users u2 ON u2.id = r.reviewed_by
            LEFT JOIN stocktake_sessions s ON s.id = r.session_id
            ORDER BY r.id DESC
        ''').fetchall()
    else:
        requests_list = db.execute('''
            SELECT r.*, c.name as category_name,
                   u.display_name as requested_by_name,
                   u2.display_name as reviewed_by_name,
                   s.title as session_title
            FROM stocktake_product_requests r
            LEFT JOIN categories c ON c.id = r.category_id
            LEFT JOIN users u ON u.id = r.requested_by
            LEFT JOIN users u2 ON u2.id = r.reviewed_by
            LEFT JOIN stocktake_sessions s ON s.id = r.session_id
            WHERE r.status = %s
            ORDER BY r.id DESC
        ''', (status_filter,)).fetchall()
    
    categories = db.execute('SELECT * FROM categories ORDER BY name').fetchall()
    counts = {
        'all': db.execute('SELECT COUNT(*) as c FROM stocktake_product_requests').fetchone()['c'],
        'pending': db.execute("SELECT COUNT(*) as c FROM stocktake_product_requests WHERE status='pending'").fetchone()['c'],
        'approved': db.execute("SELECT COUNT(*) as c FROM stocktake_product_requests WHERE status='approved'").fetchone()['c'],
        'rejected': db.execute("SELECT COUNT(*) as c FROM stocktake_product_requests WHERE status='rejected'").fetchone()['c'],
    }
    units = db.execute('SELECT * FROM units WHERE is_active = TRUE ORDER BY name').fetchall()
    
    # تحديد الطلبات المكررة مع أرقامها
    all_pending = db.execute('''
        SELECT id, barcode FROM stocktake_product_requests
        WHERE barcode IS NOT NULL AND barcode != '' AND status = 'pending'
        ORDER BY id
    ''').fetchall()
    
    # بناء خريطة: باركود → قائمة الـ IDs
    from collections import defaultdict
    barcode_ids_map = defaultdict(list)
    for r in all_pending:
        barcode_ids_map[r['barcode']].append(r['id'])
    
    duplicate_barcodes = {bc: ids for bc, ids in barcode_ids_map.items() if len(ids) > 1}
    
    # إضافة معلومات المكرر لكل طلب
    requests_with_dup = []
    for r in requests_list:
        rd = dict(r)
        bc = rd.get('barcode')
        if bc and bc in duplicate_barcodes:
            rd['is_duplicate'] = True
            # الطلبات الأخرى لنفس الباركود
            other_ids = [i for i in duplicate_barcodes[bc] if i != rd['id']]
            rd['duplicate_with'] = other_ids
        else:
            rd['is_duplicate'] = False
            rd['duplicate_with'] = []
        requests_with_dup.append(rd)
    requests_list = requests_with_dup
    
    # عدد المكررات
    counts['duplicates'] = len(duplicate_barcodes)
    
    # طلبات تعديل الأصناف
    edit_requests_raw = db.execute('''
        SELECT r.*, u.display_name as requested_by_name, u2.display_name as reviewed_by_name, p.name as current_name
        FROM stocktake_edit_requests r
        LEFT JOIN users u ON u.id = r.requested_by
        LEFT JOIN users u2 ON u2.id = r.reviewed_by
        LEFT JOIN products p ON p.id = r.product_id
        ORDER BY r.id DESC
    ''').fetchall()
    
    import json as json_mod
    edit_requests = []
    for er in edit_requests_raw:
        er_dict = dict(er)
        details = er_dict.get('details') or ''
        er_dict['note'] = ''
        er_dict['proposed'] = None
        if '=== البيانات المقترحة ===' in details:
            parts = details.split('=== البيانات المقترحة ===')
            er_dict['note'] = parts[0].strip()
            try:
                er_dict['proposed'] = json_mod.loads(parts[1])
            except:
                er_dict['proposed'] = None
        edit_requests.append(er_dict)
    
    return render_template('stocktake_requests.html', requests=requests_list, categories=categories, units=units, status_filter=status_filter, counts=counts, edit_requests=edit_requests)

@app.route('/api/stocktake/request/<int:request_id>/approve', methods=['POST'])
@login_required
def api_approve_request(request_id):
    db = get_db()
    req = db.execute('SELECT * FROM stocktake_product_requests WHERE id = %s', (request_id,)).fetchone()
    if not req:
        return jsonify({'success': False, 'message': 'الطلب غير موجود'}), 404
    
    if req['status'] != 'pending':
        return jsonify({'success': False, 'message': 'الطلب تم معالجته مسبقًا'}), 400
    
    # إنشاء الصنف الجديد
    data = request.get_json() or {}
    extra_units = data.get('extra_units', [])
    _ensure_units_registered(
        db,
        [req.get('unit')] + [eu.get('unit') for eu in extra_units if isinstance(eu, dict)]
    )

    product_code = f"NEW-{request_id}"
    new_product = db.execute('''
        INSERT INTO products (product_code, name, barcode, category_id, unit, cost_price, sell_price, current_stock, is_active)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, TRUE)
        RETURNING id
    ''', (
        product_code,
        req['product_name'] or f'صنف جديد {req["barcode"]}',
        req['barcode'],
        req['category_id'],
        req['unit'] or 'حبه',
        req['cost_price'] or 0,
        req['sell_price'] or 0,
        req['quantity_counted'] or 0
    )).fetchone()
    
    # إضافة الوحدات الإضافية إذا وجدت
    for eu in extra_units:
        if eu.get('unit'):
            db.execute('''
                INSERT INTO product_barcodes (product_id, barcode, unit, pack_size, cost_price, sell_price)
                VALUES (%s, %s, %s, %s, %s, %s)
            ''', (
                new_product['id'], eu.get('barcode') or '', eu['unit'],
                int(eu.get('pack_size') or 1),
                Decimal(str(eu['cost_price'])) if eu.get('cost_price') else None,
                Decimal(str(eu['sell_price'])) if eu.get('sell_price') else None
            ))
    
    db.execute('''
        UPDATE stocktake_product_requests 
        SET status = 'approved', reviewed_by = %s, reviewed_at = NOW()
        WHERE id = %s
    ''', (session['user_id'], request_id))
    db.commit()
    
    return jsonify({'success': True, 'message': f'تم اعتماد الصنف وإضافته برقم {new_product["id"]}', 'product_id': new_product['id']})

@app.route('/api/stocktake/request/<int:request_id>/reject', methods=['POST'])
@login_required
def api_reject_request(request_id):
    db = get_db()
    req = db.execute('SELECT * FROM stocktake_product_requests WHERE id = %s', (request_id,)).fetchone()
    if not req:
        return jsonify({'success': False, 'message': 'الطلب غير موجود'}), 404
    
    db.execute('''
        UPDATE stocktake_product_requests 
        SET status = 'rejected', reviewed_by = %s, reviewed_at = NOW()
        WHERE id = %s
    ''', (session['user_id'], request_id))
    db.commit()
    
    return jsonify({'success': True, 'message': 'تم رفض الطلب'})

@app.route('/api/stocktake/request/<int:request_id>/edit', methods=['POST'])
@login_required
def api_edit_request(request_id):
    db = get_db()
    req = db.execute('SELECT * FROM stocktake_product_requests WHERE id = %s', (request_id,)).fetchone()
    if not req:
        return jsonify({'success': False, 'message': 'الطلب غير موجود'}), 404
    
    if req['status'] != 'pending':
        return jsonify({'success': False, 'message': 'لا يمكن تعديل طلب تم معالجته'}), 400
    
    data = request.get_json() or {}
    
    db.execute('''
        UPDATE stocktake_product_requests SET
            product_name = %s, category_id = %s, unit = %s,
            quantity_counted = %s, pack_size = %s,
            cost_price = %s, sell_price = %s,
            production_date = %s, expiry_date = %s,
            batch_no = %s, notes = %s
        WHERE id = %s
    ''', (
        data.get('product_name') or req['product_name'],
        data.get('category_id') or req['category_id'],
        data.get('unit') or req['unit'],
        Decimal(str(data.get('quantity_counted') or req['quantity_counted'] or 1)),
        int(data.get('pack_size') or req['pack_size'] or 1),
        Decimal(str(data['cost_price'])) if data.get('cost_price') else req['cost_price'],
        Decimal(str(data['sell_price'])) if data.get('sell_price') else req['sell_price'],
        data.get('production_date') or req['production_date'],
        data.get('expiry_date') or req['expiry_date'],
        data.get('batch_no') or req['batch_no'],
        data.get('notes') or req['notes'],
        request_id
    ))
    db.commit()
    return jsonify({'success': True, 'message': 'تم تعديل الطلب'})

@app.route('/api/stocktake/request/<int:request_id>/link', methods=['POST'])
@login_required
def api_link_request(request_id):
    db = get_db()
    data = request.get_json() or {}
    product_id = data.get('product_id')
    
    if not product_id:
        return jsonify({'success': False, 'message': 'اختر الصنف المراد الربط به'}), 400
    
    req = db.execute('SELECT * FROM stocktake_product_requests WHERE id = %s', (request_id,)).fetchone()
    if not req:
        return jsonify({'success': False, 'message': 'الطلب غير موجود'}), 404
    
    product = db.execute('SELECT * FROM products WHERE id = %s', (product_id,)).fetchone()
    if not product:
        return jsonify({'success': False, 'message': 'الصنف غير موجود'}), 404
    
    # إضافة الباركود كوحدة جديدة للصنف
    db.execute('''
        INSERT INTO product_barcodes (product_id, barcode, unit, pack_size, cost_price, sell_price)
        VALUES (%s, %s, %s, %s, %s, %s)
    ''', (product_id, req['barcode'], req['unit'] or 'حبه', req['pack_size'] or 1, req['cost_price'], req['sell_price']))
    
    db.execute('''
        UPDATE stocktake_product_requests 
        SET status = 'approved', reviewed_by = %s, reviewed_at = NOW()
        WHERE id = %s
    ''', (session['user_id'], request_id))
    db.commit()
    
    return jsonify({'success': True, 'message': f'تم ربط الباركود بالصنف: {product["name"]}'})

@app.route('/api/products/search')
@login_required
def api_products_search():
    db = get_db()
    q = (request.args.get('q') or '').strip()
    if len(q) < 2:
        return jsonify({'products': []})
    products = db.execute('''
        SELECT id, name, barcode, unit, sell_price FROM products 
        WHERE is_active = TRUE AND (name ILIKE %s OR barcode ILIKE %s)
        ORDER BY name LIMIT 20
    ''', (f'%{q}%', f'%{q}%')).fetchall()
    
    def safe(row):
        d = dict(row)
        for k,v in d.items():
            if isinstance(v, Decimal):
                d[k] = str(v)
        return d
    
    return jsonify({'products': [safe(p) for p in products]})

@app.route('/stocktake/review/<int:session_id>')
@login_required
def stocktake_session_detail(session_id):
    db = get_db()
    
    stocktake_session = db.execute('SELECT s.*, u.username as created_by_name FROM stocktake_sessions s LEFT JOIN users u ON u.id = s.created_by WHERE s.id = %s', (session_id,)).fetchone()
    if not stocktake_session:
        return "جلسة الجرد غير موجودة", 404
    
    items = db.execute('''
        SELECT si.*, p.name as product_name_db, c.name as category_name
        FROM stocktake_items si
        LEFT JOIN products p ON p.id = si.product_id
        LEFT JOIN categories c ON c.id = p.category_id
        ORDER BY si.id DESC
    ''').fetchall()
    items = [i for i in items if i['session_id'] == session_id]
    
    requests_raw = db.execute('''
        SELECT r.*, c.name as category_name
        FROM stocktake_product_requests r
        LEFT JOIN categories c ON c.id = r.category_id
        WHERE r.session_id = %s
        ORDER BY r.id DESC
    ''', (session_id,)).fetchall()
    
    # تحديد المكررات
    from collections import Counter
    barcode_counts = Counter(r['barcode'] for r in requests_raw if r['barcode'])
    dup_barcodes = {bc for bc, cnt in barcode_counts.items() if cnt > 1}
    
    requests = []
    for r in requests_raw:
        rd = dict(r)
        rd['is_duplicate'] = rd.get('barcode') in dup_barcodes
        requests.append(rd)
    
    return render_template('stocktake_session_detail.html', stocktake_session=stocktake_session, items=items, requests=requests)

@app.route('/stocktake/export/<int:session_id>')
@login_required
def stocktake_export_excel(session_id):
    db = get_db()
    
    stocktake_session = db.execute('SELECT * FROM stocktake_sessions WHERE id = %s', (session_id,)).fetchone()
    if not stocktake_session:
        return "جلسة الجرد غير موجودة", 404
    
    items = db.execute('''
        SELECT si.*, p.name as product_name_db, c.name as category_name
        FROM stocktake_items si
        LEFT JOIN products p ON p.id = si.product_id
        LEFT JOIN categories c ON c.id = p.category_id
        WHERE si.session_id = %s
        ORDER BY si.id
    ''', (session_id,)).fetchall()
    
    requests_list = db.execute('''
        SELECT r.*, c.name as category_name
        FROM stocktake_product_requests r
        LEFT JOIN categories c ON c.id = r.category_id
        WHERE r.session_id = %s
        ORDER BY r.id
    ''', (session_id,)).fetchall()
    
    # إنشاء ملف Excel
    wb = openpyxl.Workbook()
    
    # تنسيقات
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    cell_align = Alignment(horizontal='right', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ورقة الأصناف المسجلة
    ws_items = wb.active
    ws_items.title = 'الأصناف المسجلة'
    ws_items.sheet_view.rightToLeft = True
    
    items_headers = ['#', 'اسم الصنف', 'الباركود', 'القسم', 'الوحدة', 'العبوة', 'الكمية', 'الباتش', 'تاريخ الإنتاج', 'تاريخ الانتهاء', 'المدة المتبقية (يوم)', 'ملاحظة']
    for col, header in enumerate(items_headers, 1):
        cell = ws_items.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    today_date = datetime.now().date()
    for row_idx, item in enumerate(items, 2):
        days_left = None
        if item['expiry_date']:
            days_left = (item['expiry_date'] - today_date).days
        row_data = [
            row_idx - 1,
            item['product_name'] or item['product_name_db'] or '',
            item['barcode'] or '',
            item['category_name'] or '',
            item['selected_unit'] or item['unit'] or '',
            item['pack_size'] or 1,
            float(item['counted_stock'] or 0),
            item['batch_no'] or '',
            item['production_date'].strftime('%Y-%m-%d') if item['production_date'] else '',
            item['expiry_date'].strftime('%Y-%m-%d') if item['expiry_date'] else '',
            days_left if days_left is not None else '',
            item['notes'] or ''
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws_items.cell(row=row_idx, column=col, value=value)
            cell.alignment = cell_align
            cell.border = thin_border
    
    # تعديل عرض الأعمدة
    ws_items.column_dimensions['A'].width = 5
    ws_items.column_dimensions['B'].width = 40
    ws_items.column_dimensions['C'].width = 18
    ws_items.column_dimensions['D'].width = 20
    ws_items.column_dimensions['E'].width = 12
    ws_items.column_dimensions['F'].width = 10
    ws_items.column_dimensions['G'].width = 10
    ws_items.column_dimensions['H'].width = 15
    ws_items.column_dimensions['I'].width = 14
    ws_items.column_dimensions['J'].width = 14
    ws_items.column_dimensions['K'].width = 16
    ws_items.column_dimensions['L'].width = 30
    
    # تلوين خلايا المدة المتبقية
    for row_idx, item in enumerate(items, 2):
        cell = ws_items.cell(row=row_idx, column=11)
        if isinstance(cell.value, int):
            if cell.value < 0:
                cell.fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
            elif cell.value <= 30:
                cell.fill = PatternFill(start_color='FECACA', end_color='FECACA', fill_type='solid')
            elif cell.value <= 90:
                cell.fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
    
    # ورقة الطلبات غير الموجودة
    if requests_list:
        ws_requests = wb.create_sheet('طلبات غير موجود')
        ws_requests.sheet_view.rightToLeft = True
        
        req_headers = ['#', 'الباركود', 'اسم الصنف', 'القسم', 'الوحدة', 'العبوة', 'الكمية', 'تاريخ الإنتاج', 'تاريخ الانتهاء', 'المدة المتبقية (يوم)', 'الباتش', 'سعر التكلفة', 'سعر البيع', 'الحالة', 'ملاحظة']
        for col, header in enumerate(req_headers, 1):
            cell = ws_requests.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color='F59E0B', end_color='F59E0B', fill_type='solid')
            cell.alignment = header_align
            cell.border = thin_border
        
        today = datetime.now().date()
        for row_idx, req in enumerate(requests_list, 2):
            # حساب المدة المتبقية
            days_left = None
            if req['expiry_date']:
                days_left = (req['expiry_date'] - today).days
            
            status_labels = {'pending': 'قيد المراجعة', 'approved': 'معتمد', 'rejected': 'مرفوض'}
            row_data = [
                row_idx - 1,
                req['barcode'] or '',
                req['product_name'] or '',
                req['category_name'] or '',
                req['unit'] or '',
                req['pack_size'] or 1,
                float(req['quantity_counted'] or 1),
                req['production_date'].strftime('%Y-%m-%d') if req['production_date'] else '',
                req['expiry_date'].strftime('%Y-%m-%d') if req['expiry_date'] else '',
                days_left if days_left is not None else '',
                req['batch_no'] or '',
                float(req['cost_price']) if req['cost_price'] else '',
                float(req['sell_price']) if req['sell_price'] else '',
                status_labels.get(req['status'] or 'pending', req['status'] or ''),
                req['notes'] or ''
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws_requests.cell(row=row_idx, column=col, value=value)
                cell.alignment = cell_align
                cell.border = thin_border
                # تلوين خلية المدة المتبقية
                if col == 10 and isinstance(value, int):
                    if value < 0:
                        cell.fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
                    elif value <= 30:
                        cell.fill = PatternFill(start_color='FECACA', end_color='FECACA', fill_type='solid')
                    elif value <= 90:
                        cell.fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
        
        col_widths = {'A':5,'B':18,'C':40,'D':20,'E':12,'F':8,'G':10,'H':14,'I':14,'J':16,'K':15,'L':12,'M':12,'N':14,'O':30}
        for col_letter, width in col_widths.items():
            ws_requests.column_dimensions[col_letter].width = width
    
    # حفظ في buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"stocktake_{session_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/api/stocktake/edit-request', methods=['POST'])
@login_required
def api_stocktake_edit_request():
    db = get_db()
    data = request.get_json() or {}
    
    product_id = data.get('product_id')
    if not product_id:
        return jsonify({'success': False, 'message': 'الصنف مطلوب'}), 400
    
    details = (data.get('details') or '').strip()
    if not details:
        return jsonify({'success': False, 'message': 'تفاصيل التعديل مطلوبة'}), 400
    
    db.execute('''
        INSERT INTO stocktake_edit_requests (session_id, product_id, product_name, barcode, request_type, details, requested_by)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
    ''', (
        data.get('session_id'), product_id,
        data.get('product_name') or '', data.get('barcode') or '',
        data.get('request_type') or 'edit', details, session['user_id']
    ))
    db.commit()
    return jsonify({'success': True, 'message': 'تم رفع طلب تعديل الصنف للمراجعة'})

@app.route('/api/stocktake/edit-request/<int:request_id>/done', methods=['POST'])
@login_required
def api_edit_request_done(request_id):
    db = get_db()
    db.execute('UPDATE stocktake_edit_requests SET status = %s, reviewed_by = %s, reviewed_at = NOW() WHERE id = %s',
               ('done', session['user_id'], request_id))
    db.commit()
    return jsonify({'success': True, 'message': 'تم تأكيد التعديل'})

@app.route('/api/stocktake/edit-request/<int:request_id>/reject', methods=['POST'])
@login_required
def api_edit_request_reject(request_id):
    db = get_db()
    db.execute('UPDATE stocktake_edit_requests SET status = %s, reviewed_by = %s, reviewed_at = NOW() WHERE id = %s',
               ('rejected', session['user_id'], request_id))
    db.commit()
    return jsonify({'success': True, 'message': 'تم رفض الطلب'})

@app.route('/api/stocktake/scan', methods=['POST'])
@login_required
def api_stocktake_scan():
    db = get_db()
    data = request.get_json() or {}
    session_id = data.get('session_id')
    barcode = (data.get('barcode') or '').strip()

    if not session_id or not barcode:
        return jsonify({'success': False, 'message': 'بيانات ناقصة'}), 400

    stocktake_session = db.execute('SELECT * FROM stocktake_sessions WHERE id = %s AND status = %s', (session_id, 'open')).fetchone()
    if not stocktake_session:
        return jsonify({'success': False, 'message': 'جلسة الجرد غير موجودة أو مغلقة'}), 400

    try:
        lookup = api_barcode_lookup(barcode)
        result = lookup.get_json()
    except Exception as e:
        return jsonify({'success': True, 'found': False, 'barcode': barcode, 'error': str(e)})
    
    if not result.get('found'):
        return jsonify({'success': True, 'found': False, 'barcode': barcode})

    product = result['product']
    product_id = product['id']

    # جلب المنتج الأصلي من products للحصول على الوحدة والباركود الأساسيين
    original = db.execute(
        'SELECT unit, barcode, cost_price, sell_price FROM products WHERE id = %s', (product_id,)
    ).fetchone()
    original_unit = (original.get('unit') if original else None) or 'حبه'
    original_barcode = (original.get('barcode') if original else None) or ''
    original_cost = str(original.get('cost_price') or 0) if original else '0'
    original_sell = str(original.get('sell_price') or 0) if original else '0'

    # بناء قائمة الوحدات من product_barcodes
    all_pb_units = db.execute('''
        SELECT unit, barcode, pack_size, conversion_factor, cost_price, sell_price, is_purchase, is_sale
        FROM product_barcodes WHERE product_id = %s
        ORDER BY conversion_factor ASC, sort_order ASC, pack_size ASC
    ''', (product_id,)).fetchall()

    units = []
    seen_keys = set()

    # الوحدة الأساسية أولاً (conversion_factor = 1)
    base_added = False
    for u in all_pb_units:
        conv = float(u.get('conversion_factor') or u.get('pack_size') or 1)
        if conv == 1 or conv == 1.0:
            key = (u.get('unit') or original_unit, u.get('barcode') or '')
            if key not in seen_keys:
                units.append({
                    'unit': u.get('unit') or original_unit,
                    'barcode': u.get('barcode') or original_barcode,
                    'pack_size': u.get('pack_size') or 1,
                    'conversion_factor': 1,
                    'cost_price': str(u.get('cost_price') or original_cost),
                    'sell_price': str(u.get('sell_price') or original_sell),
                    'is_sale': u.get('is_sale', True),
                    'is_purchase': u.get('is_purchase', False)
                })
                seen_keys.add(key)
                base_added = True
                break

    if not base_added:
        units.append({
            'unit': original_unit, 'barcode': original_barcode,
            'pack_size': 1, 'conversion_factor': 1,
            'cost_price': original_cost, 'sell_price': original_sell,
            'is_sale': True, 'is_purchase': False
        })
        seen_keys.add((original_unit, original_barcode))

    # باقي الوحدات (الأكبر)
    for u in all_pb_units:
        key = (u.get('unit') or '', u.get('barcode') or '')
        if key not in seen_keys:
            units.append({
                'unit': u.get('unit') or 'حبه',
                'barcode': u.get('barcode') or '',
                'pack_size': u.get('pack_size') or 1,
                'conversion_factor': u.get('conversion_factor') or u.get('pack_size') or 1,
                'cost_price': str(u.get('cost_price') or 0),
                'sell_price': str(u.get('sell_price') or 0),
                'is_purchase': u.get('is_purchase', False),
                'is_sale': u.get('is_sale', True)
            })
            seen_keys.add(key)

    # تحديد الوحدة المطابقة للباركود الممسوح
    matched_unit_index = 0
    for idx, u in enumerate(units):
        if u.get('barcode') == barcode:
            matched_unit_index = idx
            break

    return jsonify({'success': True, 'found': True, 'product': product, 'units': units, 'matched_unit_index': matched_unit_index})

@app.route('/api/stocktake/save-item', methods=['POST'])
@login_required
def api_stocktake_save_item():
    db = get_db()
    try:
        session_id = int(request.form.get('session_id') or 0)
        product_id = int(request.form.get('product_id') or 0)
        barcode = (request.form.get('barcode') or '').strip()
        product_name = (request.form.get('product_name') or '').strip()
        selected_unit = (request.form.get('selected_unit') or '').strip() or None
        pack_size = int(request.form.get('pack_size') or 1)
        counted_stock = Decimal(str(request.form.get('counted_stock') or 1))
        production_date = request.form.get('production_date') or None
        expiry_date = request.form.get('expiry_date') or None
        no_expiry_dates = str(request.form.get('no_expiry_dates') or '').strip().lower() in ('1', 'true', 'yes', 'on')
        batch_no = (request.form.get('batch_no') or '').strip() or None
        notes = (request.form.get('notes') or '').strip() or None
        expected_stock = Decimal(str(request.form.get('expected_stock') or 0))

        if not session_id or not product_id:
            return jsonify({'success': False, 'message': 'بيانات ناقصة'}), 400

        if not production_date:
            return jsonify({'success': False, 'message': 'تاريخ الإنتاج مطلوب'}), 400

        if not expiry_date:
            return jsonify({'success': False, 'message': 'تاريخ الانتهاء مطلوب'}), 400

        # التحقق من عدم تكرار الصنف في نفس الجلسة
        existing = db.execute('''
            SELECT id, counted_stock, selected_unit, batch_no FROM stocktake_items 
            WHERE session_id = %s AND product_id = %s
        ''', (session_id, product_id)).fetchone()
        
        if existing:
            return jsonify({
                'success': False, 
                'message': f'الصنف مسجل مسبقًا في هذه الجلسة (الكمية: {existing["counted_stock"]}, الوحدة: {existing["selected_unit"] or "-"}, الباتش: {existing["batch_no"] or "-"})',
                'duplicate': True,
                'existing_id': existing['id']
            }), 400

        image_path = None
        attachment_path = None
        voice_note_path = None
        for field, prefix in [('image', 'stocktake_item_image'), ('attachment', 'stocktake_item_attachment'), ('voice_note', 'stocktake_item_voice')]:
            file = request.files.get(field)
            if file and file.filename:
                ext = os.path.splitext(secure_filename(file.filename))[1] or ('.webm' if field == 'voice_note' else '.jpg')
                fname = f"{prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}{ext}"
                rel = os.path.join('static', 'uploads', fname).replace('\\', '/')
                full = os.path.join(app.config['UPLOAD_FOLDER'], fname)
                file.save(full)
                if field == 'image':
                    image_path = rel
                elif field == 'attachment':
                    attachment_path = rel
                elif field == 'voice_note':
                    voice_note_path = rel

        db.execute('''
            INSERT INTO stocktake_items (
                session_id, product_id, barcode, product_name, unit, selected_unit,
                pack_size, expected_stock, counted_stock, scan_count, created_by,
                production_date, expiry_date, batch_no, notes, image_path, attachment_path, voice_note_path
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, 1, %s, %s, %s, %s, %s, %s, %s, %s)
        ''', (
            session_id, product_id, barcode, product_name, selected_unit, selected_unit,
            pack_size, expected_stock, counted_stock, session['user_id'],
            production_date, expiry_date, batch_no, notes, image_path, attachment_path, voice_note_path
        ))
        db.commit()
        return jsonify({'success': True, 'message': 'تم حفظ الصنف في الجرد', 'voice_note_path': voice_note_path or ''})
    except Exception as e:
        db.rollback()
        return jsonify({'success': False, 'message': f'فشل الحفظ: {str(e)}'}), 500

@app.route('/api/stocktake/request-product', methods=['POST'])
@login_required
def api_stocktake_request_product():
    db = get_db()
    session_id = request.form.get('session_id')
    barcode = (request.form.get('barcode') or '').strip()
    product_name = (request.form.get('product_name') or '').strip()
    category_id = request.form.get('category_id') or None
    unit = (request.form.get('unit') or '').strip() or None
    notes = (request.form.get('notes') or '').strip() or None
    pack_size = request.form.get('pack_size') or 1
    quantity_counted = Decimal(str(request.form.get('quantity_counted') or 1))
    production_date = request.form.get('production_date') or None
    expiry_date = request.form.get('expiry_date') or None
    batch_no = (request.form.get('batch_no') or '').strip() or None
    cost_price = request.form.get('cost_price') or None
    sell_price = request.form.get('sell_price') or None

    def clean_text(v):
        if v is None:
            return None
        if not isinstance(v, str):
            v = str(v)
        return v.encode('utf-8', errors='ignore').decode('utf-8', errors='ignore').strip() or None

    barcode = clean_text(barcode) or ''
    product_name = clean_text(product_name) or ''
    unit = clean_text(unit)
    notes = clean_text(notes)
    batch_no = clean_text(batch_no)

    if not barcode:
        return jsonify({'success': False, 'message': 'الباركود مطلوب'}), 400

    image_path = None
    attachment_path = None
    for field, prefix in [('image', 'stocktake_image'), ('attachment', 'stocktake_attachment')]:
        file = request.files.get(field)
        if file and file.filename:
            ext = os.path.splitext(secure_filename(file.filename))[1] or '.jpg'
            fname = f"{prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}{ext}"
            rel = os.path.join('uploads', fname).replace('\\', '/')
            full = os.path.join(app.config['UPLOAD_FOLDER'], fname)
            file.save(full)
            if field == 'image':
                image_path = rel
            else:
                attachment_path = rel

    db.execute('''
        INSERT INTO stocktake_product_requests (
            session_id, barcode, product_name, category_id, unit, notes,
            image_path, attachment_path, status, requested_by,
            pack_size, quantity_counted, production_date, expiry_date, batch_no, cost_price, sell_price
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'pending', %s, %s, %s, %s, %s, %s, %s, %s)
    ''', (
        session_id or None, barcode, product_name or None, category_id, unit, notes,
        image_path, attachment_path, session['user_id'],
        pack_size, quantity_counted, production_date, expiry_date, batch_no,
        Decimal(str(cost_price)) if cost_price else None,
        Decimal(str(sell_price)) if sell_price else None
    ))
    db.commit()
    return jsonify({'success': True, 'message': 'تم رفع طلب إضافة الصنف للمراجعة'})

@app.route('/api/barcode/<barcode>')
@login_required
def api_barcode_lookup(barcode):
    db = get_db()
    barcode = (barcode or '').strip()
    
    # تجهيز نسخ بديلة من الباركود (مع/بدون أصفار بادئة)
    barcode_variants = [barcode]
    if barcode.startswith('0'):
        barcode_variants.append(barcode.lstrip('0'))
    else:
        barcode_variants.append('0' + barcode)
    
    def to_json_safe(row):
        if row is None:
            return None
        d = {}
        for k, v in dict(row).items():
            if isinstance(v, Decimal):
                d[k] = str(v)
            elif hasattr(v, 'isoformat'):
                d[k] = v.isoformat()
            else:
                d[k] = v
        return d
    
    # البحث في الأصناف (الباركود الرئيسي)
    for bc in barcode_variants:
        product = db.execute('''
            SELECT p.*, c.name as category_name,
                   p.sell_price as price
            FROM products p
            LEFT JOIN categories c ON c.id = p.category_id
            WHERE p.barcode = %s AND p.is_active = TRUE
        ''', (bc,)).fetchone()
        if product:
            return jsonify({'found': True, 'product': to_json_safe(product)})
    
    # البحث في جدول الباركودات المتعددة (وحدات إضافية)
    multi_barcode = None
    for bc in barcode_variants:
        multi_barcode = db.execute('''
            SELECT p.*, c.name as category_name,
                   pb.unit as barcode_unit, pb.pack_size as barcode_pack_size,
                   pb.cost_price as barcode_cost, pb.sell_price as barcode_sell,
                   pb.barcode as matched_barcode
            FROM product_barcodes pb
            JOIN products p ON p.id = pb.product_id
            LEFT JOIN categories c ON c.id = p.category_id
            WHERE pb.barcode = %s AND p.is_active = TRUE
        ''', (bc,)).fetchone()
        if multi_barcode:
            break

    if multi_barcode:
        result = to_json_safe(multi_barcode)
        result['barcode'] = result.get('matched_barcode') or barcode
        result['unit'] = result.get('barcode_unit') or result.get('unit')
        result['pack_size'] = result.get('barcode_pack_size') or 1
        if result.get('barcode_sell'):
            result['sell_price'] = result['barcode_sell']
            result['price'] = result['sell_price']
        if result.get('barcode_cost'):
            result['cost_price'] = result['barcode_cost']
        return jsonify({'found': True, 'product': result, 'source': 'barcode_unit'})
    
    # البحث في وحدات الأصناف القديمة
    product_unit = None
    for bc in barcode_variants:
        product_unit = db.execute('''
            SELECT p.*, c.name as category_name, pu.sell_price as price, pu.barcode as unit_barcode, pu.unit_name as pu_unit
            FROM product_units pu
            JOIN products p ON p.id = pu.product_id
            LEFT JOIN categories c ON c.id = p.category_id
            WHERE pu.barcode = %s AND p.is_active = TRUE
        ''', (bc,)).fetchone()
        if product_unit:
            break
    
    if product_unit:
        result = to_json_safe(product_unit)
        if result.get('pu_unit'):
            result['unit'] = result['pu_unit']
        return jsonify({'found': True, 'product': result})
    
    # البحث في أصناف الموردين
    supplier_product = None
    for bc in barcode_variants:
        supplier_product = db.execute('''
            SELECT sp.*, p.name, p.id as product_id, p.current_stock, p.unit,
                   c.name as category_name, s.name as supplier_name
            FROM supplier_products sp
            JOIN products p ON p.id = sp.product_id
            LEFT JOIN categories c ON c.id = p.category_id
            LEFT JOIN suppliers s ON s.id = sp.supplier_id
            WHERE sp.supplier_barcode = %s AND sp.is_active = TRUE
        ''', (bc,)).fetchone()
        if supplier_product:
            break
    
    if supplier_product:
        return jsonify({'found': True, 'product': to_json_safe(supplier_product), 'source': 'supplier'})
    
    return jsonify({'found': False})

@app.route('/api/products/<int:product_id>/barcodes', methods=['GET', 'POST'])
@login_required
def api_product_barcodes(product_id):
    db = get_db()
    
    if request.method == 'POST':
        data = request.json or {}
        try:
            _ensure_units_registered(db, [data.get('unit')])
            db.execute('''
                INSERT INTO product_barcodes (product_id, barcode, unit, pack_size, cost_price, sell_price)
                VALUES (%s, %s, %s, %s, %s, %s)
            ''', (product_id, data.get('barcode'), data.get('unit'), 
                  data.get('pack_size', 1), data.get('cost_price', 0), data.get('sell_price', 0)))
            db.commit()
            return jsonify({'success': True})
        except Exception as e:
            db.rollback()
            return jsonify({'success': False, 'message': str(e)})
    
    barcodes = db.execute('''
        SELECT * FROM product_barcodes WHERE product_id = %s ORDER BY pack_size
    ''', (product_id,)).fetchall()
    return jsonify([dict(b) for b in barcodes])

# ═══════════════════════════════════════════════════════════════
# وحدات العرض (Display Units & Shelves)
# ═══════════════════════════════════════════════════════════════

DISPLAY_UNIT_TYPE_LABELS = {
    'stand': 'ستاند',
    'glass_stand': 'ستاند زجاجي',
    'fridge': 'ثلاجة',
    'freezer': 'فريزر',
    'other': 'أخرى'
}

def ensure_periodic_stocktake_closed_shelves_table(db):
    db.execute('''
        CREATE TABLE IF NOT EXISTS periodic_stocktake_closed_shelves (
            id SERIAL PRIMARY KEY,
            session_id INTEGER NOT NULL REFERENCES periodic_stocktake_sessions(id) ON DELETE CASCADE,
            shelf_id INTEGER NOT NULL REFERENCES shelves(id) ON DELETE CASCADE,
            closed_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
            closed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(session_id, shelf_id)
        )
    ''')

def is_periodic_shelf_closed(db, session_id, shelf_id):
    if not session_id or not shelf_id:
        return False
    ensure_periodic_stocktake_closed_shelves_table(db)
    row = db.execute('''
        SELECT 1
        FROM periodic_stocktake_closed_shelves
        WHERE session_id = %s AND shelf_id = %s
        LIMIT 1
    ''', (session_id, shelf_id)).fetchone()
    return bool(row)

@app.route('/display-units')
@login_required
def display_units_page():
    db = get_db()
    units_raw = db.execute('SELECT * FROM display_units ORDER BY is_active DESC, id ASC').fetchall()
    display_units = []
    total_shelves = 0
    active_count = 0
    type_counts = {}
    for u in units_raw:
        ud = dict(u)
        shelves = db.execute('SELECT * FROM shelves WHERE display_unit_id = %s ORDER BY shelf_number', (u['id'],)).fetchall()
        ud['shelves'] = [dict(s) for s in shelves]
        display_units.append(ud)
        total_shelves += len(shelves)
        if u['is_active']:
            active_count += 1
        label = DISPLAY_UNIT_TYPE_LABELS.get(u['type'], u['type'])
        type_counts[label] = type_counts.get(label, 0) + 1

    stats = {
        'total': len(units_raw),
        'total_shelves': total_shelves,
        'active': active_count,
        'types': type_counts
    }
    display_units_json = json.dumps([{
        'id': u['id'], 'name': u['name'], 'type': u['type'],
        'shelves_count': u['shelves_count'], 'notes': u.get('notes') or '',
        'is_active': u['is_active'], 'shelves': [{
            'id': s['id'], 'display_unit_id': s['display_unit_id'],
            'shelf_number': s['shelf_number'], 'name': s['name'],
            'notes': s.get('notes') or '', 'is_active': s['is_active']
        } for s in u['shelves']]
    } for u in display_units], ensure_ascii=False)
    return render_template('display_units.html', display_units=display_units, stats=stats, type_labels=DISPLAY_UNIT_TYPE_LABELS, display_units_json=display_units_json)


@app.route('/api/display-units', methods=['GET'])
@login_required
def api_get_display_units():
    db = get_db()
    session_id = request.args.get('session_id', type=int)
    closed_map = {}
    if session_id:
        ensure_periodic_stocktake_closed_shelves_table(db)
        closed_rows = db.execute('''
            SELECT cs.shelf_id, cs.closed_at, COALESCE(u.display_name, u.username, '') AS closed_by_name
            FROM periodic_stocktake_closed_shelves cs
            LEFT JOIN users u ON u.id = cs.closed_by
            WHERE cs.session_id = %s
        ''', (session_id,)).fetchall()
        closed_map = {
            r['shelf_id']: {
                'closed_at': str(r['closed_at']) if r.get('closed_at') else None,
                'closed_by_name': r.get('closed_by_name') or ''
            }
            for r in closed_rows
        }

    units_raw = db.execute('SELECT * FROM display_units WHERE is_active = TRUE ORDER BY id ASC').fetchall()
    result = []
    for u in units_raw:
        ud = dict(u)
        shelves = db.execute('SELECT * FROM shelves WHERE display_unit_id = %s AND is_active = TRUE ORDER BY shelf_number', (u['id'],)).fetchall()
        ud['shelves'] = [dict(s) for s in shelves]
        # Convert datetime objects to strings for JSON
        for key in ['created_at', 'updated_at']:
            if ud.get(key):
                ud[key] = str(ud[key])
        for sh in ud['shelves']:
            if sh.get('created_at'):
                sh['created_at'] = str(sh['created_at'])
            closure = closed_map.get(sh['id'])
            sh['periodic_closed'] = bool(closure)
            sh['periodic_closed_at'] = closure['closed_at'] if closure else None
            sh['periodic_closed_by_name'] = closure['closed_by_name'] if closure else ''
        result.append(ud)
    return jsonify({'success': True, 'display_units': result})


@app.route('/api/display-units', methods=['POST'])
@login_required
def api_create_display_unit():
    if session['user_id'] != 1:
        return jsonify({'success': False, 'message': 'فقط مدير النظام يستطيع إنشاء وحدات العرض'}), 403
    db = get_db()
    try:
        data = request.get_json() or {}
        name = (data.get('name') or '').strip()
        unit_type = (data.get('type') or '').strip()
        shelves_count = int(data.get('shelves_count') or 1)
        notes = (data.get('notes') or '').strip() or None

        if not name:
            return jsonify({'success': False, 'message': 'اسم الوحدة مطلوب'}), 400
        if unit_type not in ('stand', 'glass_stand', 'fridge', 'freezer', 'other'):
            return jsonify({'success': False, 'message': 'نوع الوحدة غير صالح'}), 400
        if shelves_count < 1:
            return jsonify({'success': False, 'message': 'عدد الرفوف يجب أن يكون 1 على الأقل'}), 400

        row = db.execute('''
            INSERT INTO display_units (name, type, shelves_count, notes)
            VALUES (%s, %s, %s, %s) RETURNING *
        ''', (name, unit_type, shelves_count, notes)).fetchone()

        for i in range(1, shelves_count + 1):
            db.execute('''
                INSERT INTO shelves (display_unit_id, shelf_number, name)
                VALUES (%s, %s, %s)
            ''', (row['id'], i, f'الرف {i}'))

        db.commit()
        return jsonify({'success': True, 'message': 'تم إنشاء وحدة العرض بنجاح', 'unit_id': row['id']})
    except Exception as e:
        db.rollback()
        return jsonify({'success': False, 'message': f'فشل الإنشاء: {str(e)}'}), 500


@app.route('/api/display-units/<int:unit_id>', methods=['PUT'])
@login_required
def api_update_display_unit(unit_id):
    if session.get('role') != 'manager':
        return jsonify({'success': False, 'message': 'فقط المدراء يستطيعون تعديل وحدات العرض'}), 403
    db = get_db()
    try:
        data = request.get_json() or {}
        name = (data.get('name') or '').strip()
        unit_type = (data.get('type') or '').strip()
        new_shelves_count = int(data.get('shelves_count') or 1)
        notes = (data.get('notes') or '').strip() or None

        if not name:
            return jsonify({'success': False, 'message': 'اسم الوحدة مطلوب'}), 400
        if unit_type not in ('stand', 'glass_stand', 'fridge', 'freezer', 'other'):
            return jsonify({'success': False, 'message': 'نوع الوحدة غير صالح'}), 400

        existing = db.execute('SELECT * FROM display_units WHERE id = %s', (unit_id,)).fetchone()
        if not existing:
            return jsonify({'success': False, 'message': 'وحدة العرض غير موجودة'}), 404

        db.execute('''
            UPDATE display_units SET name = %s, type = %s, shelves_count = %s, notes = %s, updated_at = NOW()
            WHERE id = %s
        ''', (name, unit_type, new_shelves_count, notes, unit_id))

        old_count = existing['shelves_count']
        if new_shelves_count > old_count:
            for i in range(old_count + 1, new_shelves_count + 1):
                db.execute('''
                    INSERT INTO shelves (display_unit_id, shelf_number, name)
                    VALUES (%s, %s, %s)
                    ON CONFLICT (display_unit_id, shelf_number) DO NOTHING
                ''', (unit_id, i, f'الرف {i}'))
        elif new_shelves_count < old_count:
            db.execute('''
                DELETE FROM shelves WHERE display_unit_id = %s AND shelf_number > %s
            ''', (unit_id, new_shelves_count))

        db.commit()
        return jsonify({'success': True, 'message': 'تم تحديث وحدة العرض بنجاح'})
    except Exception as e:
        db.rollback()
        return jsonify({'success': False, 'message': f'فشل التحديث: {str(e)}'}), 500


@app.route('/api/display-units/<int:unit_id>', methods=['DELETE'])
@login_required
def api_delete_display_unit(unit_id):
    if session['user_id'] != 1:
        return jsonify({'success': False, 'message': 'فقط مدير النظام يستطيع تعطيل وحدات العرض'}), 403
    db = get_db()
    try:
        existing = db.execute('SELECT * FROM display_units WHERE id = %s', (unit_id,)).fetchone()
        if not existing:
            return jsonify({'success': False, 'message': 'وحدة العرض غير موجودة'}), 404
        db.execute('UPDATE display_units SET is_active = FALSE, updated_at = NOW() WHERE id = %s', (unit_id,))
        db.commit()
        return jsonify({'success': True, 'message': 'تم تعطيل وحدة العرض'})
    except Exception as e:
        db.rollback()
        return jsonify({'success': False, 'message': f'فشل التعطيل: {str(e)}'}), 500


@app.route('/api/shelves/<int:shelf_id>', methods=['PUT'])
@login_required
def api_update_shelf(shelf_id):
    db = get_db()
    try:
        data = request.get_json() or {}
        name = (data.get('name') or '').strip()
        notes = (data.get('notes') or '').strip() or None

        if not name:
            return jsonify({'success': False, 'message': 'اسم الرف مطلوب'}), 400

        existing = db.execute('SELECT * FROM shelves WHERE id = %s', (shelf_id,)).fetchone()
        if not existing:
            return jsonify({'success': False, 'message': 'الرف غير موجود'}), 404

        db.execute('UPDATE shelves SET name = %s, notes = %s WHERE id = %s', (name, notes, shelf_id))
        db.commit()
        return jsonify({'success': True, 'message': 'تم تحديث الرف'})
    except Exception as e:
        db.rollback()
        return jsonify({'success': False, 'message': f'فشل التحديث: {str(e)}'}), 500

@app.route('/api/periodic-stocktake/shelf/close', methods=['POST'])
@login_required
def api_periodic_stocktake_close_shelf():
    db = get_db()
    try:
        data = request.get_json() or {}
        try:
            session_id = int(data.get('session_id') or 0)
        except Exception:
            session_id = 0
        try:
            shelf_id = int(data.get('shelf_id') or 0)
        except Exception:
            shelf_id = 0

        if not session_id or not shelf_id:
            return jsonify({'success': False, 'message': 'بيانات ناقصة'}), 400

        stocktake_session = db.execute('''
            SELECT id FROM periodic_stocktake_sessions
            WHERE id = %s AND status = %s
        ''', (session_id, 'open')).fetchone()
        if not stocktake_session:
            return jsonify({'success': False, 'message': 'جلسة الجرد غير موجودة أو مغلقة'}), 400

        shelf = db.execute('''
            SELECT s.id, s.name
            FROM shelves s
            JOIN display_units du ON du.id = s.display_unit_id
            WHERE s.id = %s AND s.is_active = TRUE AND du.is_active = TRUE
        ''', (shelf_id,)).fetchone()
        if not shelf:
            return jsonify({'success': False, 'message': 'الرف غير موجود أو غير نشط'}), 404

        ensure_periodic_stocktake_closed_shelves_table(db)
        db.execute('''
            INSERT INTO periodic_stocktake_closed_shelves (session_id, shelf_id, closed_by, closed_at)
            VALUES (%s, %s, %s, NOW())
            ON CONFLICT (session_id, shelf_id)
            DO UPDATE SET closed_by = EXCLUDED.closed_by, closed_at = NOW()
        ''', (session_id, shelf_id, session['user_id']))
        db.commit()
        return jsonify({'success': True, 'message': f'تم إغلاق الرف "{shelf["name"]}" لهذه الجلسة'})
    except Exception as e:
        db.rollback()
        return jsonify({'success': False, 'message': f'فشل إغلاق الرف: {str(e)}'}), 500

@app.route('/api/periodic-stocktake/shelf/unlock', methods=['POST'])
@login_required
def api_periodic_stocktake_unlock_shelf():
    db = get_db()
    try:
        data = request.get_json() or {}
        try:
            session_id = int(data.get('session_id') or 0)
        except Exception:
            session_id = 0
        try:
            shelf_id = int(data.get('shelf_id') or 0)
        except Exception:
            shelf_id = 0

        if not session_id or not shelf_id:
            return jsonify({'success': False, 'message': 'بيانات ناقصة'}), 400

        stocktake_session = db.execute('''
            SELECT id FROM periodic_stocktake_sessions
            WHERE id = %s AND status = %s
        ''', (session_id, 'open')).fetchone()
        if not stocktake_session:
            return jsonify({'success': False, 'message': 'جلسة الجرد غير موجودة أو مغلقة'}), 400

        shelf = db.execute('SELECT id, name FROM shelves WHERE id = %s', (shelf_id,)).fetchone()
        if not shelf:
            return jsonify({'success': False, 'message': 'الرف غير موجود'}), 404

        ensure_periodic_stocktake_closed_shelves_table(db)
        db.execute('''
            DELETE FROM periodic_stocktake_closed_shelves
            WHERE session_id = %s AND shelf_id = %s
        ''', (session_id, shelf_id))
        db.commit()
        return jsonify({'success': True, 'message': f'تم فتح قفل الرف "{shelf["name"]}" لهذه الجلسة'})
    except Exception as e:
        db.rollback()
        return jsonify({'success': False, 'message': f'فشل فتح قفل الرف: {str(e)}'}), 500


# ═══════════════════════════════════════════════════════════════
# الجرد الدوري (Periodic Stocktake) — نسخة من الجرد الأولي
# ═══════════════════════════════════════════════════════════════

def _serialize_periodic_session(row):
    if not row:
        return None
    data = dict(row)
    created_at = data.get('created_at')
    closed_at = data.get('closed_at')

    created_iso = created_at.isoformat() if created_at and hasattr(created_at, 'isoformat') else (str(created_at) if created_at else None)
    closed_iso = closed_at.isoformat() if closed_at and hasattr(closed_at, 'isoformat') else (str(closed_at) if closed_at else None)
    created_label = created_at.strftime('%Y-%m-%d %H:%M') if created_at and hasattr(created_at, 'strftime') else (str(created_at) if created_at else '')
    closed_label = closed_at.strftime('%Y-%m-%d %H:%M') if closed_at and hasattr(closed_at, 'strftime') else (str(closed_at) if closed_at else '')

    duration_seconds = None
    if created_at and hasattr(created_at, 'strftime'):
        end_at = closed_at if (closed_at and hasattr(closed_at, 'strftime')) else datetime.now()
        try:
            duration_seconds = max(0, int((end_at - created_at).total_seconds()))
        except Exception:
            duration_seconds = None

    data['created_at_iso'] = created_iso
    data['closed_at_iso'] = closed_iso
    data['created_at_label'] = created_label
    data['closed_at_label'] = closed_label
    data['duration_seconds'] = duration_seconds
    data['created_at'] = created_iso
    data['closed_at'] = closed_iso
    return data

@app.route('/periodic-stocktake')
@login_required
def periodic_stocktake_page():
    db = get_db()
    categories = db.execute('SELECT * FROM categories ORDER BY name').fetchall()
    units = db.execute('SELECT * FROM units WHERE is_active = TRUE ORDER BY name').fetchall()

    open_session = db.execute('''
        SELECT s.*, u.display_name as creator_name FROM periodic_stocktake_sessions s
        LEFT JOIN users u ON u.id = s.created_by
        WHERE s.status = 'open'
        ORDER BY s.id DESC LIMIT 1
    ''').fetchone()
    if open_session:
        open_session = _serialize_periodic_session(open_session)

    recent_items = []
    recent_requests = []
    if open_session:
        recent_items = db.execute('''
            SELECT psi.*,
                   COALESCE(NULLIF(p.name, ''), psi.product_name) AS current_product_name,
                   du.name as display_unit_name, sh.name as shelf_name
            FROM periodic_stocktake_items psi
            LEFT JOIN products p ON p.id = psi.product_id
            LEFT JOIN display_units du ON du.id = psi.display_unit_id
            LEFT JOIN shelves sh ON sh.id = psi.shelf_id
            WHERE psi.session_id = %s
            ORDER BY psi.id DESC LIMIT 20
        ''', (open_session['id'],)).fetchall()
        for item in recent_items:
            if item.get('current_product_name'):
                item['product_name'] = item['current_product_name']
        recent_requests = db.execute('''
            SELECT r.*, c.name as category_name
            FROM periodic_stocktake_product_requests r
            LEFT JOIN categories c ON c.id = r.category_id
            WHERE r.session_id = %s
            ORDER BY r.id DESC LIMIT 20
        ''', (open_session['id'],)).fetchall()

    display_units = db.execute('SELECT * FROM display_units WHERE is_active = TRUE ORDER BY id').fetchall()
    return render_template('periodic_stocktake.html', categories=categories, units=units, open_session=open_session, recent_items=recent_items, recent_requests=recent_requests, display_units=display_units)

@app.route('/api/periodic-stocktake/recent')
@login_required
def api_periodic_stocktake_recent():
    db = get_db()
    session_id = request.args.get('session_id', type=int)
    if not session_id:
        return jsonify({'items': []})
    items = db.execute('''
        SELECT psi.id,
               COALESCE(NULLIF(p.name, ''), psi.product_name) AS product_name,
               psi.barcode, psi.selected_unit, psi.counted_stock,
               psi.production_date, psi.expiry_date, psi.batch_no, psi.notes,
               du.name as display_unit_name, sh.name as shelf_name
        FROM periodic_stocktake_items psi
        LEFT JOIN products p ON p.id = psi.product_id
        LEFT JOIN display_units du ON du.id = psi.display_unit_id
        LEFT JOIN shelves sh ON sh.id = psi.shelf_id
        WHERE psi.session_id = %s
        ORDER BY psi.id DESC LIMIT 20
    ''', (session_id,)).fetchall()
    result = []
    for i in items:
        d = dict(i)
        for k in ['production_date', 'expiry_date']:
            if d.get(k) and hasattr(d[k], 'strftime'):
                d[k] = d[k].strftime('%Y-%m-%d')
        result.append(d)
    return jsonify({'items': result})

@app.route('/api/periodic-stocktake/recent-requests')
@login_required
def api_periodic_stocktake_recent_requests():
    db = get_db()
    session_id = request.args.get('session_id', type=int)
    if not session_id:
        return jsonify({'requests': []})
    reqs = db.execute('''
        SELECT product_name, barcode, quantity_counted
        FROM periodic_stocktake_product_requests WHERE session_id = %s
        ORDER BY id DESC LIMIT 20
    ''', (session_id,)).fetchall()
    return jsonify({'requests': [dict(r) for r in reqs]})

@app.route('/api/periodic-stocktake/session', methods=['POST'])
@login_required
def api_create_periodic_stocktake_session():
    db = get_db()
    
    if session['user_id'] != 1:
        return jsonify({'success': False, 'message': 'فقط مدير النظام يستطيع فتح جلسة جرد'}), 403
    
    data = request.get_json() or {}
    title = (data.get('title') or '').strip() or f"جرد دوري {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    notes = (data.get('notes') or '').strip()

    existing = db.execute('''
        SELECT * FROM periodic_stocktake_sessions
        WHERE status = 'open'
        ORDER BY id DESC LIMIT 1
    ''').fetchone()
    if existing:
        session_data = _serialize_periodic_session(existing)
        return jsonify({'success': True, 'session_id': existing['id'], 'session': session_data, 'message': 'توجد جلسة جرد مفتوحة بالفعل'})

    row = db.execute('''
        INSERT INTO periodic_stocktake_sessions (title, notes, status, created_by)
        VALUES (%s, %s, 'open', %s)
        RETURNING *
    ''', (title, notes, session['user_id'])).fetchone()
    db.commit()
    session_data = _serialize_periodic_session(row)
    return jsonify({'success': True, 'session_id': row['id'], 'session': session_data})

@app.route('/api/periodic-stocktake/session/current')
@login_required
def api_current_periodic_stocktake_session():
    db = get_db()
    open_session = db.execute('''
        SELECT * FROM periodic_stocktake_sessions
        WHERE status = 'open'
        ORDER BY id DESC LIMIT 1
    ''').fetchone()
    session_data = _serialize_periodic_session(open_session) if open_session else None
    return jsonify({'success': True, 'session': session_data})

@app.route('/api/periodic-stocktake/session/close', methods=['POST'])
@login_required
def api_close_periodic_stocktake_session():
    db = get_db()
    
    if session['user_id'] != 1:
        return jsonify({'success': False, 'message': 'فقط مدير النظام يستطيع إغلاق جلسة الجرد'}), 403
    
    data = request.get_json() or {}
    session_id = data.get('session_id')
    
    if not session_id:
        return jsonify({'success': False, 'message': 'رقم الجلسة مطلوب'}), 400
    
    stocktake_session = db.execute('SELECT * FROM periodic_stocktake_sessions WHERE id = %s', (session_id,)).fetchone()
    if not stocktake_session:
        return jsonify({'success': False, 'message': 'جلسة الجرد غير موجودة'}), 404
    
    if stocktake_session['status'] == 'closed':
        return jsonify({'success': False, 'message': 'الجلسة مغلقة بالفعل'}), 400
    
    closed_session = db.execute('''
        UPDATE periodic_stocktake_sessions
        SET status = %s, closed_at = NOW()
        WHERE id = %s
        RETURNING *
    ''', ('closed', session_id)).fetchone()
    db.commit()
    session_data = _serialize_periodic_session(closed_session) if closed_session else None
    return jsonify({'success': True, 'message': 'تم إغلاق جلسة الجرد', 'session': session_data})

@app.route('/periodic-stocktake/review')
@login_required
def periodic_stocktake_review_page():
    db = get_db()
    
    sessions = db.execute('''
        SELECT s.*, u.username as created_by_name,
               (SELECT COUNT(*) FROM periodic_stocktake_items WHERE session_id = s.id) as items_count,
               (SELECT COUNT(*) FROM periodic_stocktake_product_requests WHERE session_id = s.id) as requests_count,
               EXTRACT(EPOCH FROM (COALESCE(s.closed_at, NOW()) - s.created_at))::BIGINT as duration_seconds
        FROM periodic_stocktake_sessions s
        LEFT JOIN users u ON u.id = s.created_by
        ORDER BY s.id DESC
        LIMIT 50
    ''').fetchall()
    
    return render_template('periodic_stocktake_review.html', sessions=sessions)

@app.route('/periodic-stocktake/requests')
@login_required
def periodic_stocktake_requests_page():
    db = get_db()
    status_filter = request.args.get('status', 'all')
    
    if status_filter == 'all':
        requests_list = db.execute('''
            SELECT r.*, c.name as category_name, 
                   u.display_name as requested_by_name,
                   u2.display_name as reviewed_by_name,
                   s.title as session_title
            FROM periodic_stocktake_product_requests r
            LEFT JOIN categories c ON c.id = r.category_id
            LEFT JOIN users u ON u.id = r.requested_by
            LEFT JOIN users u2 ON u2.id = r.reviewed_by
            LEFT JOIN periodic_stocktake_sessions s ON s.id = r.session_id
            ORDER BY r.id DESC
        ''').fetchall()
    else:
        requests_list = db.execute('''
            SELECT r.*, c.name as category_name,
                   u.display_name as requested_by_name,
                   u2.display_name as reviewed_by_name,
                   s.title as session_title
            FROM periodic_stocktake_product_requests r
            LEFT JOIN categories c ON c.id = r.category_id
            LEFT JOIN users u ON u.id = r.requested_by
            LEFT JOIN users u2 ON u2.id = r.reviewed_by
            LEFT JOIN periodic_stocktake_sessions s ON s.id = r.session_id
            WHERE r.status = %s
            ORDER BY r.id DESC
        ''', (status_filter,)).fetchall()
    
    categories = db.execute('SELECT * FROM categories ORDER BY name').fetchall()
    counts = {
        'all': db.execute('SELECT COUNT(*) as c FROM periodic_stocktake_product_requests').fetchone()['c'],
        'pending': db.execute("SELECT COUNT(*) as c FROM periodic_stocktake_product_requests WHERE status='pending'").fetchone()['c'],
        'approved': db.execute("SELECT COUNT(*) as c FROM periodic_stocktake_product_requests WHERE status='approved'").fetchone()['c'],
        'rejected': db.execute("SELECT COUNT(*) as c FROM periodic_stocktake_product_requests WHERE status='rejected'").fetchone()['c'],
    }
    units = db.execute('SELECT * FROM units WHERE is_active = TRUE ORDER BY name').fetchall()
    
    all_pending = db.execute('''
        SELECT id, barcode FROM periodic_stocktake_product_requests
        WHERE barcode IS NOT NULL AND barcode != '' AND status = 'pending'
        ORDER BY id
    ''').fetchall()
    
    from collections import defaultdict
    barcode_ids_map = defaultdict(list)
    for r in all_pending:
        barcode_ids_map[r['barcode']].append(r['id'])
    
    duplicate_barcodes = {bc: ids for bc, ids in barcode_ids_map.items() if len(ids) > 1}
    
    requests_with_dup = []
    for r in requests_list:
        rd = dict(r)
        bc = rd.get('barcode')
        if bc and bc in duplicate_barcodes:
            rd['is_duplicate'] = True
            other_ids = [i for i in duplicate_barcodes[bc] if i != rd['id']]
            rd['duplicate_with'] = other_ids
        else:
            rd['is_duplicate'] = False
            rd['duplicate_with'] = []
        requests_with_dup.append(rd)
    requests_list = requests_with_dup
    
    counts['duplicates'] = len(duplicate_barcodes)
    
    edit_requests_raw = db.execute('''
        SELECT r.*, u.display_name as requested_by_name, u2.display_name as reviewed_by_name, p.name as current_name
        FROM periodic_stocktake_edit_requests r
        LEFT JOIN users u ON u.id = r.requested_by
        LEFT JOIN users u2 ON u2.id = r.reviewed_by
        LEFT JOIN products p ON p.id = r.product_id
        ORDER BY r.id DESC
    ''').fetchall()
    
    import json as json_mod
    edit_requests = []
    for er in edit_requests_raw:
        er_dict = dict(er)
        details = er_dict.get('details') or ''
        er_dict['note'] = ''
        er_dict['proposed'] = None
        if '=== البيانات المقترحة ===' in details:
            parts = details.split('=== البيانات المقترحة ===')
            er_dict['note'] = parts[0].strip()
            try:
                er_dict['proposed'] = json_mod.loads(parts[1])
            except:
                er_dict['proposed'] = None
        edit_requests.append(er_dict)
    
    return render_template('periodic_stocktake_requests.html', requests=requests_list, categories=categories, units=units, status_filter=status_filter, counts=counts, edit_requests=edit_requests)

@app.route('/api/periodic-stocktake/request/<int:request_id>/approve', methods=['POST'])
@login_required
def api_periodic_approve_request(request_id):
    db = get_db()
    req = db.execute('SELECT * FROM periodic_stocktake_product_requests WHERE id = %s', (request_id,)).fetchone()
    if not req:
        return jsonify({'success': False, 'message': 'الطلب غير موجود'}), 404
    
    if req['status'] != 'pending':
        return jsonify({'success': False, 'message': 'الطلب تم معالجته مسبقًا'}), 400
    
    data = request.get_json() or {}
    extra_units = data.get('extra_units', [])
    _ensure_units_registered(
        db,
        [req.get('unit')] + [eu.get('unit') for eu in extra_units if isinstance(eu, dict)]
    )

    product_code = f"NEW-P{request_id}"
    new_product = db.execute('''
        INSERT INTO products (product_code, name, barcode, category_id, unit, cost_price, sell_price, current_stock, is_active)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, TRUE)
        RETURNING id
    ''', (
        product_code,
        req['product_name'] or f'صنف جديد {req["barcode"]}',
        req['barcode'],
        req['category_id'],
        req['unit'] or 'حبه',
        req['cost_price'] or 0,
        req['sell_price'] or 0,
        req['quantity_counted'] or 0
    )).fetchone()
    
    for eu in extra_units:
        if eu.get('unit'):
            db.execute('''
                INSERT INTO product_barcodes (product_id, barcode, unit, pack_size, cost_price, sell_price)
                VALUES (%s, %s, %s, %s, %s, %s)
            ''', (
                new_product['id'], eu.get('barcode') or '', eu['unit'],
                int(eu.get('pack_size') or 1),
                Decimal(str(eu['cost_price'])) if eu.get('cost_price') else None,
                Decimal(str(eu['sell_price'])) if eu.get('sell_price') else None
            ))
    
    db.execute('''
        UPDATE periodic_stocktake_product_requests 
        SET status = 'approved', reviewed_by = %s, reviewed_at = NOW()
        WHERE id = %s
    ''', (session['user_id'], request_id))
    db.commit()
    
    return jsonify({'success': True, 'message': f'تم اعتماد الصنف وإضافته برقم {new_product["id"]}', 'product_id': new_product['id']})

@app.route('/api/periodic-stocktake/request/<int:request_id>/reject', methods=['POST'])
@login_required
def api_periodic_reject_request(request_id):
    db = get_db()
    req = db.execute('SELECT * FROM periodic_stocktake_product_requests WHERE id = %s', (request_id,)).fetchone()
    if not req:
        return jsonify({'success': False, 'message': 'الطلب غير موجود'}), 404
    
    db.execute('''
        UPDATE periodic_stocktake_product_requests 
        SET status = 'rejected', reviewed_by = %s, reviewed_at = NOW()
        WHERE id = %s
    ''', (session['user_id'], request_id))
    db.commit()
    
    return jsonify({'success': True, 'message': 'تم رفض الطلب'})

@app.route('/api/periodic-stocktake/request/<int:request_id>/edit', methods=['POST'])
@login_required
def api_periodic_edit_request(request_id):
    db = get_db()
    req = db.execute('SELECT * FROM periodic_stocktake_product_requests WHERE id = %s', (request_id,)).fetchone()
    if not req:
        return jsonify({'success': False, 'message': 'الطلب غير موجود'}), 404
    
    if req['status'] != 'pending':
        return jsonify({'success': False, 'message': 'لا يمكن تعديل طلب تم معالجته'}), 400
    
    data = request.get_json() or {}
    production_date = data.get('production_date') or req['production_date']
    expiry_date = data.get('expiry_date') or req['expiry_date']
    date_validation_error = _validate_production_expiry_dates(production_date, expiry_date)
    if date_validation_error:
        return jsonify({'success': False, 'message': date_validation_error}), 400
    
    db.execute('''
        UPDATE periodic_stocktake_product_requests SET
            product_name = %s, category_id = %s, unit = %s,
            quantity_counted = %s, pack_size = %s,
            cost_price = %s, sell_price = %s,
            production_date = %s, expiry_date = %s,
            batch_no = %s, notes = %s
        WHERE id = %s
    ''', (
        data.get('product_name') or req['product_name'],
        data.get('category_id') or req['category_id'],
        data.get('unit') or req['unit'],
        Decimal(str(data.get('quantity_counted') or req['quantity_counted'] or 1)),
        int(data.get('pack_size') or req['pack_size'] or 1),
        Decimal(str(data['cost_price'])) if data.get('cost_price') else req['cost_price'],
        Decimal(str(data['sell_price'])) if data.get('sell_price') else req['sell_price'],
        production_date,
        expiry_date,
        data.get('batch_no') or req['batch_no'],
        data.get('notes') or req['notes'],
        request_id
    ))
    db.commit()
    return jsonify({'success': True, 'message': 'تم تعديل الطلب'})

@app.route('/api/periodic-stocktake/request/<int:request_id>/link', methods=['POST'])
@login_required
def api_periodic_link_request(request_id):
    db = get_db()
    data = request.get_json() or {}
    product_id = data.get('product_id')
    
    if not product_id:
        return jsonify({'success': False, 'message': 'اختر الصنف المراد الربط به'}), 400
    
    req = db.execute('SELECT * FROM periodic_stocktake_product_requests WHERE id = %s', (request_id,)).fetchone()
    if not req:
        return jsonify({'success': False, 'message': 'الطلب غير موجود'}), 404
    
    product = db.execute('SELECT * FROM products WHERE id = %s', (product_id,)).fetchone()
    if not product:
        return jsonify({'success': False, 'message': 'الصنف غير موجود'}), 404
    
    db.execute('''
        INSERT INTO product_barcodes (product_id, barcode, unit, pack_size, cost_price, sell_price)
        VALUES (%s, %s, %s, %s, %s, %s)
    ''', (product_id, req['barcode'], req['unit'] or 'حبه', req['pack_size'] or 1, req['cost_price'], req['sell_price']))
    
    db.execute('''
        UPDATE periodic_stocktake_product_requests 
        SET status = 'approved', reviewed_by = %s, reviewed_at = NOW()
        WHERE id = %s
    ''', (session['user_id'], request_id))
    db.commit()
    
    return jsonify({'success': True, 'message': f'تم ربط الباركود بالصنف: {product["name"]}'})

@app.route('/periodic-stocktake/review/<int:session_id>')
@login_required
def periodic_stocktake_session_detail(session_id):
    db = get_db()
    
    stocktake_session = db.execute('SELECT s.*, u.username as created_by_name FROM periodic_stocktake_sessions s LEFT JOIN users u ON u.id = s.created_by WHERE s.id = %s', (session_id,)).fetchone()
    if not stocktake_session:
        return "جلسة الجرد غير موجودة", 404
    
    items = db.execute('''
        SELECT si.*, p.name as product_name_db, c.name as category_name,
               du.name as display_unit_name, sh.name as shelf_name
        FROM periodic_stocktake_items si
        LEFT JOIN products p ON p.id = si.product_id
        LEFT JOIN categories c ON c.id = p.category_id
        LEFT JOIN display_units du ON du.id = si.display_unit_id
        LEFT JOIN shelves sh ON sh.id = si.shelf_id
        ORDER BY si.id DESC
    ''').fetchall()
    items = [i for i in items if i['session_id'] == session_id]
    
    requests_raw = db.execute('''
        SELECT r.*, c.name as category_name
        FROM periodic_stocktake_product_requests r
        LEFT JOIN categories c ON c.id = r.category_id
        WHERE r.session_id = %s
        ORDER BY r.id DESC
    ''', (session_id,)).fetchall()
    
    from collections import Counter
    barcode_counts = Counter(r['barcode'] for r in requests_raw if r['barcode'])
    dup_barcodes = {bc for bc, cnt in barcode_counts.items() if cnt > 1}
    
    requests = []
    for r in requests_raw:
        rd = dict(r)
        rd['is_duplicate'] = rd.get('barcode') in dup_barcodes
        requests.append(rd)
    
    # حساب الأصناف المكررة عبر الرفوف
    from collections import defaultdict
    product_locations = defaultdict(list)
    for item in items:
        pid = item['product_id']
        product_locations[pid].append(item)
    
    duplicated_products = []
    for pid, locs in product_locations.items():
        if len(locs) > 1:
            total_base_qty = sum(l['counted_stock'] * (l['pack_size'] or 1) for l in locs)
            shelves_list = []
            for l in locs:
                ps = l['pack_size'] or 1
                shelves_list.append({
                    'display_unit_name': l['display_unit_name'] or '-',
                    'shelf_name': l['shelf_name'] or '-',
                    'counted_stock': l['counted_stock'],
                    'unit': l['selected_unit'] or l['unit'] or '',
                    'pack_size': ps,
                    'base_qty': l['counted_stock'] * ps,
                    'expiry_date': l['expiry_date'].strftime('%Y-%m-%d') if l['expiry_date'] and hasattr(l['expiry_date'], 'strftime') else (l['expiry_date'] or '-'),
                    'production_date': l['production_date'].strftime('%Y-%m-%d') if l['production_date'] and hasattr(l['production_date'], 'strftime') else (l['production_date'] or '-'),
                })
            duplicated_products.append({
                'product_name': locs[0]['product_name'] or locs[0]['product_name_db'] or '',
                'barcode': locs[0]['barcode'] or '',
                'total_base_qty': total_base_qty,
                'locations_count': len(locs),
                'locations': shelves_list
            })
    duplicated_products.sort(key=lambda x: x['locations_count'], reverse=True)
    
    # ملخص تجميعي لكل صنف
    aggregated_products = []
    for pid, locs in product_locations.items():
        total_base_qty = sum(l['counted_stock'] * (l['pack_size'] or 1) for l in locs)
        locations_summary = []
        for l in locs:
            ps = l['pack_size'] or 1
            locations_summary.append({
                'display_unit_name': l['display_unit_name'] or '-',
                'shelf_name': l['shelf_name'] or '-',
                'counted_stock': l['counted_stock'],
                'unit': l['selected_unit'] or l['unit'] or '',
                'pack_size': ps,
                'base_qty': l['counted_stock'] * ps,
            })
        aggregated_products.append({
            'product_name': locs[0]['product_name'] or locs[0]['product_name_db'] or '',
            'barcode': locs[0]['barcode'] or '',
            'category_name': locs[0]['category_name'] or '',
            'total_base_qty': total_base_qty,
            'locations_count': len(locs),
            'locations': locations_summary
        })
    aggregated_products.sort(key=lambda x: x['product_name'])
    
    # كل أصناف النظام مع حالة الجرد
    all_products = db.execute('''
        SELECT p.id, p.product_code, p.name, p.brand, p.barcode, p.unit,
               p.cost_price, p.sell_price, p.min_stock, p.current_stock,
               c.id as category_id, c.name as category_name
        FROM products p
        LEFT JOIN categories c ON c.id = p.category_id
        WHERE p.is_active = TRUE
        ORDER BY p.name
    ''').fetchall()
    
    # الأقسام للفلتر
    categories = db.execute('SELECT id, name FROM categories ORDER BY name').fetchall()
    
    inventoried_ids = set(product_locations.keys())
    all_products_status = []
    for p in all_products:
        pid = p['id']
        is_inventoried = pid in inventoried_ids
        agg = None
        if is_inventoried:
            locs = product_locations[pid]
            agg = {
                'total_base_qty': sum(l['counted_stock'] * (l['pack_size'] or 1) for l in locs),
                'locations_count': len(locs)
            }
        all_products_status.append({
            'id': pid,
            'product_code': p['product_code'] or '',
            'name': p['name'],
            'brand': p['brand'] or '',
            'barcode': p['barcode'] or '',
            'unit': p['unit'] or '',
            'cost_price': float(p['cost_price']) if p['cost_price'] else 0,
            'sell_price': float(p['sell_price']) if p['sell_price'] else 0,
            'min_stock': p['min_stock'] or 0,
            'current_stock': float(p['current_stock']) if p['current_stock'] else 0,
            'category_id': p['category_id'] or 0,
            'category_name': p['category_name'] or '',
            'is_inventoried': is_inventoried,
            'agg': agg
        })
    
    inventory_stats = {
        'total': len(all_products_status),
        'inventoried': sum(1 for p in all_products_status if p['is_inventoried']),
        'not_inventoried': sum(1 for p in all_products_status if not p['is_inventoried'])
    }
    
    return render_template('periodic_stocktake_session_detail.html', stocktake_session=stocktake_session, items=items, requests=requests, duplicated_products=duplicated_products, aggregated_products=aggregated_products, all_products_status=all_products_status, inventory_stats=inventory_stats, categories=categories)

@app.route('/periodic-stocktake/export/<int:session_id>')
@login_required
def periodic_stocktake_export_excel(session_id):
    db = get_db()
    
    stocktake_session = db.execute('SELECT * FROM periodic_stocktake_sessions WHERE id = %s', (session_id,)).fetchone()
    if not stocktake_session:
        return "جلسة الجرد غير موجودة", 404
    
    items = db.execute('''
        SELECT si.*, p.name as product_name_db, c.name as category_name,
               du.name as display_unit_name, sh.name as shelf_name
        FROM periodic_stocktake_items si
        LEFT JOIN products p ON p.id = si.product_id
        LEFT JOIN categories c ON c.id = p.category_id
        LEFT JOIN display_units du ON du.id = si.display_unit_id
        LEFT JOIN shelves sh ON sh.id = si.shelf_id
        WHERE si.session_id = %s
        ORDER BY si.id
    ''', (session_id,)).fetchall()
    
    requests_list = db.execute('''
        SELECT r.*, c.name as category_name
        FROM periodic_stocktake_product_requests r
        LEFT JOIN categories c ON c.id = r.category_id
        WHERE r.session_id = %s
        ORDER BY r.id
    ''', (session_id,)).fetchall()
    
    wb = openpyxl.Workbook()
    
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    cell_align = Alignment(horizontal='right', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    ws_items = wb.active
    ws_items.title = 'الأصناف المسجلة'
    ws_items.sheet_view.rightToLeft = True
    
    items_headers = ['#', '\u0627\u0633\u0645 \u0627\u0644\u0635\u0646\u0641', '\u0627\u0644\u0628\u0627\u0631\u0643\u0648\u062f', '\u0627\u0644\u0642\u0633\u0645', '\u0648\u062d\u062f\u0629 \u0627\u0644\u0639\u0631\u0636', '\u0627\u0644\u0631\u0641', '\u0627\u0644\u0648\u062d\u062f\u0629', '\u0627\u0644\u0639\u0628\u0648\u0629', '\u0627\u0644\u0643\u0645\u064a\u0629', '\u0627\u0644\u0628\u0627\u062a\u0634', '\u062a\u0627\u0631\u064a\u062e \u0627\u0644\u0625\u0646\u062a\u0627\u062c', '\u062a\u0627\u0631\u064a\u062e \u0627\u0644\u0627\u0646\u062a\u0647\u0627\u0621', '\u0627\u0644\u0645\u062f\u0629 \u0627\u0644\u0645\u062a\u0628\u0642\u064a\u0629 (\u064a\u0648\u0645)', '\u0645\u0644\u0627\u062d\u0638\u0629']
    for col, header in enumerate(items_headers, 1):
        cell = ws_items.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    today_date = datetime.now().date()
    for row_idx, item in enumerate(items, 2):
        days_left = None
        if item['expiry_date']:
            days_left = (item['expiry_date'] - today_date).days
        row_data = [
            row_idx - 1,
            item['product_name'] or item['product_name_db'] or '',
            item['barcode'] or '',
            item['category_name'] or '',
            item['display_unit_name'] or '',
            item['shelf_name'] or '',
            item['selected_unit'] or item['unit'] or '',
            item['pack_size'] or 1,
            float(item['counted_stock'] or 0),
            item['batch_no'] or '',
            item['production_date'].strftime('%Y-%m-%d') if item['production_date'] else '',
            item['expiry_date'].strftime('%Y-%m-%d') if item['expiry_date'] else '',
            days_left if days_left is not None else '',
            item['notes'] or ''
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws_items.cell(row=row_idx, column=col, value=value)
            cell.alignment = cell_align
            cell.border = thin_border
    
    
    ws_items.column_dimensions['A'].width = 5
    ws_items.column_dimensions['B'].width = 40
    ws_items.column_dimensions['C'].width = 18
    ws_items.column_dimensions['D'].width = 20
    ws_items.column_dimensions['E'].width = 14
    ws_items.column_dimensions['F'].width = 12
    ws_items.column_dimensions['G'].width = 12
    ws_items.column_dimensions['H'].width = 10
    ws_items.column_dimensions['I'].width = 10
    ws_items.column_dimensions['J'].width = 15
    ws_items.column_dimensions['K'].width = 14
    ws_items.column_dimensions['L'].width = 14
    ws_items.column_dimensions['M'].width = 16
    ws_items.column_dimensions['N'].width = 30
    
    for row_idx, item in enumerate(items, 2):
        cell = ws_items.cell(row=row_idx, column=13)
        if isinstance(cell.value, int):
            if cell.value < 0:
                cell.fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
            elif cell.value <= 30:
                cell.fill = PatternFill(start_color='FECACA', end_color='FECACA', fill_type='solid')
            elif cell.value <= 90:
                cell.fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
    
    if requests_list:
        ws_requests = wb.create_sheet('طلبات غير موجود')
        ws_requests.sheet_view.rightToLeft = True
        
        req_headers = ['#', 'الباركود', 'اسم الصنف', 'القسم', 'الوحدة', 'العبوة', 'الكمية', 'تاريخ الإنتاج', 'تاريخ الانتهاء', 'المدة المتبقية (يوم)', 'الباتش', 'سعر التكلفة', 'سعر البيع', 'الحالة', 'ملاحظة']
        for col, header in enumerate(req_headers, 1):
            cell = ws_requests.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color='F59E0B', end_color='F59E0B', fill_type='solid')
            cell.alignment = header_align
            cell.border = thin_border
        
        today = datetime.now().date()
        for row_idx, req in enumerate(requests_list, 2):
            days_left = None
            if req['expiry_date']:
                days_left = (req['expiry_date'] - today).days
            
            status_labels = {'pending': 'قيد المراجعة', 'approved': 'معتمد', 'rejected': 'مرفوض'}
            row_data = [
                row_idx - 1,
                req['barcode'] or '',
                req['product_name'] or '',
                req['category_name'] or '',
                req['unit'] or '',
                req['pack_size'] or 1,
                float(req['quantity_counted'] or 1),
                req['production_date'].strftime('%Y-%m-%d') if req['production_date'] else '',
                req['expiry_date'].strftime('%Y-%m-%d') if req['expiry_date'] else '',
                days_left if days_left is not None else '',
                req['batch_no'] or '',
                float(req['cost_price']) if req['cost_price'] else '',
                float(req['sell_price']) if req['sell_price'] else '',
                status_labels.get(req['status'] or 'pending', req['status'] or ''),
                req['notes'] or ''
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws_requests.cell(row=row_idx, column=col, value=value)
                cell.alignment = cell_align
                cell.border = thin_border
                if col == 10 and isinstance(value, int):
                    if value < 0:
                        cell.fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
                    elif value <= 30:
                        cell.fill = PatternFill(start_color='FECACA', end_color='FECACA', fill_type='solid')
                    elif value <= 90:
                        cell.fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
        
        col_widths = {'A':5,'B':18,'C':40,'D':20,'E':12,'F':8,'G':10,'H':14,'I':14,'J':16,'K':15,'L':12,'M':12,'N':14,'O':30}
        for col_letter, width in col_widths.items():
            ws_requests.column_dimensions[col_letter].width = width
    
    # === ورقة الملخص التجميعي ===
    from collections import defaultdict as _defaultdict
    product_locations_xl = _defaultdict(list)
    for item in items:
        product_locations_xl[item['product_id']].append(item)
    
    ws_summary = wb.create_sheet('ملخص تجميعي')
    ws_summary.sheet_view.rightToLeft = True
    
    sum_headers = ['#', 'اسم الصنف', 'الباركود', 'القسم', 'عدد المواقع', 'الإجمالي (حبة)', 'التوزيع']
    sum_fill = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')
    for col, header in enumerate(sum_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = sum_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    row_idx = 2
    for pid, locs in sorted(product_locations_xl.items(), key=lambda x: x[1][0].get('product_name') or x[1][0].get('product_name_db') or ''):
        total_base = sum(l['counted_stock'] * (l['pack_size'] or 1) for l in locs)
        dist_parts = []
        for l in locs:
            ps = l['pack_size'] or 1
            base = l['counted_stock'] * ps
            unit_name = l['selected_unit'] or l['unit'] or ''
            du_name = l['display_unit_name'] or '-'
            sh_name = l['shelf_name'] or '-'
            if ps > 1:
                dist_parts.append(f"{du_name}>{sh_name}: {l['counted_stock']} {unit_name} (={base} حبة)")
            else:
                dist_parts.append(f"{du_name}>{sh_name}: {base}")
        
        row_data = [
            row_idx - 1,
            locs[0].get('product_name') or locs[0].get('product_name_db') or '',
            locs[0]['barcode'] or '',
            locs[0]['category_name'] or '',
            len(locs),
            total_base,
            ' | '.join(dist_parts)
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col, value=value)
            cell.alignment = cell_align
            cell.border = thin_border
        
        # تلوين الأصناف المتكررة
        if len(locs) > 1:
            for col in range(1, 8):
                ws_summary.cell(row=row_idx, column=col).fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
        
        row_idx += 1
    
    for col_letter, width in {'A':5,'B':40,'C':18,'D':20,'E':12,'F':16,'G':60}.items():
        ws_summary.column_dimensions[col_letter].width = width
    
    # === ورقة الأصناف المتكررة ===
    dup_products_xl = {pid: locs for pid, locs in product_locations_xl.items() if len(locs) > 1}
    if dup_products_xl:
        ws_dup = wb.create_sheet('أصناف متكررة')
        ws_dup.sheet_view.rightToLeft = True
        
        dup_headers = ['#', 'اسم الصنف', 'الباركود', 'وحدة العرض', 'الرف', 'الوحدة', 'الكمية', 'بالحبة', 'تاريخ الإنتاج', 'تاريخ الانتهاء']
        dup_fill = PatternFill(start_color='EF4444', end_color='EF4444', fill_type='solid')
        for col, header in enumerate(dup_headers, 1):
            cell = ws_dup.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = dup_fill
            cell.alignment = header_align
            cell.border = thin_border
        
        row_idx = 2
        counter = 1
        colors = ['DBEAFE', 'E0E7FF']  # ألوان متناوبة للتمييز بين المنتجات
        color_idx = 0
        for pid, locs in sorted(dup_products_xl.items(), key=lambda x: x[1][0].get('product_name') or ''):
            bg = PatternFill(start_color=colors[color_idx % 2], end_color=colors[color_idx % 2], fill_type='solid')
            color_idx += 1
            total_base = sum(l['counted_stock'] * (l['pack_size'] or 1) for l in locs)
            for i, l in enumerate(locs):
                ps = l['pack_size'] or 1
                row_data = [
                    counter if i == 0 else '',
                    (l.get('product_name') or l.get('product_name_db') or '') if i == 0 else '',
                    (l['barcode'] or '') if i == 0 else '',
                    l['display_unit_name'] or '-',
                    l['shelf_name'] or '-',
                    l['selected_unit'] or l['unit'] or '',
                    l['counted_stock'],
                    l['counted_stock'] * ps,
                    l['production_date'].strftime('%Y-%m-%d') if l['production_date'] else '',
                    l['expiry_date'].strftime('%Y-%m-%d') if l['expiry_date'] else '',
                ]
                for col, value in enumerate(row_data, 1):
                    cell = ws_dup.cell(row=row_idx, column=col, value=value)
                    cell.alignment = cell_align
                    cell.border = thin_border
                    cell.fill = bg
                row_idx += 1
            
            # صف الإجمالي
            cell = ws_dup.cell(row=row_idx, column=7, value='الإجمالي:')
            cell.font = Font(bold=True)
            cell.alignment = cell_align
            cell.border = thin_border
            cell = ws_dup.cell(row=row_idx, column=8, value=total_base)
            cell.font = Font(bold=True, color='DC2626')
            cell.alignment = cell_align
            cell.border = thin_border
            row_idx += 1
            counter += 1
        
        for col_letter, width in {'A':5,'B':40,'C':18,'D':14,'E':12,'F':12,'G':10,'H':10,'I':14,'J':14}.items():
            ws_dup.column_dimensions[col_letter].width = width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"periodic_stocktake_{session_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/api/periodic-stocktake/edit-request', methods=['POST'])
@login_required
def api_periodic_stocktake_edit_request():
    db = get_db()
    data = request.get_json() or {}
    
    product_id = data.get('product_id')
    if not product_id:
        return jsonify({'success': False, 'message': 'الصنف مطلوب'}), 400
    
    details = (data.get('details') or '').strip()
    if not details:
        return jsonify({'success': False, 'message': 'تفاصيل التعديل مطلوبة'}), 400
    
    db.execute('''
        INSERT INTO periodic_stocktake_edit_requests (session_id, product_id, product_name, barcode, request_type, details, requested_by)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
    ''', (
        data.get('session_id'), product_id,
        data.get('product_name') or '', data.get('barcode') or '',
        data.get('request_type') or 'edit', details, session['user_id']
    ))
    db.commit()
    return jsonify({'success': True, 'message': 'تم رفع طلب تعديل الصنف للمراجعة'})

@app.route('/api/periodic-stocktake/edit-request/<int:request_id>/done', methods=['POST'])
@login_required
def api_periodic_edit_request_done(request_id):
    db = get_db()
    db.execute('UPDATE periodic_stocktake_edit_requests SET status = %s, reviewed_by = %s, reviewed_at = NOW() WHERE id = %s',
               ('done', session['user_id'], request_id))
    db.commit()
    return jsonify({'success': True, 'message': 'تم تأكيد التعديل'})

@app.route('/api/periodic-stocktake/edit-request/<int:request_id>/reject', methods=['POST'])
@login_required
def api_periodic_edit_request_reject(request_id):
    db = get_db()
    db.execute('UPDATE periodic_stocktake_edit_requests SET status = %s, reviewed_by = %s, reviewed_at = NOW() WHERE id = %s',
               ('rejected', session['user_id'], request_id))
    db.commit()
    return jsonify({'success': True, 'message': 'تم رفض الطلب'})

@app.route('/api/periodic-stocktake/scan', methods=['POST'])
@login_required
def api_periodic_stocktake_scan():
    db = get_db()
    data = request.get_json() or {}
    try:
        session_id = int(data.get('session_id') or 0)
    except Exception:
        session_id = 0
    barcode = (data.get('barcode') or '').strip()
    try:
        shelf_id = int(data.get('shelf_id') or 0) or None
    except Exception:
        shelf_id = None

    if not session_id or not barcode:
        return jsonify({'success': False, 'message': 'بيانات ناقصة'}), 400

    stocktake_session = db.execute('SELECT * FROM periodic_stocktake_sessions WHERE id = %s AND status = %s', (session_id, 'open')).fetchone()
    if not stocktake_session:
        return jsonify({'success': False, 'message': 'جلسة الجرد غير موجودة أو مغلقة'}), 400
    
    if shelf_id and is_periodic_shelf_closed(db, session_id, shelf_id):
        return jsonify({'success': False, 'message': 'هذا الرف مغلق وتم اعتماد جرده، اختر رفاً آخر'}), 400

    try:
        lookup = api_barcode_lookup(barcode)
        result = lookup.get_json()
    except Exception as e:
        return jsonify({'success': True, 'found': False, 'barcode': barcode, 'error': str(e)})
    
    if not result.get('found'):
        return jsonify({'success': True, 'found': False, 'barcode': barcode})

    product = result['product']
    product_id = product['id']

    # جلب المنتج الأصلي للحصول على الوحدة والباركود الأساسيين
    original = db.execute(
        'SELECT unit, barcode, cost_price, sell_price FROM products WHERE id = %s', (product_id,)
    ).fetchone()
    original_unit = (original.get('unit') if original else None) or 'حبه'
    original_barcode = (original.get('barcode') if original else None) or ''
    original_cost = str(original.get('cost_price') or 0) if original else '0'
    original_sell = str(original.get('sell_price') or 0) if original else '0'

    all_pb_units = db.execute('''
        SELECT unit, barcode, pack_size, conversion_factor, cost_price, sell_price, is_purchase, is_sale
        FROM product_barcodes WHERE product_id = %s
        ORDER BY conversion_factor ASC, sort_order ASC, pack_size ASC
    ''', (product_id,)).fetchall()

    units = []
    seen_keys = set()

    # الوحدة الأساسية أولاً
    base_added = False
    for u in all_pb_units:
        conv = float(u.get('conversion_factor') or u.get('pack_size') or 1)
        if conv == 1 or conv == 1.0:
            key = (u.get('unit') or original_unit, u.get('barcode') or '')
            if key not in seen_keys:
                units.append({
                    'unit': u.get('unit') or original_unit,
                    'barcode': u.get('barcode') or original_barcode,
                    'pack_size': u.get('pack_size') or 1,
                    'conversion_factor': 1,
                    'cost_price': str(u.get('cost_price') or original_cost),
                    'sell_price': str(u.get('sell_price') or original_sell),
                    'is_sale': u.get('is_sale', True),
                    'is_purchase': u.get('is_purchase', False)
                })
                seen_keys.add(key)
                base_added = True
                break

    if not base_added:
        units.append({
            'unit': original_unit, 'barcode': original_barcode,
            'pack_size': 1, 'conversion_factor': 1,
            'cost_price': original_cost, 'sell_price': original_sell,
            'is_sale': True, 'is_purchase': False
        })
        seen_keys.add((original_unit, original_barcode))

    for u in all_pb_units:
        key = (u.get('unit') or '', u.get('barcode') or '')
        if key not in seen_keys:
            units.append({
                'unit': u.get('unit') or 'حبه',
                'barcode': u.get('barcode') or '',
                'pack_size': u.get('pack_size') or 1,
                'conversion_factor': u.get('conversion_factor') or u.get('pack_size') or 1,
                'cost_price': str(u.get('cost_price') or 0),
                'sell_price': str(u.get('sell_price') or 0),
                'is_purchase': u.get('is_purchase', False),
                'is_sale': u.get('is_sale', True)
            })
            seen_keys.add(key)

    matched_unit_index = 0
    for idx, u in enumerate(units):
        if u.get('barcode') == barcode:
            matched_unit_index = idx
            break

    # التحقق إذا الصنف مجرود مسبقاً في نفس الجلسة
    already_scanned = db.execute('''
        SELECT id, counted_stock, selected_unit, batch_no, production_date, expiry_date, created_at
        FROM periodic_stocktake_items
        WHERE session_id = %s AND product_id = %s
        ORDER BY created_at DESC
    ''', (session_id, product_id)).fetchall()

    return jsonify({
        'success': True, 'found': True, 'product': product, 'units': units,
        'matched_unit_index': matched_unit_index,
        'already_scanned': [dict(r) for r in already_scanned] if already_scanned else []
    })

@app.route('/api/periodic-stocktake/search-scanned', methods=['GET'])
@login_required
def api_periodic_stocktake_search_scanned():
    """البحث في الأصناف المجرودة في الجلسة الحالية"""
    db = get_db()
    session_id = request.args.get('session_id')
    q = (request.args.get('q') or '').strip()
    if not session_id:
        return jsonify({'success': False, 'message': 'الجلسة مطلوبة'}), 400

    items = db.execute('''
        SELECT
            psi.id,
            psi.product_id,
            psi.barcode,
            COALESCE(NULLIF(p.name, ''), psi.product_name) AS product_name,
            psi.selected_unit,
            psi.counted_stock,
            psi.batch_no,
            psi.production_date,
            psi.expiry_date,
            psi.created_at,
            psi.pack_size,
            COALESCE(NULLIF(p.unit, ''), 'حبة') AS primary_unit,
            (COALESCE(psi.counted_stock, 0) * COALESCE(NULLIF(psi.pack_size, 0), 1)) AS primary_qty
        FROM periodic_stocktake_items psi
        LEFT JOIN products p ON p.id = psi.product_id
        WHERE psi.session_id = %s
        AND (COALESCE(NULLIF(p.name, ''), psi.product_name) ILIKE %s OR psi.barcode ILIKE %s)
        ORDER BY psi.created_at DESC
        LIMIT 20
    ''', (session_id, f'%{q}%', f'%{q}%')).fetchall()

    return jsonify({'success': True, 'items': [dict(r) for r in items]})

def _parse_ymd_date_for_validation(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if hasattr(value, 'year') and hasattr(value, 'month') and hasattr(value, 'day') and not isinstance(value, str):
        return value
    text = str(value).strip()
    if not text:
        return None
    try:
        return datetime.strptime(text[:10], '%Y-%m-%d').date()
    except Exception:
        return None

def _validate_production_expiry_dates(production_date, expiry_date):
    if not production_date or not expiry_date:
        return None
    prod = _parse_ymd_date_for_validation(production_date)
    exp = _parse_ymd_date_for_validation(expiry_date)
    if not prod or not exp:
        return None
    if prod > exp:
        return 'تاريخ الإنتاج لا يمكن أن يكون بعد تاريخ الانتهاء'
    return None

@app.route('/api/periodic-stocktake/save-item', methods=['POST'])
@login_required
def api_periodic_stocktake_save_item():
    db = get_db()
    try:
        session_id = int(request.form.get('session_id') or 0)
        product_id = int(request.form.get('product_id') or 0)
        barcode = (request.form.get('barcode') or '').strip()
        product_name = (request.form.get('product_name') or '').strip()
        selected_unit = (request.form.get('selected_unit') or '').strip() or None
        pack_size = int(request.form.get('pack_size') or 1)
        counted_stock = Decimal(str(request.form.get('counted_stock') or 1))
        production_date = request.form.get('production_date') or None
        expiry_date = request.form.get('expiry_date') or None
        no_expiry_dates = str(request.form.get('no_expiry_dates') or '').strip().lower() in ('1', 'true', 'yes', 'on')
        batch_no = (request.form.get('batch_no') or '').strip() or None
        notes = (request.form.get('notes') or '').strip() or None
        expected_stock = Decimal(str(request.form.get('expected_stock') or 0))
        display_unit_id = request.form.get('display_unit_id') or None
        shelf_id = request.form.get('shelf_id') or None
        if display_unit_id:
            display_unit_id = int(display_unit_id)
        if shelf_id:
            shelf_id = int(shelf_id)

        if not session_id or not product_id:
            return jsonify({'success': False, 'message': 'بيانات ناقصة'}), 400
        
        if shelf_id and is_periodic_shelf_closed(db, session_id, shelf_id):
            return jsonify({'success': False, 'message': 'هذا الرف مغلق وتم اعتماد جرده، لا يمكن إضافة أصناف عليه'}), 400

        if no_expiry_dates:
            production_date = None
            expiry_date = None
        else:
            if not production_date:
                return jsonify({'success': False, 'message': 'تاريخ الإنتاج مطلوب'}), 400

            if not expiry_date:
                return jsonify({'success': False, 'message': 'تاريخ الانتهاء مطلوب'}), 400

            date_validation_error = _validate_production_expiry_dates(production_date, expiry_date)
            if date_validation_error:
                return jsonify({'success': False, 'message': date_validation_error}), 400

        # Check for duplicate: same product + same shelf in same session
        if shelf_id:
            existing = db.execute('''
                SELECT id, counted_stock, selected_unit, batch_no FROM periodic_stocktake_items 
                WHERE session_id = %s AND product_id = %s AND shelf_id = %s
            ''', (session_id, product_id, shelf_id)).fetchone()
        else:
            existing = db.execute('''
                SELECT id, counted_stock, selected_unit, batch_no FROM periodic_stocktake_items 
                WHERE session_id = %s AND product_id = %s AND shelf_id IS NULL
            ''', (session_id, product_id)).fetchone()
        
        if existing:
            return jsonify({
                'success': False, 
                'message': f'الصنف مسجل مسبقًا في هذا الموقع (الكمية: {existing["counted_stock"]}, الوحدة: {existing["selected_unit"] or "-"}, الباتش: {existing["batch_no"] or "-"})',
                'duplicate': True,
                'existing_id': existing['id']
            }), 400

        image_path = None
        attachment_path = None
        voice_note_path = None
        for field, prefix in [('image', 'periodic_stocktake_item_image'), ('attachment', 'periodic_stocktake_item_attachment'), ('voice_note', 'periodic_stocktake_item_voice')]:
            file = request.files.get(field)
            if file and file.filename:
                ext = os.path.splitext(secure_filename(file.filename))[1] or ('.webm' if field == 'voice_note' else '.jpg')
                fname = f"{prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}{ext}"
                rel = os.path.join('static', 'uploads', fname).replace('\\', '/')
                full = os.path.join(app.config['UPLOAD_FOLDER'], fname)
                file.save(full)
                if field == 'image':
                    image_path = rel
                elif field == 'attachment':
                    attachment_path = rel
                elif field == 'voice_note':
                    voice_note_path = rel

        db.execute('''
            INSERT INTO periodic_stocktake_items (
                session_id, product_id, barcode, product_name, unit, selected_unit,
                pack_size, expected_stock, counted_stock, scan_count, created_by,
                production_date, expiry_date, batch_no, notes, image_path, attachment_path, voice_note_path,
                display_unit_id, shelf_id
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, 1, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ''', (
            session_id, product_id, barcode, product_name, selected_unit, selected_unit,
            pack_size, expected_stock, counted_stock, session['user_id'],
            production_date, expiry_date, batch_no, notes, image_path, attachment_path, voice_note_path,
            display_unit_id, shelf_id
        ))
        db.commit()
        return jsonify({'success': True, 'message': 'تم حفظ الصنف في الجرد', 'voice_note_path': voice_note_path or ''})

    except Exception as e:
        db.rollback()
        return jsonify({'success': False, 'message': f'فشل الحفظ: {str(e)}'}), 500

@app.route('/api/periodic-stocktake/item/<int:item_id>', methods=['GET'])
@login_required
def api_periodic_stocktake_get_item(item_id):
    db = get_db()
    item = db.execute('''
        SELECT si.*, p.name as product_name_db, du.name as display_unit_name, sh.name as shelf_name
        FROM periodic_stocktake_items si
        LEFT JOIN products p ON p.id = si.product_id
        LEFT JOIN display_units du ON du.id = si.display_unit_id
        LEFT JOIN shelves sh ON sh.id = si.shelf_id
        WHERE si.id = %s
    ''', (item_id,)).fetchone()
    if not item:
        return jsonify({'success': False, 'message': 'العنصر غير موجود'}), 404
    d = dict(item)
    for k in ['created_at', 'production_date', 'expiry_date']:
        if d.get(k) and hasattr(d[k], 'strftime'):
            d[k] = d[k].strftime('%Y-%m-%d')
    return jsonify({'success': True, 'item': d})


@app.route('/api/periodic-stocktake/item/<int:item_id>', methods=['PUT'])
@login_required
def api_periodic_stocktake_update_item(item_id):
    db = get_db()
    try:
        item = db.execute('SELECT * FROM periodic_stocktake_items WHERE id = %s', (item_id,)).fetchone()
        if not item:
            return jsonify({'success': False, 'message': 'العنصر غير موجود'}), 404
        item_session_id = int(item['session_id'] or 0)
        item_shelf_id = int(item['shelf_id']) if item.get('shelf_id') else None
        if item_shelf_id and is_periodic_shelf_closed(db, item_session_id, item_shelf_id):
            return jsonify({'success': False, 'message': 'الرف الذي يحتوي هذا الصنف مغلق. افتح قفل الرف أولاً للتعديل'}), 400

        data = request.get_json() or {}
        counted_stock = int(data.get('counted_stock', item['counted_stock']))
        production_date = data.get('production_date') or (item['production_date'].strftime('%Y-%m-%d') if item['production_date'] and hasattr(item['production_date'], 'strftime') else item['production_date'])
        expiry_date = data.get('expiry_date') or (item['expiry_date'].strftime('%Y-%m-%d') if item['expiry_date'] and hasattr(item['expiry_date'], 'strftime') else item['expiry_date'])
        batch_no = data.get('batch_no', item['batch_no'])
        notes = data.get('notes', item['notes'])
        display_unit_id = data.get('display_unit_id', item['display_unit_id'])
        shelf_id = data.get('shelf_id', item['shelf_id'])
        if shelf_id:
            try:
                shelf_id = int(shelf_id)
            except Exception:
                shelf_id = None
        if shelf_id and is_periodic_shelf_closed(db, item_session_id, shelf_id):
            return jsonify({'success': False, 'message': 'لا يمكن نقل/تعديل الصنف داخل رف مغلق. افتح القفل أولاً'}), 400

        date_validation_error = _validate_production_expiry_dates(production_date, expiry_date)
        if date_validation_error:
            return jsonify({'success': False, 'message': date_validation_error}), 400

        db.execute('''
            UPDATE periodic_stocktake_items 
            SET counted_stock = %s, production_date = %s, expiry_date = %s, 
                batch_no = %s, notes = %s, display_unit_id = %s, shelf_id = %s
            WHERE id = %s
        ''', (counted_stock, production_date, expiry_date, batch_no, notes, display_unit_id, shelf_id, item_id))
        db.commit()
        return jsonify({'success': True, 'message': 'تم تحديث العنصر'})
    except Exception as e:
        db.rollback()
        return jsonify({'success': False, 'message': f'فشل التحديث: {str(e)}'}), 500


@app.route('/api/periodic-stocktake/item/<int:item_id>', methods=['DELETE'])
@login_required
def api_periodic_stocktake_delete_item(item_id):
    db = get_db()
    try:
        item = db.execute('SELECT * FROM periodic_stocktake_items WHERE id = %s', (item_id,)).fetchone()
        if not item:
            return jsonify({'success': False, 'message': 'العنصر غير موجود'}), 404
        item_session_id = int(item['session_id'] or 0)
        item_shelf_id = int(item['shelf_id']) if item.get('shelf_id') else None
        if item_shelf_id and is_periodic_shelf_closed(db, item_session_id, item_shelf_id):
            return jsonify({'success': False, 'message': 'الرف الذي يحتوي هذا الصنف مغلق. افتح قفل الرف أولاً للحذف'}), 400
        db.execute('DELETE FROM periodic_stocktake_items WHERE id = %s', (item_id,))
        db.commit()
        return jsonify({'success': True, 'message': 'تم حذف العنصر'})
    except Exception as e:
        db.rollback()
        return jsonify({'success': False, 'message': f'فشل الحذف: {str(e)}'}), 500


@app.route('/api/periodic-stocktake/request-product', methods=['POST'])
@login_required
def api_periodic_stocktake_request_product():
    db = get_db()
    try:
        session_id = int(request.form.get('session_id') or 0)
    except Exception:
        session_id = 0
    barcode = (request.form.get('barcode') or '').strip()
    product_name = (request.form.get('product_name') or '').strip()
    category_id = request.form.get('category_id') or None
    unit = (request.form.get('unit') or '').strip() or None
    notes = (request.form.get('notes') or '').strip() or None
    pack_size = request.form.get('pack_size') or 1
    quantity_counted = Decimal(str(request.form.get('quantity_counted') or 1))
    production_date = request.form.get('production_date') or None
    expiry_date = request.form.get('expiry_date') or None
    batch_no = (request.form.get('batch_no') or '').strip() or None
    cost_price = request.form.get('cost_price') or None
    sell_price = request.form.get('sell_price') or None
    display_unit_id = request.form.get('display_unit_id') or None
    shelf_id = request.form.get('shelf_id') or None
    if display_unit_id:
        display_unit_id = int(display_unit_id)
    if shelf_id:
        shelf_id = int(shelf_id)

    if shelf_id and session_id and is_periodic_shelf_closed(db, session_id, shelf_id):
        return jsonify({'success': False, 'message': 'هذا الرف مغلق وتم اعتماد جرده، لا يمكن رفع طلبات جديدة عليه'}), 400

    def clean_text(v):
        if v is None:
            return None
        if not isinstance(v, str):
            v = str(v)
        return v.encode('utf-8', errors='ignore').decode('utf-8', errors='ignore').strip() or None

    barcode = clean_text(barcode) or ''
    product_name = clean_text(product_name) or ''
    unit = clean_text(unit)
    notes = clean_text(notes)
    batch_no = clean_text(batch_no)

    if not barcode:
        return jsonify({'success': False, 'message': 'الباركود مطلوب'}), 400

    date_validation_error = _validate_production_expiry_dates(production_date, expiry_date)
    if date_validation_error:
        return jsonify({'success': False, 'message': date_validation_error}), 400

    image_path = None
    attachment_path = None
    for field, prefix in [('image', 'periodic_stocktake_image'), ('attachment', 'periodic_stocktake_attachment')]:
        file = request.files.get(field)
        if file and file.filename:
            ext = os.path.splitext(secure_filename(file.filename))[1] or '.jpg'
            fname = f"{prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}{ext}"
            rel = os.path.join('uploads', fname).replace('\\', '/')
            full = os.path.join(app.config['UPLOAD_FOLDER'], fname)
            file.save(full)
            if field == 'image':
                image_path = rel
            else:
                attachment_path = rel

    db.execute('''
        INSERT INTO periodic_stocktake_product_requests (
            session_id, barcode, product_name, category_id, unit, notes,
            image_path, attachment_path, status, requested_by,
            pack_size, quantity_counted, production_date, expiry_date, batch_no, cost_price, sell_price,
            display_unit_id, shelf_id
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'pending', %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    ''', (
        session_id or None, barcode, product_name or None, category_id, unit, notes,
        image_path, attachment_path, session['user_id'],
        pack_size, quantity_counted, production_date, expiry_date, batch_no,
        Decimal(str(cost_price)) if cost_price else None,
        Decimal(str(sell_price)) if sell_price else None,
        display_unit_id, shelf_id
    ))
    db.commit()
    return jsonify({'success': True, 'message': 'تم رفع طلب إضافة الصنف للمراجعة'})

# ═══════════════════════════════════════════════════════════════
# تفاصيل الصنف
# ═══════════════════════════════════════════════════════════════

@app.route('/products/<int:product_id>')
@login_required
def product_detail(product_id):
    db = get_db()
    
    # بيانات الصنف
    product = db.execute('''
        SELECT p.*, c.name as category_name
        FROM products p
        LEFT JOIN categories c ON c.id = p.category_id
        WHERE p.id = %s
    ''', (product_id,)).fetchone()
    
    if not product:
        flash('الصنف غير موجود', 'error')
        return redirect(url_for('products'))
    
    # الوحدات والباركودات
    barcodes = db.execute('''
        SELECT * FROM product_barcodes WHERE product_id = %s ORDER BY pack_size
    ''', (product_id,)).fetchall()
    
    # حركات المخزون
    try:
        movements = db.execute('''
            SELECT im.*, u.display_name as user_name
            FROM inventory_movements im
            LEFT JOIN users u ON u.id = im.user_id
            WHERE im.product_id = %s
            ORDER BY im.created_at DESC
            LIMIT 20
        ''', (product_id,)).fetchall()
    except:
        movements = []
    
    # تاريخ الأسعار
    try:
        price_history = db.execute('''
            SELECT * FROM pricing_history
            WHERE product_id = %s
            ORDER BY created_at DESC
            LIMIT 20
        ''', (product_id,)).fetchall()
    except:
        price_history = []
    
    # أسعار الموردين
    try:
        supplier_prices = db.execute('''
            SELECT sp.*, s.name as supplier_name
            FROM supplier_prices sp
            JOIN suppliers s ON s.id = sp.supplier_id
            WHERE sp.product_id = %s
            ORDER BY sp.recorded_at DESC
        ''', (product_id,)).fetchall()
    except:
        supplier_prices = []
    
    return render_template('product_detail.html',
                         product=product, barcodes=barcodes,
                         movements=movements, price_history=price_history,
                         supplier_prices=supplier_prices)

@app.route('/api/barcodes/<int:barcode_id>', methods=['DELETE'])
@login_required
def api_delete_barcode(barcode_id):
    db = get_db()
    try:
        db.execute('DELETE FROM product_barcodes WHERE id = %s', (barcode_id,))
        db.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.rollback()
        return jsonify({'success': False, 'message': str(e)})

# ═══════════════════════════════════════════════════════════════
# إدارة الأصناف
# ═══════════════════════════════════════════════════════════════

@app.route('/products')
@role_required('manager')
def products():
    db = get_db()
    
    # Pagination
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    search = request.args.get('q', '').strip()
    cat_filter = request.args.get('cat', '', type=str)
    brand_filter = request.args.get('brand', '').strip()
    auto_open_add = request.args.get('open', '').strip().lower() == 'add'
    auto_edit_id = request.args.get('edit_id', type=int)
    return_to = (request.args.get('return_to', '') or '').strip()
    if not (return_to.startswith('/') and not return_to.startswith('//') and '://' not in return_to):
        return_to = ''
    
    # Build query
    where = ['p.is_active = TRUE']
    params = []
    
    if search:
        where.append("(p.name LIKE %s OR p.barcode LIKE %s OR p.product_code LIKE %s OR p.brand LIKE %s)")
        s = f'%{search}%'
        params.extend([s, s, s, s])
    
    if cat_filter:
        where.append("p.category_id = %s")
        params.append(cat_filter)
    
    if brand_filter:
        where.append("p.brand = %s")
        params.append(brand_filter)
    
    where_sql = ' AND '.join(where)
    
    # Total count
    total = db.execute(f'SELECT COUNT(*) FROM products p WHERE {where_sql}', params).fetchone()['count']
    total_pages = (total + per_page - 1) // per_page
    
    # Get page
    offset = (page - 1) * per_page
    products_list = db.execute(f'''
        SELECT p.*, c.name as category_name
        FROM products p
        LEFT JOIN categories c ON c.id = p.category_id
        WHERE {where_sql}
        ORDER BY p.category_id, p.name
        LIMIT %s OFFSET %s
    ''', params + [per_page, offset]).fetchall()
    
    categories = db.execute('SELECT * FROM categories ORDER BY id').fetchall()
    
    # Get unique brands for filter
    brands = db.execute('SELECT DISTINCT brand FROM products WHERE brand IS NOT NULL AND is_active=TRUE ORDER BY brand').fetchall()
    
    try:
        units_list = db.execute('SELECT * FROM units WHERE is_active = TRUE ORDER BY name').fetchall()
    except:
        units_list = []
    
    return render_template('products.html', 
                         products=products_list, categories=categories, units=units_list,
                         brands=brands, page=page, per_page=per_page, total=total, 
                         total_pages=total_pages, search=search, cat_filter=cat_filter,
                         brand_filter=brand_filter, auto_open_add=auto_open_add,
                         auto_edit_id=auto_edit_id, return_to=return_to)

def _normalize_unit_name(value):
    return (str(value or '')).strip()

def _collect_product_payload_unit_names(data):
    payload = data or {}
    unit_names = []

    base_unit = _normalize_unit_name(payload.get('unit'))
    if base_unit:
        unit_names.append(base_unit)

    extra_units = payload.get('extra_units') or []
    for eu in extra_units:
        if not isinstance(eu, dict):
            continue
        unit_name = _normalize_unit_name(eu.get('unit'))
        if unit_name:
            unit_names.append(unit_name)

    legacy_units = payload.get('units') or []
    for lu in legacy_units:
        if not isinstance(lu, dict):
            continue
        unit_name = _normalize_unit_name(lu.get('unit_name') or lu.get('unit'))
        if unit_name:
            unit_names.append(unit_name)

    return unit_names

def _ensure_units_registered(db, unit_names):
    added = 0
    reactivated = 0
    seen = set()

    for raw_name in unit_names or []:
        unit_name = _normalize_unit_name(raw_name)
        if not unit_name:
            continue
        key = unit_name.lower()
        if key in seen:
            continue
        seen.add(key)

        existing = db.execute(
            'SELECT id, is_active FROM units WHERE LOWER(name) = LOWER(%s) ORDER BY id LIMIT 1',
            (unit_name,)
        ).fetchone()
        if existing:
            if not existing.get('is_active'):
                db.execute('UPDATE units SET is_active = TRUE WHERE id = %s', (existing['id'],))
                reactivated += 1
            continue

        db.execute('INSERT INTO units (name, symbol, is_active) VALUES (%s, %s, TRUE)', (unit_name, ''))
        added += 1

    return {'added': added, 'reactivated': reactivated}

def _sync_units_catalog_from_products(db):
    unit_names = []

    rows = db.execute('''
        SELECT DISTINCT TRIM(unit) AS unit_name
        FROM products
        WHERE COALESCE(TRIM(unit), '') <> ''
    ''').fetchall()
    unit_names.extend([(r.get('unit_name') or '') for r in rows])

    rows = db.execute('''
        SELECT DISTINCT TRIM(unit) AS unit_name
        FROM product_barcodes
        WHERE COALESCE(TRIM(unit), '') <> ''
    ''').fetchall()
    unit_names.extend([(r.get('unit_name') or '') for r in rows])

    try:
        rows = db.execute('''
            SELECT DISTINCT TRIM(unit_name) AS unit_name
            FROM product_units
            WHERE COALESCE(TRIM(unit_name), '') <> ''
        ''').fetchall()
        unit_names.extend([(r.get('unit_name') or '') for r in rows])
    except Exception:
        pass

    return _ensure_units_registered(db, unit_names)

def _normalize_barcode_value(value):
    return (str(value or '')).strip()

def _build_barcode_conflict_message(barcode, owner):
    product_name = (owner.get('product_name') or '').strip() or 'غير معروف'
    unit_name = (owner.get('unit_name') or '').strip() or 'غير محددة'
    return f'الباركود "{barcode}" مسجل مسبقاً للصنف "{product_name}" بوحدة "{unit_name}" ولا يمكن حفظه.'

def _collect_product_payload_barcodes(data):
    payload = data or {}
    items = []

    main_barcode = _normalize_barcode_value(payload.get('barcode'))
    if main_barcode:
        items.append({
            'barcode': main_barcode,
            'unit_name': (payload.get('unit') or 'الوحدة الأساسية')
        })

    extra_units = payload.get('extra_units') or []
    for idx, eu in enumerate(extra_units):
        if not isinstance(eu, dict):
            continue
        unit_name = (eu.get('unit') or f'وحدة إضافية {idx + 1}')
        raw_barcodes = eu.get('barcodes')
        if isinstance(raw_barcodes, str):
            raw_barcodes = [raw_barcodes]
        if not isinstance(raw_barcodes, list):
            raw_barcodes = []
        fallback_barcode = eu.get('barcode')
        if fallback_barcode and not raw_barcodes:
            raw_barcodes = [fallback_barcode]

        for bc in raw_barcodes:
            normalized = _normalize_barcode_value(bc)
            if normalized:
                items.append({'barcode': normalized, 'unit_name': unit_name})

    # توافق خلفي مع نموذج product_units القديم
    legacy_units = payload.get('units') or []
    for idx, lu in enumerate(legacy_units):
        if not isinstance(lu, dict):
            continue
        normalized = _normalize_barcode_value(lu.get('barcode'))
        if normalized:
            unit_name = (lu.get('unit_name') or lu.get('unit') or f'وحدة {idx + 1}')
            items.append({'barcode': normalized, 'unit_name': unit_name})

    return items

def _find_barcode_owner(db, barcode, exclude_product_id=None):
    normalized = _normalize_barcode_value(barcode)
    if not normalized:
        return None

    try:
        exclude_id = int(exclude_product_id) if exclude_product_id not in (None, '', 0, '0') else None
    except Exception:
        exclude_id = None

    extra_filter = ''
    params = [normalized]
    if exclude_id:
        extra_filter = ' AND p.id <> %s'
        params.append(exclude_id)

    checks = [
        f'''
            SELECT p.id AS product_id,
                   p.name AS product_name,
                   COALESCE(NULLIF(p.unit, ''), 'الوحدة الأساسية') AS unit_name,
                   'products' AS source
            FROM products p
            WHERE p.is_active = TRUE AND p.barcode = %s{extra_filter}
            LIMIT 1
        ''',
        f'''
            SELECT p.id AS product_id,
                   p.name AS product_name,
                   COALESCE(NULLIF(pb.unit, ''), NULLIF(p.unit, ''), 'الوحدة الأساسية') AS unit_name,
                   'product_barcodes' AS source
            FROM product_barcodes pb
            JOIN products p ON p.id = pb.product_id
            WHERE p.is_active = TRUE AND pb.barcode = %s{extra_filter}
            LIMIT 1
        ''',
        f'''
            SELECT p.id AS product_id,
                   p.name AS product_name,
                   COALESCE(NULLIF(pb.unit, ''), NULLIF(p.unit, ''), 'الوحدة الأساسية') AS unit_name,
                   'unit_barcodes' AS source
            FROM unit_barcodes ub
            JOIN product_barcodes pb ON pb.id = ub.product_barcode_id
            JOIN products p ON p.id = pb.product_id
            WHERE p.is_active = TRUE AND ub.barcode = %s{extra_filter}
            LIMIT 1
        ''',
        f'''
            SELECT p.id AS product_id,
                   p.name AS product_name,
                   COALESCE(NULLIF(pu.unit_name, ''), NULLIF(p.unit, ''), 'الوحدة الأساسية') AS unit_name,
                   'product_units' AS source
            FROM product_units pu
            JOIN products p ON p.id = pu.product_id
            WHERE p.is_active = TRUE AND pu.barcode = %s{extra_filter}
            LIMIT 1
        '''
    ]

    for query in checks:
        owner = db.execute(query, tuple(params)).fetchone()
        if owner:
            return owner
    return None

def _validate_product_barcodes(db, data, exclude_product_id=None):
    barcode_items = _collect_product_payload_barcodes(data)
    if not barcode_items:
        return None

    seen = {}
    for item in barcode_items:
        bc = item['barcode']
        first = seen.get(bc)
        if first:
            return {
                'success': False,
                'duplicate': True,
                'barcode': bc,
                'message': f'الباركود "{bc}" مكرر داخل نفس الصنف بين وحدة "{first["unit_name"]}" ووحدة "{item["unit_name"]}".'
            }
        seen[bc] = item

    for item in barcode_items:
        owner = _find_barcode_owner(db, item['barcode'], exclude_product_id=exclude_product_id)
        if owner:
            return {
                'success': False,
                'duplicate': True,
                'barcode': item['barcode'],
                'existing_product_id': owner.get('product_id'),
                'existing_product_name': owner.get('product_name'),
                'existing_unit_name': owner.get('unit_name'),
                'input_unit_name': item.get('unit_name'),
                'message': _build_barcode_conflict_message(item['barcode'], owner)
            }
    return None

@app.route('/api/barcodes/check-duplicate', methods=['POST'])
@role_required('manager')
def api_check_barcode_duplicate():
    db = get_db()
    data = request.json or {}
    barcode = _normalize_barcode_value(data.get('barcode'))
    exclude_product_id = data.get('exclude_product_id')

    if not barcode:
        return jsonify({'success': True, 'duplicate': False, 'barcode': ''})

    owner = _find_barcode_owner(db, barcode, exclude_product_id=exclude_product_id)
    if owner:
        return jsonify({
            'success': True,
            'duplicate': True,
            'barcode': barcode,
            'existing_product_id': owner.get('product_id'),
            'existing_product_name': owner.get('product_name'),
            'existing_unit_name': owner.get('unit_name'),
            'message': _build_barcode_conflict_message(barcode, owner)
        })

    return jsonify({'success': True, 'duplicate': False, 'barcode': barcode})

@app.route('/api/products', methods=['GET', 'POST'])
@role_required('manager')
def api_products():
    db = get_db()
    
    if request.method == 'POST':
        data = request.json or {}
        _ensure_units_registered(db, _collect_product_payload_unit_names(data))
        conflict = _validate_product_barcodes(db, data)
        if conflict:
            return jsonify(conflict), 400
        try:
            cursor = db.execute('''
                INSERT INTO products (name, name_en, receipt_name, barcode, product_code, brand, category_id, unit, pack_size, cost_price, sell_price, min_stock, notes)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ''', (
                data['name'], data.get('name_en'), data.get('receipt_name'),
                data.get('barcode'), data.get('product_code'), data.get('brand'),
                data.get('category_id') or None, data.get('unit', 'حبه'),
                data.get('pack_size', 1), data.get('cost_price', 0), data.get('sell_price', 0),
                data.get('min_stock', 0), data.get('notes')
            ))
            product_id = db.execute('SELECT lastval() AS id').fetchone()['id']
            
            # حفظ الوحدات الإضافية (product_barcodes)
            extra_units = data.get('extra_units', [])
            for idx, eu in enumerate(extra_units):
                if eu.get('unit'):
                    main_bc = eu.get('barcodes', [eu.get('barcode', '')])[0] if eu.get('barcodes') or eu.get('barcode') else ''
                    pb_row = db.execute('''
                        INSERT INTO product_barcodes (product_id, barcode, unit, pack_size, conversion_factor, cost_price, sell_price, is_purchase, is_sale, sort_order)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        RETURNING id
                    ''', (product_id, main_bc, eu['unit'],
                          int(eu.get('conversion_factor') or eu.get('pack_size') or 1),
                          int(eu.get('conversion_factor') or eu.get('pack_size') or 1),
                          Decimal(str(eu.get('cost_price') or 0)),
                          Decimal(str(eu.get('sell_price') or 0)),
                          eu.get('is_purchase', True),
                          eu.get('is_sale', False),
                          idx + 2
                    )).fetchone()
                    # Save extra barcodes
                    if pb_row:
                        for bc in eu.get('barcodes', []):
                            if bc:
                                db.execute('INSERT INTO unit_barcodes (product_barcode_id, barcode, is_primary) VALUES (%s, %s, %s)',
                                          (pb_row['id'], bc, bc == main_bc))
            
            # حفظ الوحدات
            units = data.get('units', [])
            for unit_data in units:
                db.execute('''
                    INSERT INTO product_units (product_id, unit_id, conversion_factor, barcode, price, is_default)
                    VALUES (%s, %s, %s, %s, %s, %s)
                ''', (
                    product_id, unit_data['unit_id'],
                    unit_data.get('conversion_factor', 1),
                    unit_data.get('barcode', ''),
                    unit_data.get('price', 0),
                    1 if unit_data.get('is_default') else 0
                ))
            
            db.commit()
            return jsonify({'success': True, 'message': 'تم إضافة الصنف', 'id': product_id})
        except Exception as e:
            db.rollback()
            import traceback
            traceback.print_exc()
            err_msg = str(e)
            if 'unique' in err_msg.lower() or 'duplicate' in err_msg.lower():
                return jsonify({'success': False, 'message': 'الباركود أو البيانات مكررة'}), 400
            if 'foreign key' in err_msg.lower() or 'violates foreign key' in err_msg.lower():
                return jsonify({'success': False, 'message': 'القسم المحدد غير موجود'}), 400
            return jsonify({'success': False, 'message': f'خطأ في حفظ الصنف: {err_msg}'}), 500
    
    products_list = db.execute('''
        SELECT p.*, c.name as category_name
        FROM products p
        LEFT JOIN categories c ON c.id = p.category_id
        WHERE p.is_active = TRUE
        ORDER BY p.name
    ''').fetchall()
    return jsonify([dict(p) for p in products_list])

@app.route('/api/products/<int:id>', methods=['GET', 'PUT', 'DELETE'])
@role_required('manager')
def api_product(id):
    db = get_db()

    if request.method == 'GET':
        product = db.execute('''
            SELECT id, name, brand, product_code, barcode, category_id, unit, cost_price, sell_price
            FROM products
            WHERE id = %s AND is_active = TRUE
            LIMIT 1
        ''', (id,)).fetchone()
        if not product:
            return jsonify({'success': False, 'message': 'الصنف غير موجود'}), 404
        return jsonify({'success': True, 'product': dict(product)})
    
    if request.method == 'PUT':
        data = request.json or {}
        _ensure_units_registered(db, _collect_product_payload_unit_names(data))
        conflict = _validate_product_barcodes(db, data, exclude_product_id=id)
        if conflict:
            return jsonify(conflict), 400
        try:
            db.execute('''
                UPDATE products SET name=%s, name_en=%s, receipt_name=%s, barcode=%s,
                       category_id=%s, unit=%s, min_stock=%s, notes=%s
                WHERE id=%s
            ''', (
                data['name'], data.get('name_en'), data.get('receipt_name'),
                data.get('barcode'), data.get('category_id') or None, data.get('unit'),
                data.get('min_stock', 0), data.get('notes'), id
            ))
            
            # تحديث الوحدات الإضافية (product_barcodes)
            extra_units = data.get('extra_units')
            if extra_units is not None:
                db.execute('DELETE FROM unit_barcodes WHERE product_barcode_id IN (SELECT id FROM product_barcodes WHERE product_id = %s)', (id,))
                db.execute('DELETE FROM product_barcodes WHERE product_id = %s', (id,))
                for idx, eu in enumerate(extra_units):
                    if eu.get('unit'):
                        main_bc = eu.get('barcodes', [''])[0] if eu.get('barcodes') else ''
                        pb_row = db.execute('''
                            INSERT INTO product_barcodes (product_id, barcode, unit, pack_size, conversion_factor, cost_price, sell_price, is_purchase, is_sale, sort_order)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                            RETURNING id
                        ''', (id, main_bc, eu['unit'],
                              int(eu.get('conversion_factor') or eu.get('pack_size') or 1),
                              int(eu.get('conversion_factor') or eu.get('pack_size') or 1),
                              Decimal(str(eu.get('cost_price') or 0)),
                              Decimal(str(eu.get('sell_price') or 0)),
                              eu.get('is_purchase', True),
                              eu.get('is_sale', False),
                              idx + 2
                        )).fetchone()
                        if pb_row:
                            for bc in eu.get('barcodes', []):
                                if bc:
                                    db.execute('INSERT INTO unit_barcodes (product_barcode_id, barcode, is_primary) VALUES (%s, %s, %s)',
                                              (pb_row['id'], bc, bc == main_bc))
            
            # تحديث الوحدات القديمة
            units = data.get('units', [])
            if units:
                # حذف الوحدات القديمة وإضافة الجديدة
                db.execute('DELETE FROM product_units WHERE product_id = %s', (id,))
                for unit_data in units:
                    db.execute('''
                        INSERT INTO product_units (product_id, unit_id, conversion_factor, barcode, price, is_default)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    ''', (
                        id, unit_data['unit_id'],
                        unit_data.get('conversion_factor', 1),
                        unit_data.get('barcode', ''),
                        unit_data.get('price', 0),
                        1 if unit_data.get('is_default') else 0
                    ))
            
            db.commit()
            return jsonify({'success': True, 'message': 'تم تحديث الصنف'})
        except Exception as e:
            db.rollback()
            import traceback
            traceback.print_exc()
            err_msg = str(e)
            if 'unique' in err_msg.lower() or 'duplicate' in err_msg.lower():
                return jsonify({'success': False, 'message': 'الباركود أو البيانات مكررة'}), 400
            return jsonify({'success': False, 'message': f'خطأ في تحديث الصنف: {err_msg}'}), 500
    
    elif request.method == 'DELETE':
        db.execute('UPDATE products SET is_active=FALSE WHERE id=%s', (id,))
        db.commit()
        return jsonify({'success': True, 'message': 'تم حذف الصنف'})

# ═══════════════════════════════════════════════════════════════
# وحدات الصنف المتعددة
# ═══════════════════════════════════════════════════════════════

@app.route('/api/products/<int:product_id>/units', methods=['GET', 'POST'])
@role_required('manager')
def api_product_units(product_id):
    db = get_db()
    
    if request.method == 'POST':
        data = request.json
        unit_id = data.get('unit_id')
        
        # التحقق من عدم التكرار
        existing = db.execute(
            'SELECT id FROM product_units WHERE product_id = %s AND unit_id = %s',
            (product_id, unit_id)
        ).fetchone()
        if existing:
            return jsonify({'success': False, 'message': 'هذه الوحدة مضافة مسبقاً لهذا الصنف'})
        
        # إذا كانت أول وحدة، اجعلها افتراضية
        count = db.execute('SELECT COUNT(*) as count FROM product_units WHERE product_id = %s', (product_id,)).fetchone()['count']
        is_default = 1 if count == 0 else (1 if data.get('is_default') else 0)
        
        # إذا تم تعيينها كافتراضية، ألغِ الافتراضي السابق
        if is_default:
            db.execute('UPDATE product_units SET is_default = 0 WHERE product_id = %s', (product_id,))
        
        db.execute('''
            INSERT INTO product_units (product_id, unit_id, conversion_factor, barcode, price, is_default)
            VALUES (%s, %s, %s, %s, %s, %s)
        ''', (
            product_id, unit_id,
            data.get('conversion_factor', 1),
            data.get('barcode', ''),
            data.get('price', 0),
            is_default
        ))
        db.commit()
        return jsonify({'success': True, 'message': 'تم إضافة الوحدة للصنف'})
    
    # GET - جلب وحدات الصنف
    units = db.execute('''
        SELECT pu.*, u.name as unit_name, u.symbol
        FROM product_units pu
        JOIN units u ON u.id = pu.unit_id
        WHERE pu.product_id = %s
        ORDER BY pu.is_default DESC, u.name
    ''', (product_id,)).fetchall()
    return jsonify([dict(u) for u in units])

@app.route('/api/product-units/<int:id>', methods=['PUT', 'DELETE'])
@role_required('manager')
def api_product_unit(id):
    db = get_db()
    
    if request.method == 'PUT':
        data = request.json
        
        # إذا تم تعيينها كافتراضية، ألغِ الافتراضي السابق
        if data.get('is_default'):
            product_id = db.execute('SELECT product_id FROM product_units WHERE id = %s', (id,)).fetchone()['count']
            db.execute('UPDATE product_units SET is_default = 0 WHERE product_id = %s', (product_id,))
        
        db.execute('''
            UPDATE product_units SET conversion_factor=%s, barcode=%s, price=%s, is_default=%s
            WHERE id=%s
        ''', (
            data.get('conversion_factor', 1),
            data.get('barcode', ''),
            data.get('price', 0),
            1 if data.get('is_default') else 0,
            id
        ))
        db.commit()
        return jsonify({'success': True, 'message': 'تم تحديث الوحدة'})
    
    elif request.method == 'DELETE':
        # التحقق من أنها ليست الوحدة الوحيدة
        pu = db.execute('SELECT product_id, is_default FROM product_units WHERE id = %s', (id,)).fetchone()
        if pu:
            count = db.execute('SELECT COUNT(*) as count FROM product_units WHERE product_id = %s', (pu['product_id'],)).fetchone()['count']
            if count <= 1:
                return jsonify({'success': False, 'message': 'لا يمكن حذف الوحدة الوحيدة للصنف'})
            
            db.execute('DELETE FROM product_units WHERE id = %s', (id,))
            
            # إذا كانت الافتراضية، اجعل أخرى افتراضية
            if pu['is_default']:
                db.execute('''
                    UPDATE product_units SET is_default = 1 
                    WHERE product_id = %s AND id = (SELECT MIN(id) FROM product_units WHERE product_id = %s)
                ''', (pu['product_id'], pu['product_id']))
            
            db.commit()
        return jsonify({'success': True, 'message': 'تم حذف الوحدة'})

# ═══════════════════════════════════════════════════════════════
# ربط أسماء المورد + إعدادات الفرز
# ═══════════════════════════════════════════════════════════════

@app.route('/supplier-products')
@role_required('manager')
def supplier_products():
    db = get_db()
    supplier_id = request.args.get('supplier_id', type=int)
    
    query = '''
        SELECT sp.*, s.name as supplier_name,
               p.name as product_name,
               (SELECT COUNT(*) FROM sorting_rules WHERE supplier_product_id = sp.id) as linked_count
        FROM supplier_products sp
        JOIN suppliers s ON s.id = sp.supplier_id
        LEFT JOIN products p ON p.id = sp.product_id
    '''
    if supplier_id:
        query += ' WHERE sp.supplier_id = %s'
        items = db.execute(query + ' ORDER BY sp.supplier_product_name', (supplier_id,)).fetchall()
    else:
        items = db.execute(query + ' ORDER BY s.name, sp.supplier_product_name').fetchall()
    
    suppliers = db.execute('SELECT * FROM suppliers WHERE is_active=TRUE ORDER BY name').fetchall()
    products_list = db.execute('SELECT * FROM products WHERE is_active=TRUE ORDER BY name').fetchall()
    units_list = db.execute('SELECT * FROM units WHERE is_active=TRUE ORDER BY name').fetchall()
    
    return render_template('supplier_products.html', 
                         items=items, suppliers=suppliers, 
                         products=products_list, selected_supplier=supplier_id,
                         units=units_list)

@app.route('/api/supplier-products', methods=['POST'])
@role_required('manager')
def api_supplier_products():
    db = get_db()
    data = request.json
    
    # دعم الإضافة السريعة من نافذة الفاتورة
    if data.get('quick_add'):
        supplier_id = data.get('supplier_id')
        product_id = data.get('product_id') or None
        supplier_product_name = data.get('supplier_product_name', '').strip()
        pack_size = data.get('pack_size', 1)
        supplier_unit = (data.get('supplier_unit') or 'كرتون').strip()
        is_kit = bool(data.get('is_kit'))
        
        if not supplier_id or not supplier_product_name:
            return jsonify({'success': False, 'message': 'بيانات ناقصة'})
        if not is_kit and not product_id:
            return jsonify({'success': False, 'message': 'اختر صنفاً داخلياً أو فعّل خيار الطقم'})
        
        row = db.execute('''
            INSERT INTO supplier_products (supplier_id, product_id, supplier_product_name, supplier_unit, pack_size, is_kit)
            VALUES (%s, %s, %s, %s, %s, %s)
            RETURNING id
        ''', (supplier_id, product_id, supplier_product_name, supplier_unit, pack_size, is_kit)).fetchone()
        db.commit()
        
        new_id = row['id']
        
        # إرجاع البيانات الكاملة
        new_item = db.execute('''
            SELECT sp.*, p.name as product_name
            FROM supplier_products sp
            LEFT JOIN products p ON p.id = sp.product_id
            WHERE sp.id = %s
        ''', (new_id,)).fetchone()
        
        return jsonify({'success': True, 'id': new_id, 'product': dict(new_item)})
    
    # الطريقة القديمة
    allowed_products = data.get('allowed_products', [])
    primary_product_id = allowed_products[0] if allowed_products else None
    
    cursor = db.execute('''
        INSERT INTO supplier_products (supplier_id, supplier_product_name, supplier_product_code, 
                                       supplier_barcode, supplier_unit, product_id, pack_size, 
                                       purchase_price, min_order_qty, notes)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    ''', (
        data['supplier_id'], data['supplier_product_name'], 
        data.get('supplier_product_code'), data.get('supplier_barcode'),
        data.get('supplier_unit', 'كرتون'), primary_product_id,
        data.get('pack_size', 1), data.get('purchase_price', 0),
        data.get('min_order_qty', 1), data.get('notes')
    ))
    supplier_product_id = cursor.lastrowid
    
    for product_id in allowed_products:
        db.execute('''
            INSERT OR IGNORE INTO sorting_rules (supplier_product_id, allowed_product_id)
            VALUES (%s, %s)
        ''', (supplier_product_id, product_id))
    
    db.commit()
    msg = f'تم إضافة الصنف مع ربط {len(allowed_products)} أصناف' if allowed_products else 'تم إضافة الصنف'
    return jsonify({'success': True, 'id': supplier_product_id, 'message': msg})

@app.route('/api/supplier-products/<int:id>/link', methods=['POST'])
@role_required('manager')
def api_supplier_product_link(id):
    """ربط صنف المورد بصنف داخلي"""
    db = get_db()
    data = request.json
    product_id = data.get('product_id')
    db.execute('UPDATE supplier_products SET product_id = %s WHERE id = %s', (product_id, id))
    db.commit()
    return jsonify({'success': True, 'message': 'تم تحديث الربط'})

@app.route('/api/sorting-rules/<int:supplier_product_id>', methods=['GET', 'POST'])
@role_required('manager')
def api_sorting_rules(supplier_product_id):
    db = get_db()
    
    if request.method == 'POST':
        data = request.json
        # حذف القواعد القديمة
        db.execute('DELETE FROM sorting_rules WHERE supplier_product_id = %s', (supplier_product_id,))
        # إضافة الجديدة
        for product_id in data.get('allowed_products', []):
            db.execute('''
                INSERT INTO sorting_rules (supplier_product_id, allowed_product_id)
                VALUES (%s, %s)
            ''', (supplier_product_id, product_id))
        db.commit()
        return jsonify({'success': True, 'message': 'تم حفظ إعدادات الفرز'})
    
    rules = db.execute('''
        SELECT sr.*, p.name as product_name
        FROM sorting_rules sr
        JOIN products p ON p.id = sr.allowed_product_id
        WHERE sr.supplier_product_id = %s
    ''', (supplier_product_id,)).fetchall()
    
    return jsonify([dict(r) for r in rules])

@app.route('/api/supplier-products/<int:id>', methods=['GET', 'PUT', 'DELETE'])
@role_required('manager')
def api_supplier_product(id):
    db = get_db()
    
    if request.method == 'GET':
        item = db.execute('''
            SELECT sp.*, s.name as supplier_name,
                   p.name as product_name
            FROM supplier_products sp
            JOIN suppliers s ON s.id = sp.supplier_id
            LEFT JOIN products p ON p.id = sp.product_id
            WHERE sp.id = %s
        ''', (id,)).fetchone()
        
        if not item:
            return jsonify({'success': False, 'message': 'الصنف غير موجود'}), 404
        
        # جلب قواعد الفرز
        rules = db.execute('''
            SELECT allowed_product_id FROM sorting_rules WHERE supplier_product_id = %s
        ''', (id,)).fetchall()
        
        result = dict(item)
        result['allowed_products'] = [r['allowed_product_id'] for r in rules]
        return jsonify(result)
    
    elif request.method == 'PUT':
        data = request.json
        
        # تحديث بيانات صنف المورد
        db.execute('''
            UPDATE supplier_products SET 
                supplier_product_name = %s,
                supplier_product_code = %s,
                supplier_barcode = %s,
                supplier_unit = %s,
                pack_size = %s,
                purchase_price = %s,
                min_order_qty = %s,
                notes = %s
            WHERE id = %s
        ''', (
            data['supplier_product_name'],
            data.get('supplier_product_code'),
            data.get('supplier_barcode'),
            data.get('supplier_unit', 'كرتون'),
            data.get('pack_size', 1),
            data.get('purchase_price', 0),
            data.get('min_order_qty', 1),
            data.get('notes'),
            id
        ))
        
        # تحديث قواعد الفرز
        allowed_products = data.get('allowed_products', [])
        if allowed_products:
            db.execute('DELETE FROM sorting_rules WHERE supplier_product_id = %s', (id,))
            # تحديث الصنف الرئيسي
            db.execute('UPDATE supplier_products SET product_id = %s WHERE id = %s', (allowed_products[0], id))
            for product_id in allowed_products:
                db.execute('''
                    INSERT INTO sorting_rules (supplier_product_id, allowed_product_id)
                    VALUES (%s, %s)
                ''', (id, product_id))
        
        db.commit()
        return jsonify({'success': True, 'message': 'تم تحديث الصنف'})
    
    elif request.method == 'DELETE':
        # التحقق من عدم استخدامه في فواتير
        used = db.execute('''
            SELECT COUNT(*) as count FROM supplier_invoice_items WHERE supplier_product_id = %s
        ''', (id,)).fetchone()
        
        if used['count'] > 0:
            return jsonify({'success': False, 'message': 'لا يمكن حذف هذا الصنف لأنه مستخدم في فواتير'})
        
        db.execute('DELETE FROM sorting_rules WHERE supplier_product_id = %s', (id,))
        db.execute('DELETE FROM supplier_products WHERE id = %s', (id,))
        db.commit()
        return jsonify({'success': True, 'message': 'تم حذف الصنف'})

# ═══════════════════════════════════════════════════════════════
# أسعار الموردين + المقارنة
# ═══════════════════════════════════════════════════════════════

@app.route('/supplier-prices')
@role_required('manager')
def supplier_prices():
    db = get_db()
    
    # مقارنة أسعار الموردين لكل صنف
    comparison = db.execute('''
        SELECT p.id, p.name, p.unit,
               STRING_AGG(s.name || ':' || COALESCE(sp.unit_price::TEXT, '0'), '|') as prices
        FROM products p
        JOIN supplier_prices sp ON sp.product_id = p.id
        JOIN suppliers s ON s.id = sp.supplier_id
        WHERE p.is_active = TRUE
        GROUP BY p.id
        ORDER BY p.name
    ''').fetchall()
    
    suppliers = db.execute('SELECT * FROM suppliers WHERE is_active=TRUE ORDER BY name').fetchall()
    products_list = db.execute('SELECT * FROM products WHERE is_active=TRUE ORDER BY name').fetchall()
    
    return render_template('supplier_prices.html', 
                         comparison=comparison, 
                         suppliers=suppliers, 
                         products=products_list)

@app.route('/api/supplier-prices', methods=['POST'])
@role_required('manager')
def api_supplier_prices():
    db = get_db()
    data = request.json
    
    # تحديث أو إضافة السعر
    existing = db.execute('''
        SELECT id FROM supplier_prices 
        WHERE supplier_id = %s AND product_id = %s
    ''', (data['supplier_id'], data['product_id'])).fetchone()
    
    if existing:
        db.execute('''
            UPDATE supplier_prices SET unit_price=%s, pack_price=%s, pack_size=%s, effective_date=%s
            WHERE id=%s
        ''', (data['unit_price'], data.get('pack_price'), data.get('pack_size', 1), 
              datetime.now().isoformat(), existing['id']))
    else:
        db.execute('''
            INSERT INTO supplier_prices (supplier_id, product_id, unit_price, pack_price, pack_size)
            VALUES (%s, %s, %s, %s, %s)
        ''', (data['supplier_id'], data['product_id'], data['unit_price'],
              data.get('pack_price'), data.get('pack_size', 1)))
    
    db.commit()
    return jsonify({'success': True, 'message': 'تم حفظ السعر'})

# ═══════════════════════════════════════════════════════════════
# أسعار المنافسين
# ═══════════════════════════════════════════════════════════════

@app.route('/competitor-prices')
@role_required('manager')
def competitor_prices():
    db = get_db()
    competitors = db.execute('SELECT * FROM competitors WHERE is_active=TRUE ORDER BY name').fetchall()
    products_list = db.execute('SELECT * FROM products WHERE is_active=TRUE ORDER BY name').fetchall()
    
    prices = db.execute('''
        SELECT cp.*, c.name as competitor_name, p.name as product_name
        FROM competitor_prices cp
        JOIN competitors c ON c.id = cp.competitor_id
        JOIN products p ON p.id = cp.product_id
        ORDER BY p.name, c.name
    ''').fetchall()
    
    return render_template('competitor_prices.html', 
                         competitors=competitors, 
                         products=products_list, 
                         prices=prices)

@app.route('/api/competitor-prices', methods=['POST'])
@role_required('manager')
def api_competitor_prices():
    db = get_db()
    data = request.json
    
    db.execute('''
        INSERT INTO competitor_prices (competitor_id, product_id, price)
        VALUES (%s, %s, %s)
    ''', (data['competitor_id'], data['product_id'], data['price']))
    db.commit()
    
    return jsonify({'success': True, 'message': 'تم حفظ السعر'})

# ═══════════════════════════════════════════════════════════════
# تسعيرات البيع
# ═══════════════════════════════════════════════════════════════

@app.route('/pricing')
@role_required('manager')
def pricing():
    db = get_db()
    price_lists = db.execute('SELECT * FROM price_lists ORDER BY is_default DESC, name').fetchall()
    products_list = db.execute('SELECT * FROM products WHERE is_active=TRUE ORDER BY name').fetchall()
    
    # جلب كل الأسعار
    prices = db.execute('''
        SELECT pp.*, p.name as product_name, pl.name as price_list_name
        FROM product_prices pp
        JOIN products p ON p.id = pp.product_id
        JOIN price_lists pl ON pl.id = pp.price_list_id
        ORDER BY p.name, pl.is_default DESC
    ''').fetchall()
    
    return render_template('pricing.html', 
                         price_lists=price_lists, 
                         products=products_list, 
                         prices=prices)

@app.route('/api/product-prices', methods=['POST'])
@role_required('manager')
def api_product_prices():
    db = get_db()
    data = request.json
    
    # تحديث أو إضافة السعر
    existing = db.execute('''
        SELECT id FROM product_prices 
        WHERE product_id = %s AND price_list_id = %s
    ''', (data['product_id'], data['price_list_id'])).fetchone()
    
    if existing:
        db.execute('UPDATE product_prices SET price=%s WHERE id=%s', (data['price'], existing['id']))
    else:
        db.execute('''
            INSERT INTO product_prices (product_id, price_list_id, price)
            VALUES (%s, %s, %s)
        ''', (data['product_id'], data['price_list_id'], data['price']))
    
    db.commit()
    return jsonify({'success': True, 'message': 'تم حفظ السعر'})

# ═══════════════════════════════════════════════════════════════
# إدارة المخازن
# ═══════════════════════════════════════════════════════════════

@app.route('/warehouses')
@role_required('manager')
def warehouses_page():
    db = get_db()
    warehouses_list = db.execute('''
        SELECT w.*, u.display_name as manager_name
        FROM warehouses w
        LEFT JOIN users u ON u.id = w.manager_id
        ORDER BY w.id
    ''').fetchall()
    users_list = db.execute("SELECT id, display_name FROM users WHERE is_active=TRUE ORDER BY display_name").fetchall()
    warehouses_json = [dict(w) for w in warehouses_list]
    return render_template('warehouses.html', warehouses=warehouses_list, warehouses_json=warehouses_json, users=users_list)

@app.route('/api/warehouses-crud', methods=['POST'])
@role_required('manager')
def api_warehouses_crud():
    db = get_db()
    data = request.get_json(silent=True) or {}
    name = (data.get('name') or '').strip()
    wtype = data.get('type', 'warehouse')
    address = data.get('address', '')
    manager_id = data.get('manager_id') or None
    
    if not name:
        return jsonify({'success': False, 'message': 'اسم المخزن مطلوب'})
    
    row = db.execute('''
        INSERT INTO warehouses (name, type, address, manager_id) VALUES (%s, %s, %s, %s) RETURNING id
    ''', (name, wtype, address, manager_id)).fetchone()
    db.commit()
    return jsonify({'success': True, 'id': row['id'], 'message': 'تم إضافة المخزن'})

@app.route('/api/warehouses-crud/<int:wid>', methods=['PUT', 'DELETE'])
@role_required('manager')
def api_warehouse_crud_detail(wid):
    db = get_db()
    if request.method == 'DELETE':
        db.execute('UPDATE warehouses SET is_active = FALSE WHERE id = %s', (wid,))
        db.commit()
        return jsonify({'success': True, 'message': 'تم حذف المخزن'})
    
    data = request.get_json(silent=True) or {}
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'success': False, 'message': 'اسم المخزن مطلوب'})
    
    db.execute('''
        UPDATE warehouses SET name=%s, type=%s, address=%s, manager_id=%s WHERE id=%s
    ''', (name, data.get('type', 'warehouse'), data.get('address', ''), data.get('manager_id') or None, wid))
    db.commit()
    return jsonify({'success': True, 'message': 'تم تحديث المخزن'})

# ═══════════════════════════════════════════════════════════════
# التحويلات الداخلية
# ═══════════════════════════════════════════════════════════════

@app.route('/internal-transfers')
@role_required('manager', 'warehouse')
def internal_transfers_page():
    db = get_db()
    transfers = db.execute('''
        SELECT it.*, fw.name as from_name, tw.name as to_name, u.display_name as created_by_name
        FROM internal_transfers it
        JOIN warehouses fw ON fw.id = it.from_warehouse_id
        JOIN warehouses tw ON tw.id = it.to_warehouse_id
        LEFT JOIN users u ON u.id = it.created_by
        ORDER BY it.created_at DESC
    ''').fetchall()
    warehouses_list = db.execute("SELECT id, name, type FROM warehouses WHERE is_active = TRUE ORDER BY name").fetchall()
    return render_template('internal_transfers.html', transfers=transfers, warehouses=warehouses_list)

@app.route('/api/internal-transfers', methods=['POST'])
@role_required('manager', 'warehouse')
def api_create_transfer():
    db = get_db()
    data = request.get_json(silent=True) or {}
    from_id = data.get('from_warehouse_id')
    to_id = data.get('to_warehouse_id')
    items = data.get('items', [])
    notes = data.get('notes', '')
    
    if not from_id or not to_id or from_id == to_id:
        return jsonify({'success': False, 'message': 'اختر مخزنين مختلفين'})
    if not items:
        return jsonify({'success': False, 'message': 'أضف صنف واحد على الأقل'})
    
    row = db.execute('''
        INSERT INTO internal_transfers (from_warehouse_id, to_warehouse_id, notes, created_by)
        VALUES (%s, %s, %s, %s) RETURNING id
    ''', (from_id, to_id, notes, session['user_id'])).fetchone()
    transfer_id = row['id']
    
    for item in items:
        pid = item.get('product_id')
        qty = float(item.get('quantity', 0))
        if pid and qty > 0:
            db.execute('''
                INSERT INTO internal_transfer_items (transfer_id, product_id, quantity)
                VALUES (%s, %s, %s)
            ''', (transfer_id, pid, qty))
    
    db.commit()
    return jsonify({'success': True, 'id': transfer_id, 'message': 'تم إنشاء التحويل'})

@app.route('/api/internal-transfers/<int:tid>/complete', methods=['POST'])
@role_required('manager', 'warehouse')
def api_complete_transfer(tid):
    db = get_db()
    transfer = db.execute('SELECT * FROM internal_transfers WHERE id = %s', (tid,)).fetchone()
    if not transfer:
        return jsonify({'success': False, 'message': 'التحويل غير موجود'})
    if transfer['status'] != 'pending':
        return jsonify({'success': False, 'message': 'التحويل ليس في حالة انتظار'})
    
    items = db.execute('SELECT * FROM internal_transfer_items WHERE transfer_id = %s', (tid,)).fetchall()
    
    for item in items:
        pid = item['product_id']
        qty = float(item['quantity'])
        
        # خصم من المخزن المرسل
        db.execute('''
            INSERT INTO warehouse_stock (product_id, warehouse_id, quantity) VALUES (%s, %s, 0)
            ON CONFLICT (product_id, warehouse_id) DO NOTHING
        ''', (pid, transfer['from_warehouse_id']))
        db.execute('''
            UPDATE warehouse_stock SET quantity = quantity - %s
            WHERE product_id = %s AND warehouse_id = %s
        ''', (qty, pid, transfer['from_warehouse_id']))
        
        # إضافة للمخزن المستلم
        db.execute('''
            INSERT INTO warehouse_stock (product_id, warehouse_id, quantity) VALUES (%s, %s, 0)
            ON CONFLICT (product_id, warehouse_id) DO NOTHING
        ''', (pid, transfer['to_warehouse_id']))
        db.execute('''
            UPDATE warehouse_stock SET quantity = quantity + %s
            WHERE product_id = %s AND warehouse_id = %s
        ''', (qty, pid, transfer['to_warehouse_id']))
        
        # حركات المخزون
        db.execute('''
            INSERT INTO inventory_movements (product_id, movement_type, quantity, reference_type, reference_id, warehouse_id, created_by)
            VALUES (%s, 'out', %s, 'transfer', %s, %s, %s)
        ''', (pid, qty, tid, transfer['from_warehouse_id'], session['user_id']))
        db.execute('''
            INSERT INTO inventory_movements (product_id, movement_type, quantity, reference_type, reference_id, warehouse_id, created_by)
            VALUES (%s, 'in', %s, 'transfer', %s, %s, %s)
        ''', (pid, qty, tid, transfer['to_warehouse_id'], session['user_id']))
    
    db.execute("UPDATE internal_transfers SET status = 'completed', completed_at = NOW() WHERE id = %s", (tid,))
    db.commit()
    return jsonify({'success': True, 'message': 'تم استلام التحويل وتحديث المخزون'})

@app.route('/api/internal-transfers/<int:tid>/cancel', methods=['POST'])
@role_required('manager', 'warehouse')
def api_cancel_transfer(tid):
    db = get_db()
    db.execute("UPDATE internal_transfers SET status = 'cancelled' WHERE id = %s AND status = 'pending'", (tid,))
    db.commit()
    return jsonify({'success': True, 'message': 'تم إلغاء التحويل'})

@app.route('/api/internal-transfers/<int:tid>/items')
@role_required('manager', 'warehouse')
def api_transfer_items(tid):
    db = get_db()
    items = db.execute('''
        SELECT iti.*, p.name as product_name
        FROM internal_transfer_items iti
        JOIN products p ON p.id = iti.product_id
        WHERE iti.transfer_id = %s
    ''', (tid,)).fetchall()
    return jsonify([dict(i) for i in items])

# ═══════════════════════════════════════════════════════════════
# API أطقم الموردين
# ═══════════════════════════════════════════════════════════════

@app.route('/api/supplier-products/<int:sp_id>/kit-items', methods=['GET', 'POST'])
@role_required('manager')
def api_kit_items(sp_id):
    db = get_db()
    
    if request.method == 'POST':
        data = request.get_json(silent=True) or {}
        items = data.get('items', [])
        # حذف القديمة
        db.execute('DELETE FROM sorting_rules WHERE supplier_product_id = %s', (sp_id,))
        # إضافة الجديدة
        for item in items:
            pid = item.get('product_id')
            if pid:
                db.execute('''
                    INSERT INTO sorting_rules (supplier_product_id, allowed_product_id)
                    VALUES (%s, %s) ON CONFLICT DO NOTHING
                ''', (sp_id, pid))
        # تحديث is_kit
        db.execute('UPDATE supplier_products SET is_kit = %s WHERE id = %s', (len(items) > 0, sp_id))
        db.commit()
        return jsonify({'success': True, 'message': 'تم تحديث مكونات الطقم'})
    
    # GET
    items = db.execute('''
        SELECT sr.allowed_product_id as product_id, p.name as product_name
        FROM sorting_rules sr
        JOIN products p ON p.id = sr.allowed_product_id
        WHERE sr.supplier_product_id = %s
    ''', (sp_id,)).fetchall()
    return jsonify([dict(i) for i in items])

# ═══════════════════════════════════════════════════════════════
# فواتير المورد
# ═══════════════════════════════════════════════════════════════

@app.route('/supplier-invoices')
@role_required('manager', 'agent')
def supplier_invoices():
    db = get_db()
    
    invoices = db.execute('''
        SELECT si.*, s.name as supplier_name, u.display_name as created_by_name,
               w.name as warehouse_name
        FROM supplier_invoices si
        JOIN suppliers s ON s.id = si.supplier_id
        JOIN users u ON u.id = si.created_by
        LEFT JOIN warehouses w ON w.id = si.warehouse_id
        ORDER BY si.created_at DESC
    ''').fetchall()
    
    suppliers = db.execute('SELECT * FROM suppliers WHERE is_active=TRUE ORDER BY name').fetchall()
    
    return render_template('supplier_invoices.html', invoices=invoices, suppliers=suppliers)

@app.route('/supplier-invoices/<int:id>')
@role_required('manager', 'agent')
def view_supplier_invoice(id):
    db = get_db()
    
    invoice = db.execute('''
        SELECT si.*, s.name as supplier_name, s.phone as supplier_phone,
               u.display_name as created_by_name
        FROM supplier_invoices si
        JOIN suppliers s ON s.id = si.supplier_id
        JOIN users u ON u.id = si.created_by
        WHERE si.id = %s
    ''', (id,)).fetchone()
    
    if not invoice:
        flash('الفاتورة غير موجودة', 'error')
        return redirect(url_for('supplier_invoices'))
    
    items = db.execute('''
        SELECT sii.*, sp.supplier_product_name, sp.pack_size, p.name as product_name
        FROM supplier_invoice_items sii
        JOIN supplier_products sp ON sp.id = sii.supplier_product_id
        JOIN products p ON p.id = sp.product_id
        WHERE sii.invoice_id = %s
    ''', (id,)).fetchall()
    
    return render_template('supplier_invoice_view.html', invoice=invoice, items=items)

@app.route('/api/supplier-invoices/<int:id>/pay', methods=['POST'])
@role_required('manager', 'agent')
def pay_supplier_invoice(id):
    db = get_db()
    data = request.get_json(silent=True) or {}
    try:
        amount = float(data.get('amount', 0))
    except (TypeError, ValueError):
        return jsonify({'success': False, 'message': 'المبلغ غير صحيح'})
    
    if amount <= 0:
        return jsonify({'success': False, 'message': 'المبلغ غير صحيح'})
    
    invoice = db.execute(
        'SELECT id, total_amount, paid_amount FROM supplier_invoices WHERE id = %s',
        (id,)
    ).fetchone()
    
    if not invoice:
        return jsonify({'success': False, 'message': 'الفاتورة غير موجودة'}), 404
    
    remaining = float(invoice['total_amount']) - float(invoice['paid_amount'])
    if remaining <= 0:
        return jsonify({'success': False, 'message': 'الفاتورة مسددة بالكامل'})
    
    if amount > remaining:
        return jsonify({'success': False, 'message': f'المبلغ أكبر من المتبقي ({remaining:,.0f} ريال)'})
    
    db.execute('UPDATE supplier_invoices SET paid_amount = paid_amount + %s WHERE id = %s', (amount, id))
    db.commit()
    
    return jsonify({'success': True, 'message': f'تم تسجيل دفعة {amount:,.0f} ريال'})

@app.route('/supplier-invoices/new', methods=['GET', 'POST'])
@role_required('manager', 'agent')
def new_supplier_invoice():
    db = get_db()
    
    if request.method == 'POST':
        data = request.get_json(silent=True) or {}
        items = data.get('items') or []
        
        try:
            supplier_id = int(data.get('supplier_id'))
        except (TypeError, ValueError):
            return jsonify({'success': False, 'message': 'المورد غير صحيح'})
        
        if not data.get('invoice_date'):
            return jsonify({'success': False, 'message': 'تاريخ الفاتورة مطلوب'})
        
        warehouse_id = data.get('warehouse_id')
        if not warehouse_id:
            return jsonify({'success': False, 'message': 'وجهة الاستلام مطلوبة'})
        
        if not items:
            return jsonify({'success': False, 'message': 'يجب إضافة بند واحد على الأقل'})
        
        supplier = db.execute(
            'SELECT id FROM suppliers WHERE id = %s AND is_active = TRUE',
            (supplier_id,)
        ).fetchone()
        if not supplier:
            return jsonify({'success': False, 'message': 'المورد غير موجود أو غير نشط'})
        
        # حساب الإجمالي
        total_amount = 0
        for item in items:
            total_amount += float(item.get('total_price', 0))
        
        # إنشاء رقم فاتورة تلقائي
        last = db.execute('SELECT MAX(id) as max_id FROM supplier_invoices').fetchone()['max_id'] or 0
        invoice_number = f"SI-{datetime.now().strftime('%Y%m')}-{last+1:04d}"
        
        row = db.execute('''
            INSERT INTO supplier_invoices (invoice_number, supplier_id, invoice_date, total_amount, warehouse_id, notes, created_by)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        ''', (
            invoice_number, supplier_id, data['invoice_date'],
            total_amount, warehouse_id, data.get('notes'), session['user_id']
        )).fetchone()
        invoice_id = row['id']
        
        # إضافة البنود
        for item in items:
            item_type = item.get('type', 'direct')
            
            if item_type == 'direct':
                product_id = item.get('product_id')
                quantity = float(item.get('quantity', 0))
                unit_price = float(item.get('unit_price', 0))
                line_total = float(item.get('total_price', 0))
                unit = item.get('unit', '')
                conversion_factor = float(item.get('conversion_factor', 1))
                base_quantity = quantity * conversion_factor
                supplier_product_id = item.get('supplier_product_id') or None
                
                db.execute('''
                    INSERT INTO supplier_invoice_items 
                    (invoice_id, product_id, supplier_product_id, quantity, unit_price, total_price, unit, conversion_factor, base_quantity, is_kit_item)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, FALSE)
                ''', (invoice_id, product_id, supplier_product_id, quantity, unit_price, line_total, unit, conversion_factor, base_quantity))
            
            elif item_type == 'kit':
                supplier_product_id = item.get('supplier_product_id')
                quantity = float(item.get('quantity', 0))
                unit_price = float(item.get('unit_price', 0))
                line_total = float(item.get('total_price', 0))
                
                # بند الطقم الأب
                parent_row = db.execute('''
                    INSERT INTO supplier_invoice_items 
                    (invoice_id, supplier_product_id, quantity, unit_price, total_price, is_kit_item)
                    VALUES (%s, %s, %s, %s, %s, FALSE)
                    RETURNING id
                ''', (invoice_id, supplier_product_id, quantity, unit_price, line_total)).fetchone()
                parent_item_id = parent_row['id']
                
                # البنود الفرعية
                sub_items = item.get('sub_items', [])
                for sub in sub_items:
                    sub_pid = sub.get('product_id')
                    sub_qty = float(sub.get('quantity', 0))
                    if sub_pid and sub_qty > 0:
                        db.execute('''
                            INSERT INTO supplier_invoice_items 
                            (invoice_id, product_id, quantity, unit_price, total_price, is_kit_item, parent_item_id)
                            VALUES (%s, %s, %s, 0, 0, TRUE, %s)
                        ''', (invoice_id, sub_pid, sub_qty, parent_item_id))
        
        db.commit()
        return jsonify({'success': True, 'id': invoice_id, 'invoice_number': invoice_number})
    
    suppliers = db.execute('SELECT * FROM suppliers WHERE is_active=TRUE ORDER BY name').fetchall()
    warehouses_list = db.execute('SELECT id, name, type FROM warehouses WHERE is_active = TRUE ORDER BY name').fetchall()
    return render_template('supplier_invoice_form.html', suppliers=suppliers, warehouses=warehouses_list)

@app.route('/api/supplier/<int:supplier_id>/products')
@role_required('manager', 'agent')
def api_supplier_product_list(supplier_id):
    db = get_db()
    items = db.execute('''
        SELECT sp.*, p.name as product_name
        FROM supplier_products sp
        LEFT JOIN products p ON p.id = sp.product_id
        WHERE sp.supplier_id = %s
        ORDER BY sp.supplier_product_name
    ''', (supplier_id,)).fetchall()
    return jsonify([dict(i) for i in items])

# ═══════════════════════════════════════════════════════════════
# فاتورة المندوب (ملغاة — redirect)
# ═══════════════════════════════════════════════════════════════

@app.route('/agent-invoices')
@role_required('manager', 'agent')
def agent_invoices():
    flash('تم نقل نظام الفرز إلى نظام المشتريات الجديد', 'info')
    return redirect(url_for('supplier_invoices'))

# (sort_invoice, submit_agent_invoice routes removed — system migrated to direct receiving)

# ═══════════════════════════════════════════════════════════════
# استلام المخزن (نظام جديد — استلام مباشر من فواتير المورد)
# ═══════════════════════════════════════════════════════════════

@app.route('/warehouse')
@role_required('manager', 'warehouse')
def warehouse():
    db = get_db()
    warehouse_id = request.args.get('warehouse_id', type=int)
    
    # فواتير بانتظار الاستلام
    q = '''
        SELECT si.*, s.name as supplier_name, u.display_name as created_by_name,
               w.name as warehouse_name
        FROM supplier_invoices si
        JOIN suppliers s ON s.id = si.supplier_id
        JOIN users u ON u.id = si.created_by
        LEFT JOIN warehouses w ON w.id = si.warehouse_id
        WHERE si.status = 'pending'
    '''
    params = []
    if warehouse_id:
        q += ' AND si.warehouse_id = %s'
        params.append(warehouse_id)
    q += ' ORDER BY si.created_at'
    pending = db.execute(q, params).fetchall() if params else db.execute(q).fetchall()
    
    # آخر الاستلامات
    q2 = '''
        SELECT si.*, s.name as supplier_name, u.display_name as created_by_name,
               w.name as warehouse_name
        FROM supplier_invoices si
        JOIN suppliers s ON s.id = si.supplier_id
        JOIN users u ON u.id = si.created_by
        LEFT JOIN warehouses w ON w.id = si.warehouse_id
        WHERE si.status = 'received'
    '''
    if warehouse_id:
        q2 += ' AND si.warehouse_id = %s'
    q2 += ' ORDER BY si.created_at DESC LIMIT 10'
    recent = db.execute(q2, (warehouse_id,)).fetchall() if warehouse_id else db.execute(q2).fetchall()
    
    warehouses_list = db.execute("SELECT id, name, type FROM warehouses WHERE is_active = TRUE ORDER BY name").fetchall()
    return render_template('warehouse.html', pending=pending, recent=recent, 
                         warehouses=warehouses_list, selected_warehouse_id=warehouse_id)

@app.route('/api/warehouse/receive/<int:invoice_id>', methods=['POST'])
@role_required('manager', 'warehouse')
def receive_invoice(invoice_id):
    db = get_db()
    
    invoice = db.execute('SELECT * FROM supplier_invoices WHERE id = %s', (invoice_id,)).fetchone()
    if not invoice:
        return jsonify({'success': False, 'message': 'الفاتورة غير موجودة'}), 404
    
    if invoice['status'] == 'received':
        return jsonify({'success': False, 'message': 'تم استلام هذه الفاتورة مسبقا'})
    
    if invoice['status'] != 'pending':
        return jsonify({'success': False, 'message': 'لا يمكن استلام فاتورة ليست في حالة انتظار'})
    
    warehouse_id = invoice['warehouse_id']
    
    # جلب بنود الفاتورة (بنود مباشرة + بنود فرعية للأطقم)
    items = db.execute('''
        SELECT sii.product_id, sii.quantity, sii.base_quantity, sii.conversion_factor, sii.is_kit_item
        FROM supplier_invoice_items sii
        WHERE sii.invoice_id = %s AND sii.product_id IS NOT NULL
    ''', (invoice_id,)).fetchall()
    
    if not items:
        return jsonify({'success': False, 'message': 'لا توجد بنود للاستلام'})
    
    for item in items:
        pid = item['product_id']
        # للبنود المباشرة: نستخدم base_quantity (الكمية × معامل التحويل)
        # للبنود الفرعية: quantity هي الكمية النهائية
        if item['is_kit_item']:
            qty = float(item['quantity'])
        else:
            qty = float(item['base_quantity']) if item['base_quantity'] else float(item['quantity'])
        
        # تحديث المخزون العام
        db.execute('UPDATE products SET current_stock = current_stock + %s WHERE id = %s', (qty, pid))
        
        # تحديث مخزون المخزن
        if warehouse_id:
            db.execute('''
                INSERT INTO warehouse_stock (product_id, warehouse_id, quantity) VALUES (%s, %s, 0)
                ON CONFLICT (product_id, warehouse_id) DO NOTHING
            ''', (pid, warehouse_id))
            db.execute('''
                UPDATE warehouse_stock SET quantity = quantity + %s
                WHERE product_id = %s AND warehouse_id = %s
            ''', (qty, pid, warehouse_id))
        
        # حركة المخزون
        db.execute('''
            INSERT INTO inventory_movements (product_id, movement_type, quantity, reference_type, reference_id, warehouse_id, created_by)
            VALUES (%s, 'in', %s, 'supplier_invoice', %s, %s, %s)
        ''', (pid, qty, invoice_id, warehouse_id, session['user_id']))
    
    # تحديث حالة الفاتورة
    db.execute("UPDATE supplier_invoices SET status = 'received' WHERE id = %s", (invoice_id,))
    db.commit()
    return jsonify({'success': True, 'message': 'تم استلام البضاعة وتحديث المخزون'})

# ═══════════════════════════════════════════════════════════════
# المخزون
# ═══════════════════════════════════════════════════════════════

@app.route('/inventory', strict_slashes=False)
@role_required('manager', 'warehouse')
def inventory():
    db = get_db()
    warehouse_id = request.args.get('warehouse_id', type=int)
    
    products_list = db.execute('''
        SELECT p.*, c.name as category_name
        FROM products p
        LEFT JOIN categories c ON c.id = p.category_id
        WHERE p.is_active = TRUE
        ORDER BY p.name
    ''').fetchall()
    
    # جلب وحدات كل صنف
    products_with_units = []
    for p in products_list:
        product = dict(p)
        units = db.execute('''
            SELECT pu.*, u.name as unit_name, u.symbol
            FROM product_units pu
            JOIN units u ON u.id = pu.unit_id
            WHERE pu.product_id = %s
            ORDER BY pu.conversion_factor
        ''', (p['id'],)).fetchall()
        product['units'] = [dict(u) for u in units]
        
        # مخزون مخزن محدد
        if warehouse_id:
            ws = db.execute('''
                SELECT quantity FROM warehouse_stock WHERE product_id = %s AND warehouse_id = %s
            ''', (p['id'], warehouse_id)).fetchone()
            product['warehouse_qty'] = float(ws['quantity']) if ws else 0
        
        products_with_units.append(product)
    
    # حساب الإحصائيات
    stats = {
        'total': len(products_with_units),
        'above_min': sum(1 for p in products_with_units if p['current_stock'] > p['min_stock']),
        'below_min': sum(1 for p in products_with_units if p['current_stock'] <= p['min_stock'] and p['current_stock'] > 0),
        'zero_stock': sum(1 for p in products_with_units if p['current_stock'] == 0),
    }
    
    warehouses_list = db.execute("SELECT id, name, type FROM warehouses WHERE is_active = TRUE ORDER BY name").fetchall()
    return render_template('inventory.html', products=products_with_units, stats=stats,
                         warehouses=warehouses_list, selected_warehouse_id=warehouse_id)

@app.route('/api/inventory/movements/<int:product_id>')
@role_required('manager', 'warehouse')
def product_movements(product_id):
    db = get_db()
    movements = db.execute('''
        SELECT im.*, u.display_name as created_by_name, w.name as warehouse_name
        FROM inventory_movements im
        LEFT JOIN users u ON u.id = im.created_by
        LEFT JOIN warehouses w ON w.id = im.warehouse_id
        WHERE im.product_id = %s
        ORDER BY im.created_at DESC
        LIMIT 50
    ''', (product_id,)).fetchall()
    return jsonify([dict(m) for m in movements])

# ═══════════════════════════════════════════════════════════════
# التقارير
# ═══════════════════════════════════════════════════════════════

@app.route('/reports')
@role_required('manager')
def reports():
    return render_template('reports.html')

@app.route('/api/reports/purchases')
@role_required('manager')
def report_purchases():
    db = get_db()
    start = request.args.get('start', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    end = request.args.get('end', datetime.now().strftime('%Y-%m-%d'))
    
    data = db.execute('''
        SELECT s.name as supplier_name, SUM(si.total_amount) as total,
               COUNT(si.id) as invoice_count
        FROM supplier_invoices si
        JOIN suppliers s ON s.id = si.supplier_id
        WHERE si.invoice_date BETWEEN %s AND %s
        GROUP BY s.id
        ORDER BY total DESC
    ''', (start, end)).fetchall()
    
    return jsonify([dict(d) for d in data])

@app.route('/api/reports/supplier-debts')
@role_required('manager')
def report_supplier_debts():
    db = get_db()
    data = db.execute('''
        SELECT s.name, s.balance,
               SUM(si.total_amount) as total_invoices,
               SUM(si.paid_amount) as total_paid
        FROM suppliers s
        LEFT JOIN supplier_invoices si ON si.supplier_id = s.id
        WHERE s.is_active = TRUE
        GROUP BY s.id
        HAVING (SUM(si.total_amount) - SUM(si.paid_amount)) > 0
        ORDER BY (SUM(si.total_amount) - SUM(si.paid_amount)) DESC
    ''').fetchall()
    return jsonify([dict(d) for d in data])

# ═══════════════════════════════════════════════════════════════
# الإعدادات
# ═══════════════════════════════════════════════════════════════

@app.route('/settings')
@role_required('manager')
def settings():
    db = get_db()
    settings_list = db.execute('SELECT * FROM settings').fetchall()
    settings_dict = {s['key']: s['value'] for s in settings_list}
    return render_template('settings.html', settings=settings_dict)

@app.route('/api/settings', methods=['POST'])
@role_required('manager')
def api_settings():
    db = get_db()
    data = request.json
    
    for key, value in data.items():
        db.execute('UPDATE settings SET value = %s WHERE key = %s', (value, key))
    
    db.commit()
    return jsonify({'success': True, 'message': 'تم حفظ الإعدادات'})

@app.route('/api/backup')
@role_required('manager')
def api_backup():
    import shutil
    from datetime import datetime
    
    db_path = app.config['DATABASE']
    backup_name = f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
    backup_path = os.path.join(os.path.dirname(db_path), 'backups', backup_name)
    
    os.makedirs(os.path.dirname(backup_path), exist_ok=True)
    shutil.copy2(db_path, backup_path)
    
    return jsonify({'success': True, 'message': f'تم النسخ الاحتياطي: {backup_name}'})

# ═══════════════════════════════════════════════════════════════
# تشغيل التطبيق
# ═══════════════════════════════════════════════════════════════

def generate_ssl_cert(force=False):
    """إنشاء شهادة SSL ذاتية التوقيع مع SAN"""
    from OpenSSL import crypto
    
    cert_file = 'cert.pem'
    key_file = 'key.pem'
    
    if not force and os.path.exists(cert_file) and os.path.exists(key_file):
        return cert_file, key_file
    
    # إنشاء مفتاح
    key = crypto.PKey()
    key.generate_key(crypto.TYPE_RSA, 2048)
    
    # إنشاء شهادة
    cert = crypto.X509()
    cert.get_subject().C = "YE"
    cert.get_subject().ST = "Sanaa"
    cert.get_subject().L = "Sanaa"
    cert.get_subject().O = "Supermarket"
    cert.get_subject().OU = "IT"
    cert.get_subject().CN = "localhost"
    cert.set_serial_number(2000)
    cert.gmtime_adj_notBefore(0)
    cert.gmtime_adj_notAfter(365*24*60*60)
    cert.set_issuer(cert.get_subject())
    cert.set_pubkey(key)
    
    # إضافة SAN لدعم IP والميكروفون
    san = b"DNS:localhost,IP:127.0.0.1,IP:192.168.8.38"
    cert.add_extensions([
        crypto.X509Extension(b"subjectAltName", False, san),
        crypto.X509Extension(b"basicConstraints", True, b"CA:TRUE"),
    ])
    
    cert.sign(key, 'sha256')
    
    with open(cert_file, "wb") as f:
        f.write(crypto.dump_certificate(crypto.FILETYPE_PEM, cert))
    with open(key_file, "wb") as f:
        f.write(crypto.dump_privatekey(crypto.FILETYPE_PEM, key))
    
    return cert_file, key_file

if __name__ == '__main__':
    import sys
    
    # PostgreSQL: tables already created, skip init_db
    # with app.app_context():
    #     init_db()
    
    # التحقق من وضع HTTPS
    use_https = '--https' in sys.argv or os.environ.get('USE_HTTPS', '').lower() == 'true'
    host = os.environ.get('FLASK_RUN_HOST', '0.0.0.0')  # تغيير للسماح بالوصول الخارجي
    port = int(os.environ.get('FLASK_RUN_PORT', '5555'))
    
    print('=' * 50)
    print('Supermarket Management System')
    print('=' * 50)
    
    if use_https:
        cert_file, key_file = generate_ssl_cert()
        print(f'HTTPS Mode')
        print(f'URL: https://localhost:{port}')
        print(f'URL: https://<your-tailscale-ip>:{port}')
    else:
        print(f'URL: http://localhost:{port}')
    
    print('User: admin')
    print('Password: admin')
    print('=' * 50)
    print('For mobile + camera: python app.py --https')
    print('=' * 50)
    
    if use_https:
        app.run(
            debug=False,
            host=host,
            port=port,
            ssl_context=(cert_file, key_file),
            use_reloader=False,
            threaded=True
        )
    else:
        app.run(
            debug=app.config['DEBUG'],
            host=host,
            port=port,
            use_reloader=False,
            threaded=True
        )
