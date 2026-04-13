# -*- coding: utf-8 -*-
"""
تحديث قاعدة البيانات من ملف Excel:
1. تحديث سعر التكلفة من Sheet1
2. إضافة الوحدات الكبيرة من الجاهز وورقة1 إلى product_barcodes
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
import psycopg2
import psycopg2.extras

# === اتصال ===
conn = psycopg2.connect(host='localhost', database='supermarket', user='postgres', password='774424555')
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

# === تحميل بيانات DB ===
print("تحميل بيانات قاعدة البيانات...")

# الأصناف الفعّالة
cur.execute("SELECT id, name, barcode, product_code, cost_price, sell_price, unit FROM products WHERE is_active = TRUE")
db_products = cur.fetchall()

# الباركودات
cur.execute("SELECT id, product_id, barcode, unit, pack_size, cost_price, sell_price FROM product_barcodes")
db_barcodes = cur.fetchall()

# Indexes
by_barcode = {}
for p in db_products:
    if p['barcode']:
        by_barcode[p['barcode'].strip()] = p
for b in db_barcodes:
    if b['barcode']:
        by_barcode[b['barcode'].strip()] = b

by_code = {}
for p in db_products:
    if p['product_code']:
        by_code[p['product_code'].strip()] = p

by_name = {}
for p in db_products:
    n = p['name'].strip() if p['name'] else ''
    if n and n not in by_name:
        by_name[n] = p

print(f"  أصناف: {len(db_products)} | باركودات: {len(db_barcodes)}")
print(f"  فهرس باركود: {len(by_barcode)} | فهرس كود: {len(by_code)} | فهرس اسم: {len(by_name)}")

# === قراءة Excel ===
print("\nقراءة ملف Excel...")
path = r"C:\Users\qbas\Desktop\مجلد جديد\نسخة من اصناف_مواد_غذائية(1).xlsx"
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

def find_product(barcode, code, name):
    """البحث عن صنف بالباركود أولاً، ثم الكود، ثم الاسم"""
    if barcode and barcode in by_barcode:
        match = by_barcode[barcode]
        # قد يكون من products أو product_barcodes
        pid = match.get('product_id') or match.get('id')
        return pid
    if code and code in by_code:
        return by_code[code]['id']
    if name and name in by_name:
        return by_name[name]['id']
    return None

# ==============================
# المرحلة 1: تحديث التكلفة من Sheet1
# ==============================
print("\n" + "=" * 60)
print("المرحلة 1: تحديث سعر التكلفة من Sheet1")
print("=" * 60)

cost_updated = 0
sell_updated = 0
not_found_s1 = 0

ws1 = wb['Sheet1']
for row in ws1.iter_rows(min_row=2, values_only=True):
    if not row[4]:
        continue
    
    code = str(row[3]).strip() if row[3] else ''
    name = str(row[4]).strip()
    barcode = str(row[7]).strip() if row[7] and str(row[7]).strip() not in ('None', 'بدون', '') else ''
    cost = row[11] if row[11] and isinstance(row[11], (int, float)) and row[11] > 0 else 0
    sell = row[12] if row[12] and isinstance(row[12], (int, float)) and row[12] > 0 else 0
    
    pid = find_product(barcode, code, name)
    if not pid:
        not_found_s1 += 1
        continue
    
    # تحديث التكلفة إذا كانت فارغة
    if cost > 0:
        cur.execute("""
            UPDATE products SET cost_price = %s 
            WHERE id = %s AND (cost_price IS NULL OR cost_price = 0)
        """, (round(cost, 2), pid))
        if cur.rowcount > 0:
            cost_updated += 1
    
    # تحديث البيع إذا كان فارغ
    if sell > 0:
        cur.execute("""
            UPDATE products SET sell_price = %s 
            WHERE id = %s AND (sell_price IS NULL OR sell_price = 0)
        """, (round(sell, 2), pid))
        if cur.rowcount > 0:
            sell_updated += 1

conn.commit()
print(f"  تكلفة محدّثة: {cost_updated}")
print(f"  بيع محدّث: {sell_updated}")
print(f"  غير موجود: {not_found_s1}")

# ==============================
# المرحلة 2: إضافة وحدات من الجاهز
# ==============================
print("\n" + "=" * 60)
print("المرحلة 2: إضافة الوحدات من شيت 'الجاهز'")
print("=" * 60)

units_added_j = 0
units_skipped_j = 0
not_found_j = 0

# بناء فهرس الوحدات الموجودة لكل صنف
cur.execute("SELECT product_id, unit, barcode FROM product_barcodes")
existing_units = {}
for r in cur.fetchall():
    key = (r['product_id'], (r['unit'] or '').strip().lower())
    existing_units[key] = True
    if r['barcode']:
        existing_units[('bc', r['barcode'].strip())] = True

# أيضاً وحدة المنتج الأساسية
for p in db_products:
    key = (p['id'], (p['unit'] or 'حبه').strip().lower())
    existing_units[key] = True

ws2 = wb['الجاهز']
for row in ws2.iter_rows(min_row=2, values_only=True):
    if not row[2]:
        continue
    
    code = str(row[1]).strip() if row[1] else ''
    name = str(row[2]).strip()
    unit = str(row[3]).strip() if row[3] else 'حبه'
    barcode = str(row[4]).strip() if row[4] and str(row[4]).strip() not in ('None', 'بدون', '') else ''
    pack = int(row[5]) if row[5] and isinstance(row[5], (int, float)) else 1
    cost = round(row[8], 2) if row[8] and isinstance(row[8], (int, float)) and row[8] > 0 else 0
    sell = round(row[9], 2) if row[9] and isinstance(row[9], (int, float)) and row[9] > 0 else 0
    
    pid = find_product(barcode, code, name)
    if not pid:
        not_found_j += 1
        continue
    
    # تحقق: هل الوحدة موجودة بالفعل؟
    unit_key = (pid, unit.lower())
    bc_key = ('bc', barcode) if barcode else None
    
    if unit_key in existing_units or (bc_key and bc_key in existing_units):
        units_skipped_j += 1
        continue
    
    # إضافة وحدة جديدة
    try:
        cur.execute("""
            INSERT INTO product_barcodes (product_id, barcode, unit, pack_size, cost_price, sell_price)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (pid, barcode if barcode else None, unit, pack, cost if cost > 0 else None, sell if sell > 0 else None))
        units_added_j += 1
        existing_units[unit_key] = True
        if barcode:
            existing_units[('bc', barcode)] = True
    except Exception as e:
        conn.rollback()  # reset transaction on error

conn.commit()
print(f"  وحدات مضافة: {units_added_j}")
print(f"  موجودة مسبقاً: {units_skipped_j}")
print(f"  غير موجود: {not_found_j}")

# ==============================
# المرحلة 3: إضافة وحدات من ورقة1
# ==============================
print("\n" + "=" * 60)
print("المرحلة 3: إضافة الوحدات من شيت 'ورقة1'")
print("=" * 60)

units_added_w = 0
units_skipped_w = 0
not_found_w = 0

ws3 = wb['ورقة1']
for row in ws3.iter_rows(min_row=2, values_only=True):
    if not row[2]:
        continue
    
    code = str(row[1]).strip() if row[1] else ''
    name = str(row[2]).strip()
    unit = str(row[3]).strip() if row[3] else 'حبه'
    barcode = str(row[4]).strip() if row[4] and str(row[4]).strip() not in ('None', 'بدون', '') else ''
    pack = int(row[5]) if row[5] and isinstance(row[5], (int, float)) else 1
    cost = round(row[8], 2) if row[8] and isinstance(row[8], (int, float)) and row[8] > 0 else 0
    sell = round(row[9], 2) if row[9] and isinstance(row[9], (int, float)) and row[9] > 0 else 0
    
    pid = find_product(barcode, code, name)
    if not pid:
        not_found_w += 1
        continue
    
    unit_key = (pid, unit.lower())
    bc_key = ('bc', barcode) if barcode else None
    
    if unit_key in existing_units or (bc_key and bc_key in existing_units):
        units_skipped_w += 1
        continue
    
    try:
        cur.execute("""
            INSERT INTO product_barcodes (product_id, barcode, unit, pack_size, cost_price, sell_price)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (pid, barcode if barcode else None, unit, pack, cost if cost > 0 else None, sell if sell > 0 else None))
        units_added_w += 1
        existing_units[unit_key] = True
        if barcode:
            existing_units[('bc', barcode)] = True
    except:
        conn.rollback()

conn.commit()
print(f"  وحدات مضافة: {units_added_w}")
print(f"  موجودة مسبقاً: {units_skipped_w}")
print(f"  غير موجود: {not_found_w}")

wb.close()

# ==============================
# النتيجة النهائية
# ==============================
print("\n" + "=" * 60)
print("النتيجة النهائية")
print("=" * 60)

cur.execute("SELECT COUNT(*) as c FROM products WHERE is_active=TRUE AND cost_price IS NOT NULL AND cost_price > 0")
new_cost = cur.fetchone()['c']
cur.execute("SELECT COUNT(*) as c FROM products WHERE is_active=TRUE AND sell_price IS NOT NULL AND sell_price > 0")
new_sell = cur.fetchone()['c']
cur.execute("SELECT COUNT(*) as c FROM products WHERE is_active=TRUE")
total_active = cur.fetchone()['c']
cur.execute("SELECT COUNT(*) as c FROM product_barcodes")
total_bc = cur.fetchone()['c']

print(f"أصناف فعّالة: {total_active}")
print(f"بسعر تكلفة: {new_cost} / {total_active} (كان 116)")
print(f"بسعر بيع: {new_sell} / {total_active} (كان 4112)")
print(f"باركودات/وحدات: {total_bc} (كان 3813)")
print(f"\nتكلفة محدّثة: {cost_updated}")
print(f"وحدات مضافة: {units_added_j + units_added_w}")

conn.close()
