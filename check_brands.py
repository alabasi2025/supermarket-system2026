# -*- coding: utf-8 -*-
"""
البحث عن بيانات العائلات/العلامات التجارية في الملف
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

path = r"C:\Users\qbas\Desktop\مجلد جديد\نسخة من اصناف_مواد_غذائية(1).xlsx"
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

# === Sheet1: فحص كل الأعمدة ===
print("=" * 70)
print("Sheet1 — فحص كل الأعمدة (أول 3 صفوف كاملة)")
print("=" * 70)
ws1 = wb['Sheet1']
for ri, row in enumerate(ws1.iter_rows(min_row=1, max_row=4)):
    label = "HEADER" if ri == 0 else f"Row {ri+1}"
    print(f"\n{label}:")
    for ci, cell in enumerate(row):
        if cell.value is not None:
            print(f"  Col {chr(65+ci) if ci<26 else str(ci+1)} ({ci+1}): {str(cell.value)[:60]}")

# === الجاهز: فحص كل الأعمدة ===
print(f"\n{'='*70}")
print("الجاهز — فحص كل الأعمدة (أول 3 صفوف كاملة)")
print("=" * 70)
ws2 = wb['الجاهز']
for ri, row in enumerate(ws2.iter_rows(min_row=1, max_row=4)):
    label = "HEADER" if ri == 0 else f"Row {ri+1}"
    print(f"\n{label}:")
    for ci, cell in enumerate(row):
        if cell.value is not None:
            print(f"  Col {chr(65+ci) if ci<26 else str(ci+1)} ({ci+1}): {str(cell.value)[:60]}")

# === ورقة1 ===
print(f"\n{'='*70}")
print("ورقة1 — فحص كل الأعمدة (أول 3 صفوف)")
print("=" * 70)
ws3 = wb['ورقة1']
for ri, row in enumerate(ws3.iter_rows(min_row=1, max_row=4)):
    label = "HEADER" if ri == 0 else f"Row {ri+1}"
    print(f"\n{label}:")
    for ci, cell in enumerate(row):
        if cell.value is not None:
            print(f"  Col {chr(65+ci) if ci<26 else str(ci+1)} ({ci+1}): {str(cell.value)[:60]}")

wb.close()

# === فحص الملف الكبير الأصلي (من data/) ===
print(f"\n{'='*70}")
print("الملف الكبير الأصلي (اصناف سوبر ماركت) — هل فيه عمود مورد/عائلة؟")
print("=" * 70)

import glob, os
data_dir = r"D:\supermarket-system\data"
files = sorted(glob.glob(os.path.join(data_dir, "*.xlsx")), key=lambda f: os.path.getsize(f), reverse=True)
f = files[0]
print(f"ملف: {os.path.basename(f)} ({os.path.getsize(f)//1024} KB)")

wb2 = openpyxl.load_workbook(f, read_only=True)
ws = wb2[wb2.sheetnames[0]]
for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=4)):
    label = "HEADER" if ri == 0 else f"Row {ri+1}"
    print(f"\n{label}:")
    for ci, cell in enumerate(row):
        if cell.value is not None:
            print(f"  Col {chr(65+ci) if ci<26 else str(ci+1)} ({ci+1}): {str(cell.value)[:60]}")

# عدد الموردين الفريدين
print("\n--- الموردين في الملف الكبير ---")
suppliers = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[3]:
        continue
    s = str(row[5]).strip() if len(row) > 5 and row[5] else ''
    if s:
        suppliers[s] = suppliers.get(s, 0) + 1

print(f"عدد الموردين: {len(suppliers)}")
for s, c in sorted(suppliers.items(), key=lambda x: x[1], reverse=True)[:25]:
    print(f"  {s}: {c} صنف")

wb2.close()

# === فحص DB الحالي ===
print(f"\n{'='*70}")
print("قاعدة البيانات — حالة العائلات (brand)")
print("=" * 70)

import psycopg2, psycopg2.extras
conn = psycopg2.connect(host='localhost', database='supermarket', user='postgres', password='774424555')
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

cur.execute("SELECT COUNT(*) as c FROM products WHERE is_active=TRUE AND brand IS NOT NULL AND brand != ''")
with_brand = cur.fetchone()['c']
cur.execute("SELECT COUNT(*) as c FROM products WHERE is_active=TRUE AND (brand IS NULL OR brand = '')")
no_brand = cur.fetchone()['c']
cur.execute("SELECT COUNT(DISTINCT brand) as c FROM products WHERE is_active=TRUE AND brand IS NOT NULL AND brand != ''")
unique_brands = cur.fetchone()['c']

print(f"بعائلة: {with_brand}")
print(f"بدون عائلة: {no_brand}")
print(f"عائلات فريدة: {unique_brands}")

if unique_brands > 0:
    cur.execute("""
        SELECT brand, COUNT(*) as cnt FROM products 
        WHERE is_active=TRUE AND brand IS NOT NULL AND brand != ''
        GROUP BY brand ORDER BY cnt DESC LIMIT 20
    """)
    print("\nأكثر العائلات:")
    for r in cur.fetchall():
        print(f"  {r['brand']}: {r['cnt']}")

conn.close()
