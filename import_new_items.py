# -*- coding: utf-8 -*-
import openpyxl, sys, psycopg2, psycopg2.extras
from decimal import Decimal
sys.stdout.reconfigure(encoding='utf-8')

conn = psycopg2.connect(host='localhost', database='supermarket', user='postgres', password='774424555')
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

# 1) قراءة الملف
wb = openpyxl.load_workbook('D:/supermarket-system/web/new_items.xlsx', data_only=True)
ws = wb.active

items = {}
for row in range(2, ws.max_row + 1):
    code = str(ws.cell(row=row, column=1).value or '').strip()
    name = str(ws.cell(row=row, column=2).value or '').strip()
    unit = str(ws.cell(row=row, column=3).value or '').strip()
    barcode = str(ws.cell(row=row, column=4).value or '').strip()
    pack = str(ws.cell(row=row, column=5).value or '1').strip()
    
    if not code or barcode == 'None':
        barcode = ''
    
    if code not in items:
        items[code] = {'name': name, 'units': []}
    items[code]['units'].append({'unit': unit, 'barcode': barcode, 'pack': int(float(pack)) if pack else 1})

# تجاهل أصناف بدون أي باركود
items_with_bc = {}
for code, item in items.items():
    has_bc = any(u['barcode'] for u in item['units'])
    if has_bc:
        items_with_bc[code] = item

print(f"أصناف في الملف (لها باركود): {len(items_with_bc)}")

# 2) جلب باركودات النظام
cur.execute("SELECT id, barcode FROM products WHERE is_active = TRUE AND barcode IS NOT NULL AND barcode != ''")
db_products_by_bc = {}
for r in cur.fetchall():
    db_products_by_bc[r['barcode']] = r['id']
    # أضف بالصفر البادئ
    if r['barcode'].startswith('0'):
        db_products_by_bc[r['barcode'].lstrip('0')] = r['id']
    else:
        db_products_by_bc['0' + r['barcode']] = r['id']

cur.execute("SELECT product_id, barcode FROM product_barcodes WHERE barcode IS NOT NULL AND barcode != ''")
for r in cur.fetchall():
    db_products_by_bc[r['barcode']] = r['product_id']
    if r['barcode'].startswith('0'):
        db_products_by_bc[r['barcode'].lstrip('0')] = r['product_id']
    else:
        db_products_by_bc['0' + r['barcode']] = r['product_id']

# 3) جلب الوحدات الموجودة لكل صنف
cur.execute("SELECT product_id, unit FROM product_barcodes")
existing_units = {}
for r in cur.fetchall():
    if r['product_id'] not in existing_units:
        existing_units[r['product_id']] = set()
    existing_units[r['product_id']].add(r['unit'])

# 4) تصنيف: جديد أو تحديث
new_products = []
update_products = []

for code, item in items_with_bc.items():
    # ابحث عن أي باركود متطابق
    matched_product_id = None
    for u in item['units']:
        if u['barcode'] and u['barcode'] in db_products_by_bc:
            matched_product_id = db_products_by_bc[u['barcode']]
            break
    
    if matched_product_id:
        update_products.append({'product_id': matched_product_id, 'code': code, 'item': item})
    else:
        new_products.append({'code': code, 'item': item})

print(f"أصناف جديدة: {len(new_products)}")
print(f"أصناف موجودة (تحديث وحدات): {len(update_products)}")

# 5) إضافة الأصناف الجديدة
added = 0
for np in new_products:
    item = np['item']
    # الوحدة الأساسية = أول وحدة لها باركود أو أول وحدة
    base_unit = item['units'][0]
    base_barcode = ''
    for u in item['units']:
        if u['barcode']:
            base_unit = u
            base_barcode = u['barcode']
            break
    
    try:
        cur.execute("""
            INSERT INTO products (product_code, name, barcode, unit, is_active)
            VALUES (%s, %s, %s, %s, TRUE)
            RETURNING id
        """, (np['code'], item['name'], base_barcode, base_unit['unit']))
        product_id = cur.fetchone()['id']
        
        # إضافة الوحدات الإضافية
        cumulative = 1
        for i, u in enumerate(item['units']):
            if i == 0:
                continue  # الأساسية محفوظة في products
            pack = u['pack'] or 1
            cumulative *= pack
            cur.execute("""
                INSERT INTO product_barcodes (product_id, barcode, unit, pack_size, conversion_factor, sort_order)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (product_id, u['barcode'] or '', u['unit'], pack, cumulative, i + 1))
        
        added += 1
    except Exception as e:
        conn.rollback()
        print(f"  خطأ إضافة {np['code']}: {e}")
        conn = psycopg2.connect(host='localhost', database='supermarket', user='postgres', password='774424555')
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

conn.commit()
print(f"\n✅ أضفت {added} صنف جديد")

# 6) تحديث الوحدات الناقصة
units_added = 0
for up in update_products:
    pid = up['product_id']
    item = up['item']
    current_units = existing_units.get(pid, set())
    
    cumulative = 1
    for i, u in enumerate(item['units']):
        if i == 0:
            continue
        pack = u['pack'] or 1
        cumulative *= pack
        
        # هل الوحدة موجودة؟
        if u['unit'] not in current_units:
            try:
                cur.execute("""
                    INSERT INTO product_barcodes (product_id, barcode, unit, pack_size, conversion_factor, sort_order)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """, (pid, u['barcode'] or '', u['unit'], pack, cumulative, i + 1))
                units_added += 1
            except Exception as e:
                conn.rollback()
                conn = psycopg2.connect(host='localhost', database='supermarket', user='postgres', password='774424555')
                cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

conn.commit()
print(f"✅ أضفت {units_added} وحدة ناقصة للأصناف الموجودة")

# 7) إحصائيات نهائية
cur.execute("SELECT COUNT(*) as c FROM products WHERE is_active = TRUE")
print(f"\n=== الإحصائيات النهائية ===")
print(f"  إجمالي الأصناف الفعّالة: {cur.fetchone()['c']}")
cur.execute("SELECT COUNT(*) as c FROM product_barcodes")
print(f"  إجمالي الباركودات/الوحدات: {cur.fetchone()['c']}")

conn.close()
print("\n✅ تم!")
