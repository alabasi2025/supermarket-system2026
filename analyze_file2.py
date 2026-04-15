# -*- coding: utf-8 -*-
import openpyxl, sys, psycopg2
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('D:/supermarket-system/web/new_items.xlsx', data_only=True)
ws = wb.active

print(f"الملف: {ws.max_row - 1} سطر × {ws.max_column} أعمدة")
print()

# الأعمدة:
# 1 = كود الصنف
# 2 = اسم الصنف
# 3 = الوحدة
# 4 = الباركود
# 5 = العبوة/التحويل
# 6 = ؟

print("=== هيكل الملف ===")
print("  Col 1: كود الصنف")
print("  Col 2: اسم الصنف")
print("  Col 3: الوحدة")
print("  Col 4: الباركود")
print("  Col 5: العبوة/التحويل")
print("  Col 6: غير واضح")

# تحليل البيانات
items = {}  # code -> {name, units: [{unit, barcode, pack}]}
for row in range(2, ws.max_row + 1):
    code = str(ws.cell(row=row, column=1).value or '').strip()
    name = str(ws.cell(row=row, column=2).value or '').strip()
    unit = str(ws.cell(row=row, column=3).value or '').strip()
    barcode = str(ws.cell(row=row, column=4).value or '').strip()
    pack = str(ws.cell(row=row, column=5).value or '1').strip()
    col6 = str(ws.cell(row=row, column=6).value or '').strip()
    
    if not code:
        continue
    
    if code not in items:
        items[code] = {'name': name, 'units': []}
    items[code]['units'].append({
        'unit': unit, 'barcode': barcode if barcode != 'None' else '',
        'pack': pack, 'col6': col6
    })

print(f"\n=== إحصائيات الملف ===")
print(f"  أصناف فريدة: {len(items)}")

# عدد الوحدات لكل صنف
multi = sum(1 for v in items.values() if len(v['units']) > 1)
single = sum(1 for v in items.values() if len(v['units']) == 1)
print(f"  أصناف بوحدة واحدة: {single}")
print(f"  أصناف بوحدات متعددة: {multi}")

# الوحدات المستخدمة
all_units = set()
for v in items.values():
    for u in v['units']:
        if u['unit']:
            all_units.add(u['unit'])
print(f"  الوحدات المستخدمة: {sorted(all_units)}")

# باركودات
barcodes_in_file = set()
for v in items.values():
    for u in v['units']:
        if u['barcode']:
            barcodes_in_file.add(u['barcode'])
print(f"  باركودات فريدة: {len(barcodes_in_file)}")

# مقارنة مع قاعدة البيانات
conn = psycopg2.connect(host='localhost', database='supermarket', user='postgres', password='774424555')
cur = conn.cursor()

# باركودات النظام
cur.execute("SELECT barcode FROM products WHERE is_active = TRUE AND barcode IS NOT NULL AND barcode != ''")
db_barcodes = set(r[0] for r in cur.fetchall())
cur.execute("SELECT barcode FROM product_barcodes WHERE barcode IS NOT NULL AND barcode != ''")
db_barcodes.update(r[0] for r in cur.fetchall())

# توسيع بالأصفار
db_expanded = set()
for bc in db_barcodes:
    db_expanded.add(bc)
    if bc.startswith('0'):
        db_expanded.add(bc.lstrip('0'))
    else:
        db_expanded.add('0' + bc)

# أكواد الأصناف في النظام
cur.execute("SELECT product_code FROM products WHERE is_active = TRUE AND product_code IS NOT NULL")
db_codes = set(r[0] for r in cur.fetchall())

# مقارنة بالباركود
found_bc = barcodes_in_file & db_expanded
not_found_bc = barcodes_in_file - db_expanded

# مقارنة بالكود
file_codes = set(items.keys())
found_code = file_codes & db_codes
not_found_code = file_codes - db_codes

print(f"\n=== المقارنة بالباركود ===")
print(f"  موجود في النظام: {len(found_bc)}")
print(f"  غير موجود: {len(not_found_bc)}")

print(f"\n=== المقارنة بكود الصنف ===")
print(f"  موجود في النظام: {len(found_code)}")
print(f"  غير موجود: {len(not_found_code)}")

# مثال على أصناف جديدة
print(f"\n=== أمثلة أصناف جديدة (أول 15) ===")
count = 0
for code in sorted(not_found_code):
    if count >= 15:
        break
    item = items[code]
    units_str = ' + '.join([f"{u['unit']}({u['barcode'] or '-'})" for u in item['units']])
    print(f"  {code} | {item['name']} | {units_str}")
    count += 1

conn.close()
