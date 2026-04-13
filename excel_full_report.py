# -*- coding: utf-8 -*-
"""
تقرير كامل عن محتوى ملف نسخة من اصناف_مواد_غذائية(1).xlsx
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

path = r"C:\Users\qbas\Desktop\مجلد جديد\نسخة من اصناف_مواد_غذائية(1).xlsx"
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

print("=" * 70)
print("محتوى ملف: نسخة من اصناف_مواد_غذائية(1).xlsx")
print("=" * 70)
print(f"عدد الشيتات: {len(wb.sheetnames)} → {wb.sheetnames}")

# ======== Sheet1 ========
print(f"\n{'='*70}")
print("شيت 1: Sheet1")
print("=" * 70)
ws1 = wb['Sheet1']
print("الأعمدة:")
print("  A: رقم المجموعة (القسم)")
print("  B: رقم الصنف (كود قديم)")
print("  C: (فارغ)")
print("  D: رقم جديد (كود جديد)")
print("  E: اسم الصنف")
print("  F: الوحده")
print("  G: الجرد (كمية المخزون)")
print("  H: رقم الباركود")
print("  I: العبوه (كمية في العبوة)")
print("  J: وحدة رئيسية")
print("  K: نوع الصنف")
print("  L: سعر التكلفة تقديري")
print("  M: سعر البيع")

# إحصائيات مفصلة
total = 0
groups = {}
units = {}
has_barcode = 0
has_cost = 0
has_sell = 0
has_stock = 0

for row in ws1.iter_rows(min_row=2, values_only=True):
    if not row[4]:
        continue
    total += 1
    
    g = str(row[0]) if row[0] else '?'
    groups[g] = groups.get(g, 0) + 1
    
    u = str(row[5]).strip() if row[5] else 'حبه'
    units[u] = units.get(u, 0) + 1
    
    if row[7] and str(row[7]).strip() not in ('', 'None', 'بدون'):
        has_barcode += 1
    if row[11] and isinstance(row[11], (int, float)) and row[11] > 0:
        has_cost += 1
    if row[12] and isinstance(row[12], (int, float)) and row[12] > 0:
        has_sell += 1
    if row[6] and isinstance(row[6], (int, float)) and row[6] > 0:
        has_stock += 1

print(f"\nالإجمالي: {total} صنف")
print(f"بباركود: {has_barcode}")
print(f"بسعر تكلفة: {has_cost}")
print(f"بسعر بيع: {has_sell}")
print(f"بجرد/مخزون: {has_stock}")
print(f"\nالأقسام ({len(groups)}):")
for g in sorted(groups.keys()):
    print(f"  {g}: {groups[g]} صنف")
print(f"\nالوحدات:")
for u, c in sorted(units.items(), key=lambda x: x[1], reverse=True):
    print(f"  {u}: {c}")

print(f"\nأول 10 أصناف:")
count = 0
for row in ws1.iter_rows(min_row=2, values_only=True):
    if not row[4] or count >= 10:
        break
    count += 1
    name = str(row[4])[:40]
    unit = row[5] or 'حبه'
    bc = row[7] or '-'
    cost = row[11] if isinstance(row[11], (int, float)) else '-'
    sell = row[12] if isinstance(row[12], (int, float)) else '-'
    stock = row[6] if isinstance(row[6], (int, float)) else '-'
    print(f"  {count}. {name:<42} | {unit:<6} | bc={str(bc)[:13]:<13} | تكلفة={str(cost):<10} | بيع={str(sell):<10} | جرد={stock}")

# ======== الجاهز ========
print(f"\n{'='*70}")
print("شيت 2: الجاهز")
print("=" * 70)
ws2 = wb['الجاهز']
print("الأعمدة:")
print("  A: رقم المجموعة")
print("  B: رقم الصنف")
print("  C: اسم الصنف")
print("  D: الوحده")
print("  E: رقم الباركود")
print("  F: العبوه")
print("  G: وحدة رئيسية")
print("  H: نوع الصنف")
print("  I: سعر التكلفة تقديري")
print("  J: سعر البيع")
print("  K: نفيثنث (عدد الوحدات)")

total2 = 0
units2 = {}
has_bc2 = 0
has_cost2 = 0
has_sell2 = 0

for row in ws2.iter_rows(min_row=2, values_only=True):
    if not row[2]:
        continue
    total2 += 1
    u = str(row[3]).strip() if row[3] else 'حبه'
    units2[u] = units2.get(u, 0) + 1
    if row[4] and str(row[4]).strip() not in ('', 'None', 'بدون'):
        has_bc2 += 1
    if row[8] and isinstance(row[8], (int, float)) and row[8] > 0:
        has_cost2 += 1
    if row[9] and isinstance(row[9], (int, float)) and row[9] > 0:
        has_sell2 += 1

print(f"\nالإجمالي: {total2} صف")
print(f"بباركود: {has_bc2}")
print(f"بسعر تكلفة: {has_cost2}")
print(f"بسعر بيع: {has_sell2}")
print(f"\nالوحدات:")
for u, c in sorted(units2.items(), key=lambda x: x[1], reverse=True):
    print(f"  {u}: {c}")

# أمثلة: نفس الصنف بوحدات مختلفة
print(f"\nأمثلة أصناف بوحدات مختلفة:")
items_by_code = {}
for row in ws2.iter_rows(min_row=2, values_only=True):
    if not row[2]:
        continue
    code = str(row[1]).strip() if row[1] else ''
    items_by_code.setdefault(code, []).append(row)

shown = 0
for code, rows in items_by_code.items():
    all_units = set(str(r[3]).strip() for r in rows if r[3])
    if len(all_units) > 1 and shown < 5:
        shown += 1
        name = str(rows[0][2])[:45]
        print(f"\n  📦 {name} (كود: {code})")
        for r in rows:
            u = r[3] or '-'
            bc = r[4] or 'بدون'
            pack = r[5] or '-'
            cost = r[8] if isinstance(r[8], (int, float)) else '-'
            sell = r[9] if isinstance(r[9], (int, float)) else '-'
            print(f"     → {u:<8} | عبوة={str(pack):<4} | تكلفة={str(cost):<12} | بيع={str(sell):<8} | bc={str(bc)[:15]}")

# ======== ورقة1 ========
print(f"\n{'='*70}")
print("شيت 3: ورقة1")
print("=" * 70)
ws3 = wb['ورقة1']
print("نفس أعمدة 'الجاهز'")

total3 = 0
units3 = {}
for row in ws3.iter_rows(min_row=2, values_only=True):
    if not row[2]:
        continue
    total3 += 1
    u = str(row[3]).strip() if row[3] else 'حبه'
    units3[u] = units3.get(u, 0) + 1

print(f"\nالإجمالي: {total3} صف")
print(f"\nالوحدات:")
for u, c in sorted(units3.items(), key=lambda x: x[1], reverse=True):
    print(f"  {u}: {c}")

# أمثلة
print(f"\nأمثلة أصناف بوحدات متعددة:")
items3 = {}
for row in ws3.iter_rows(min_row=2, values_only=True):
    if not row[2]:
        continue
    code = str(row[1]).strip() if row[1] else ''
    items3.setdefault(code, []).append(row)

shown3 = 0
for code, rows in items3.items():
    if len(rows) > 1 and shown3 < 5:
        shown3 += 1
        name = str(rows[0][2])[:45]
        print(f"\n  📦 {name} (كود: {code})")
        for r in rows:
            u = r[3] or '-'
            bc = r[4] or 'بدون'
            pack = r[5] or '-'
            cost = r[8] if isinstance(r[8], (int, float)) else '-'
            sell = r[9] if isinstance(r[9], (int, float)) else '-'
            print(f"     → {u:<8} | عبوة={str(pack):<4} | تكلفة={str(cost):<12} | بيع={str(sell):<8} | bc={str(bc)[:15]}")

wb.close()

# ملخص
print(f"\n{'='*70}")
print("الملخص النهائي")
print("=" * 70)
print(f"Sheet1:  {total} صنف  — الأصناف الأساسية (حبه غالباً) + باركود + تكلفة + بيع + جرد")
print(f"الجاهز:  {total2} صف   — نفس الأصناف + الوحدات الكبيرة (كرتون/باكت/شده/عروسه)")
print(f"ورقة1:   {total3} صف   — وحدات إضافية (باكت/كرتون/شده)")
print(f"\nالبيانات في الملف:")
print(f"  ✅ اسم الصنف")
print(f"  ✅ رقم الصنف (كود قديم + جديد)")
print(f"  ✅ القسم (19 قسم)")
print(f"  ✅ الوحدة (حبه/كرتون/باكت/شده/عروسه/كيس/علبه...)")
print(f"  ✅ الباركود")
print(f"  ✅ حجم العبوة")
print(f"  ✅ سعر التكلفة")
print(f"  ✅ سعر البيع")
print(f"  ✅ الجرد/المخزون (في Sheet1 فقط)")
