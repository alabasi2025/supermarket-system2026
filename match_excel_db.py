# -*- coding: utf-8 -*-
"""
مطابقة بيانات Excel مع قاعدة PostgreSQL
"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
import psycopg2
import psycopg2.extras

# === اتصال بقاعدة البيانات ===
conn = psycopg2.connect(
    host='localhost', database='supermarket',
    user='postgres', password='774424555'
)
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

# === قراءة البيانات من DB ===
print("=" * 70)
print("قراءة البيانات من PostgreSQL...")
print("=" * 70)

# الأصناف الفعّالة
cur.execute("SELECT COUNT(*) as cnt FROM products WHERE is_active = TRUE")
active = cur.fetchone()['cnt']
cur.execute("SELECT COUNT(*) as cnt FROM products WHERE is_active = FALSE")
inactive = cur.fetchone()['cnt']
print(f"أصناف فعّالة: {active}")
print(f"أصناف معطّلة: {inactive}")

# الباركودات
cur.execute("SELECT COUNT(*) as cnt FROM product_barcodes")
barcodes_count = cur.fetchone()['cnt']
print(f"باركودات: {barcodes_count}")

# أصناف بأسعار
cur.execute("SELECT COUNT(*) as cnt FROM products WHERE is_active = TRUE AND sell_price IS NOT NULL AND sell_price > 0")
with_sell = cur.fetchone()['cnt']
cur.execute("SELECT COUNT(*) as cnt FROM products WHERE is_active = TRUE AND cost_price IS NOT NULL AND cost_price > 0")
with_cost = cur.fetchone()['cnt']
cur.execute("SELECT COUNT(*) as cnt FROM products WHERE is_active = TRUE AND (sell_price IS NULL OR sell_price = 0)")
no_sell = cur.fetchone()['cnt']
print(f"أصناف بسعر بيع: {with_sell}")
print(f"أصناف بسعر تكلفة: {with_cost}")
print(f"أصناف بدون سعر بيع: {no_sell}")

# تحميل كل الأصناف الفعّالة
cur.execute("""
    SELECT id, name, barcode, product_code, cost_price, sell_price, unit, category_id
    FROM products WHERE is_active = TRUE
""")
db_products = cur.fetchall()

# تحميل كل الباركودات
cur.execute("SELECT * FROM product_barcodes")
db_barcodes = cur.fetchall()

# بناء index
db_by_name = {}
for p in db_products:
    name = p['name'].strip() if p['name'] else ''
    if name not in db_by_name:
        db_by_name[name] = []
    db_by_name[name].append(p)

db_by_barcode = {}
for p in db_products:
    if p['barcode']:
        db_by_barcode[p['barcode'].strip()] = p
for b in db_barcodes:
    if b['barcode']:
        db_by_barcode[b['barcode'].strip()] = b

db_by_code = {}
for p in db_products:
    if p['product_code']:
        db_by_code[p['product_code'].strip()] = p

print(f"\nفهرس الأسماء: {len(db_by_name)} اسم فريد")
print(f"فهرس الباركودات: {len(db_by_barcode)}")
print(f"فهرس الأكواد: {len(db_by_code)}")

# === قراءة Excel ===
print("\n" + "=" * 70)
print("قراءة ملف Excel...")
print("=" * 70)

excel_path = r"C:\Users\qbas\Desktop\مجلد جديد\نسخة من اصناف_مواد_غذائية(1).xlsx"
wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

all_excel_items = []

# Sheet1: headers in row 1
ws1 = wb['Sheet1']
for row in ws1.iter_rows(min_row=2, values_only=True):
    if not row[4]:  # اسم الصنف (col E)
        continue
    item = {
        'group': str(row[0]).strip() if row[0] else '',
        'old_code': str(row[1]).strip() if row[1] else '',
        'new_code': str(row[3]).strip() if row[3] else '',
        'name': str(row[4]).strip() if row[4] else '',
        'unit': str(row[5]).strip() if row[5] else 'حبه',
        'barcode': str(row[7]).strip() if row[7] else '',
        'pack_size': row[8] if row[8] else 1,
        'cost_price': row[11] if row[11] and isinstance(row[11], (int, float)) else 0,
        'sell_price': row[12] if row[12] and isinstance(row[12], (int, float)) else 0,
        'sheet': 'Sheet1'
    }
    # Clean barcode
    if item['barcode'] in ('None', '', 'بدون'):
        item['barcode'] = ''
    all_excel_items.append(item)

# الجاهز: different column layout
ws2 = wb['الجاهز']
for row in ws2.iter_rows(min_row=2, values_only=True):
    if not row[2]:  # اسم الصنف (col C)
        continue
    item = {
        'group': str(row[0]).strip() if row[0] else '',
        'old_code': '',
        'new_code': str(row[1]).strip() if row[1] else '',
        'name': str(row[2]).strip() if row[2] else '',
        'unit': str(row[3]).strip() if row[3] else 'حبه',
        'barcode': str(row[4]).strip() if row[4] else '',
        'pack_size': row[5] if row[5] else 1,
        'cost_price': row[8] if row[8] and isinstance(row[8], (int, float)) else 0,
        'sell_price': row[9] if row[9] and isinstance(row[9], (int, float)) else 0,
        'sheet': 'الجاهز'
    }
    if item['barcode'] in ('None', '', 'بدون'):
        item['barcode'] = ''
    all_excel_items.append(item)

# ورقة1
ws3 = wb['ورقة1']
for row in ws3.iter_rows(min_row=2, values_only=True):
    if not row[2]:
        continue
    item = {
        'group': str(row[0]).strip() if row[0] else '',
        'old_code': '',
        'new_code': str(row[1]).strip() if row[1] else '',
        'name': str(row[2]).strip() if row[2] else '',
        'unit': str(row[3]).strip() if row[3] else 'حبه',
        'barcode': str(row[4]).strip() if row[4] else '',
        'pack_size': row[5] if row[5] else 1,
        'cost_price': row[8] if row[8] and isinstance(row[8], (int, float)) else 0,
        'sell_price': row[9] if row[9] and isinstance(row[9], (int, float)) else 0,
        'sheet': 'ورقة1'
    }
    if item['barcode'] in ('None', '', 'بدون'):
        item['barcode'] = ''
    all_excel_items.append(item)

wb.close()

print(f"إجمالي أصناف Excel: {len(all_excel_items)}")
print(f"  Sheet1: {sum(1 for i in all_excel_items if i['sheet']=='Sheet1')}")
print(f"  الجاهز: {sum(1 for i in all_excel_items if i['sheet']=='الجاهز')}")
print(f"  ورقة1: {sum(1 for i in all_excel_items if i['sheet']=='ورقة1')}")

# أصناف بأسعار
with_price = sum(1 for i in all_excel_items if i['sell_price'] and i['sell_price'] > 0)
print(f"أصناف Excel بسعر بيع: {with_price}")

# === المطابقة ===
print("\n" + "=" * 70)
print("المطابقة...")
print("=" * 70)

matched_by_barcode = 0
matched_by_code = 0
matched_by_name = 0
not_matched = 0
matched_items = []
unmatched_items = []

for item in all_excel_items:
    match = None
    match_type = ''
    
    # 1. مطابقة بالباركود
    if item['barcode'] and item['barcode'] in db_by_barcode:
        match = db_by_barcode[item['barcode']]
        match_type = 'barcode'
        matched_by_barcode += 1
    
    # 2. مطابقة بالكود
    elif item['new_code'] and item['new_code'] in db_by_code:
        match = db_by_code[item['new_code']]
        match_type = 'code'
        matched_by_code += 1
    
    # 3. مطابقة بالاسم (exact)
    elif item['name'] in db_by_name:
        match = db_by_name[item['name']][0]
        match_type = 'name'
        matched_by_name += 1
    
    else:
        not_matched += 1
        unmatched_items.append(item)
        continue
    
    matched_items.append({
        'excel': item,
        'db': match,
        'match_type': match_type
    })

print(f"\n--- نتائج المطابقة ---")
print(f"مطابق بالباركود: {matched_by_barcode}")
print(f"مطابق بالكود: {matched_by_code}")
print(f"مطابق بالاسم: {matched_by_name}")
print(f"إجمالي مطابق: {matched_by_barcode + matched_by_code + matched_by_name}")
print(f"غير مطابق: {not_matched}")

# === تحليل التحديثات الممكنة ===
print("\n" + "=" * 70)
print("تحليل التحديثات الممكنة...")
print("=" * 70)

can_update_sell = 0
can_update_cost = 0
can_update_both = 0
can_add_unit = 0
price_differs = 0

for m in matched_items:
    excel = m['excel']
    db = m['db']
    
    db_id = db.get('product_id') or db.get('id')
    db_sell = db.get('sell_price') or 0
    db_cost = db.get('cost_price') or 0
    
    has_new_sell = excel['sell_price'] and excel['sell_price'] > 0
    has_new_cost = excel['cost_price'] and excel['cost_price'] > 0
    
    if has_new_sell and (not db_sell or db_sell == 0):
        can_update_sell += 1
    if has_new_cost and (not db_cost or db_cost == 0):
        can_update_cost += 1
    if has_new_sell and has_new_cost and (not db_sell or db_sell == 0) and (not db_cost or db_cost == 0):
        can_update_both += 1
    
    # سعر مختلف
    if has_new_sell and db_sell and float(db_sell) > 0 and abs(float(db_sell) - float(excel['sell_price'])) > 1:
        price_differs += 1

print(f"يمكن تحديث سعر البيع (كان فارغ): {can_update_sell}")
print(f"يمكن تحديث سعر التكلفة (كان فارغ): {can_update_cost}")
print(f"يمكن تحديث كلا السعرين: {can_update_both}")
print(f"سعر بيع مختلف عن الموجود: {price_differs}")

# === عينة من غير المطابق ===
print(f"\n--- عينة من غير المطابق ({len(unmatched_items)}) ---")
for item in unmatched_items[:15]:
    print(f"  [{item['sheet']}] {item['new_code']} | {item['name'][:40]} | {item['unit']} | bc={item['barcode']} | sell={item['sell_price']}")

# === عينة من المطابق بأسعار جديدة ===
print(f"\n--- عينة من المطابق يمكن تحديث أسعاره ---")
count = 0
for m in matched_items:
    if count >= 10:
        break
    excel = m['excel']
    db = m['db']
    db_sell = db.get('sell_price') or 0
    if excel['sell_price'] and excel['sell_price'] > 0 and (not db_sell or db_sell == 0):
        print(f"  [{m['match_type']}] {excel['name'][:35]} | DB sell={db_sell} → Excel sell={excel['sell_price']} cost={excel['cost_price']}")
        count += 1

conn.close()
