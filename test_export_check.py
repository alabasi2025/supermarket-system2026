import requests, re, urllib3
urllib3.disable_warnings()
base = 'https://localhost:5555'
s = requests.Session()
r = s.get(base+'/login', verify=False, timeout=15)
csrf = re.search(r'name="csrf_token" value="([^"]+)"', r.text).group(1)
s.post(base+'/login', data={'username':'1','password':'1234','csrf_token':csrf}, verify=False, allow_redirects=True)
r = s.get(base+'/stocktake/export/1', verify=False)
print('Export status:', r.status_code)
print('Content-Type:', r.headers.get('Content-Type'))
print('File size:', len(r.content), 'bytes')
with open('test_export_new.xlsx', 'wb') as f:
    f.write(r.content)
import openpyxl
wb = openpyxl.load_workbook('test_export_new.xlsx')
print('Sheets:', wb.sheetnames)
for sheet in wb.sheetnames:
    ws = wb[sheet]
    print(f'Sheet {sheet}: {ws.max_row-1} rows x {ws.max_column} cols')
    print('Headers:', [ws.cell(1,c).value for c in range(1, ws.max_column+1)])
