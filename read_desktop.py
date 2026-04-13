# -*- coding: utf-8 -*-
import os, sys, glob
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

folder = r"C:\Users\qbas\Desktop\مجلد جديد"

print(f"=== Files in: {folder} ===")
for f in os.listdir(folder):
    fp = os.path.join(folder, f)
    size = os.path.getsize(fp) // 1024
    print(f"  {f} ({size} KB)")

print()

# Read each xlsx
for f in os.listdir(folder):
    if not f.endswith('.xlsx'):
        continue
    fp = os.path.join(folder, f)
    print(f"\n{'='*60}")
    print(f"FILE: {f} ({os.path.getsize(fp)//1024} KB)")
    print(f"{'='*60}")
    
    wb = openpyxl.load_workbook(fp, read_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n  Sheet: '{sheet_name}' rows={ws.max_row} cols={ws.max_column}")
        
        # Headers
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1):
            headers = [str(cell.value)[:40] if cell.value else '' for cell in row]
        non_empty = [(i, h) for i, h in enumerate(headers) if h and h != 'None']
        print(f"  Headers ({len(non_empty)} non-empty):")
        for i, h in non_empty:
            print(f"    Col {chr(65+i) if i<26 else str(i+1)} ({i+1}): {h}")
        
        # Sample rows
        print(f"\n  Sample rows:")
        count = 0
        for row in ws.iter_rows(min_row=2, max_row=8):
            vals = [str(cell.value)[:35] if cell.value else '' for cell in row[:len(non_empty)+2]]
            # Only non-empty vals
            vals_clean = [v for v in vals if v]
            if vals_clean:
                print(f"    {vals_clean}")
                count += 1
        
        # Count total rows
        total = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row[:5]):
                total += 1
        print(f"\n  Total data rows: {total}")
    
    wb.close()
