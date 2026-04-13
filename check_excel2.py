# -*- coding: utf-8 -*-
import os, glob, sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

data_dir = r"D:\supermarket-system\data"
files = sorted(glob.glob(os.path.join(data_dir, "*.xlsx")), key=lambda f: os.path.getsize(f), reverse=True)

# Check the largest file
f = files[0]
fname = os.path.basename(f)
print(f"File: {fname} ({os.path.getsize(f)//1024} KB)")

wb = openpyxl.load_workbook(f, read_only=True)
ws = wb[wb.sheetnames[0]]
print(f"Sheet: '{wb.sheetnames[0]}' rows={ws.max_row}")

# Headers
headers = [str(cell.value) if cell.value else '' for cell in next(ws.iter_rows(min_row=1, max_row=1))]
print(f"\nHeaders ({len([h for h in headers if h])}):")
for i, h in enumerate(headers):
    if h:
        print(f"  Col {i+1}: {h}")

# Sample rows
print("\nSample rows:")
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=8)):
    vals = [str(cell.value)[:40] if cell.value else '' for cell in row[:8]]
    print(f"  {vals}")

# Count unique barcodes and names
print("\nCounting unique items...")
barcodes = set()
names = set()
total = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue
    total += 1
    if row[1]:
        barcodes.add(str(row[1]))
    if row[3]:
        names.add(str(row[3]))

print(f"Total rows: {total}")
print(f"Unique barcodes: {len(barcodes)}")
print(f"Unique names: {len(names)}")

wb.close()

# Also check export.xlsx
export_files = [f for f in files if 'export' in os.path.basename(f).lower()]
if export_files:
    f2 = export_files[0]
    print(f"\n\n=== {os.path.basename(f2)} ({os.path.getsize(f2)//1024} KB) ===")
    wb2 = openpyxl.load_workbook(f2, read_only=True)
    ws2 = wb2[wb2.sheetnames[0]]
    headers2 = [str(cell.value) if cell.value else '' for cell in next(ws2.iter_rows(min_row=1, max_row=1))]
    print(f"Headers: {[h for h in headers2 if h]}")
    for i, row in enumerate(ws2.iter_rows(min_row=2, max_row=5)):
        vals = [str(cell.value)[:40] if cell.value else '' for cell in row[:12]]
        print(f"  {vals}")
    wb2.close()
