# -*- coding: utf-8 -*-
"""
مقارنة الشيتات الثلاثة: هل هي نفس الأصناف بوحدات مختلفة؟
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

path = r"C:\Users\qbas\Desktop\مجلد جديد\نسخة من اصناف_مواد_غذائية(1).xlsx"
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

# === جمع البيانات من كل شيت ===

# Sheet1
sheet1_items = {}  # code -> {name, unit, sell, cost, barcode}
ws1 = wb['Sheet1']
for row in ws1.iter_rows(min_row=2, values_only=True):
    if not row[4]:
        continue
    code = str(row[3]).strip() if row[3] else ''  # رقم جديد
    name = str(row[4]).strip()
    unit = str(row[5]).strip() if row[5] else 'حبه'
    barcode = str(row[7]).strip() if row[7] else ''
    cost = row[11] if row[11] and isinstance(row[11], (int, float)) else 0
    sell = row[12] if row[12] and isinstance(row[12], (int, float)) else 0
    pack = row[8] if row[8] else 1
    
    if code not in sheet1_items:
        sheet1_items[code] = []
    sheet1_items[code].append({
        'name': name, 'unit': unit, 'barcode': barcode,
        'cost': cost, 'sell': sell, 'pack': pack
    })

# الجاهز
jahiz_items = {}
ws2 = wb['الجاهز']
for row in ws2.iter_rows(min_row=2, values_only=True):
    if not row[2]:
        continue
    code = str(row[1]).strip() if row[1] else ''
    name = str(row[2]).strip()
    unit = str(row[3]).strip() if row[3] else 'حبه'
    barcode = str(row[4]).strip() if row[4] else ''
    cost = row[8] if row[8] and isinstance(row[8], (int, float)) else 0
    sell = row[9] if row[9] and isinstance(row[9], (int, float)) else 0
    pack = row[5] if row[5] else 1
    
    if code not in jahiz_items:
        jahiz_items[code] = []
    jahiz_items[code].append({
        'name': name, 'unit': unit, 'barcode': barcode,
        'cost': cost, 'sell': sell, 'pack': pack
    })

# ورقة1
warqa_items = {}
ws3 = wb['ورقة1']
for row in ws3.iter_rows(min_row=2, values_only=True):
    if not row[2]:
        continue
    code = str(row[1]).strip() if row[1] else ''
    name = str(row[2]).strip()
    unit = str(row[3]).strip() if row[3] else 'حبه'
    barcode = str(row[4]).strip() if row[4] else ''
    cost = row[8] if row[8] and isinstance(row[8], (int, float)) else 0
    sell = row[9] if row[9] and isinstance(row[9], (int, float)) else 0
    pack = row[5] if row[5] else 1
    
    if code not in warqa_items:
        warqa_items[code] = []
    warqa_items[code].append({
        'name': name, 'unit': unit, 'barcode': barcode,
        'cost': cost, 'sell': sell, 'pack': pack
    })

wb.close()

# === المقارنة ===
print("=" * 70)
print("مقارنة الشيتات")
print("=" * 70)

s1_codes = set(sheet1_items.keys())
j_codes = set(jahiz_items.keys())
w_codes = set(warqa_items.keys())

print(f"\nSheet1: {len(s1_codes)} كود فريد ({sum(len(v) for v in sheet1_items.values())} صف)")
print(f"الجاهز: {len(j_codes)} كود فريد ({sum(len(v) for v in jahiz_items.values())} صف)")
print(f"ورقة1:  {len(w_codes)} كود فريد ({sum(len(v) for v in warqa_items.values())} صف)")

# التقاطع
s1_j = s1_codes & j_codes
s1_w = s1_codes & w_codes
j_w = j_codes & w_codes
all3 = s1_codes & j_codes & w_codes

print(f"\n--- التقاطع ---")
print(f"مشترك بين Sheet1 و الجاهز: {len(s1_j)}")
print(f"مشترك بين Sheet1 و ورقة1:  {len(s1_w)}")
print(f"مشترك بين الجاهز و ورقة1:  {len(j_w)}")
print(f"مشترك في الثلاثة:           {len(all3)}")

# فقط في كل شيت
only_s1 = s1_codes - j_codes - w_codes
only_j = j_codes - s1_codes - w_codes
only_w = w_codes - s1_codes - j_codes

print(f"\nفقط في Sheet1:  {len(only_s1)}")
print(f"فقط في الجاهز:  {len(only_j)}")
print(f"فقط في ورقة1:   {len(only_w)}")

# === تحليل: هل الوحدات مختلفة؟ ===
print(f"\n{'='*70}")
print("تحليل الوحدات للأصناف المشتركة")
print("=" * 70)

# الأصناف المشتركة بين Sheet1 و الجاهز
diff_unit_count = 0
same_unit_count = 0
examples_diff = []
examples_same = []

for code in sorted(s1_j):
    s1_units = set(i['unit'] for i in sheet1_items[code])
    j_units = set(i['unit'] for i in jahiz_items[code])
    
    if s1_units != j_units:
        diff_unit_count += 1
        if len(examples_diff) < 10:
            name = sheet1_items[code][0]['name']
            examples_diff.append((code, name, s1_units, j_units))
    else:
        same_unit_count += 1
        if len(examples_same) < 5:
            name = sheet1_items[code][0]['name']
            examples_same.append((code, name, s1_units, j_units))

print(f"\nSheet1 ↔ الجاهز ({len(s1_j)} مشترك):")
print(f"  وحدات مختلفة: {diff_unit_count} ← هذي أصناف بوحدات متعددة!")
print(f"  وحدات متطابقة: {same_unit_count}")

print(f"\nأمثلة وحدات مختلفة:")
for code, name, s1u, ju in examples_diff:
    s1_info = sheet1_items[code][0]
    j_info = jahiz_items[code][0]
    print(f"  {code}: {name[:45]}")
    print(f"    Sheet1: {s1u} سعر={s1_info['sell']}")
    print(f"    الجاهز: {ju} سعر={j_info['sell']}")

print(f"\nأمثلة وحدات متطابقة:")
for code, name, s1u, ju in examples_same:
    s1_info = sheet1_items[code][0]
    j_info = jahiz_items[code][0]
    print(f"  {code}: {name[:45]}")
    print(f"    Sheet1: {s1u} سعر={s1_info['sell']}")
    print(f"    الجاهز: {ju} سعر={j_info['sell']}")

# === تحليل ورقة1 ===
print(f"\n{'='*70}")
print("تحليل ورقة1 (هل هي وحدات إضافية؟)")
print("=" * 70)

# أصناف بأكثر من صف في ورقة1 (نفس الكود وحدات مختلفة)
multi_in_warqa = {k: v for k, v in warqa_items.items() if len(v) > 1}
print(f"أصناف بأكثر من وحدة في ورقة1: {len(multi_in_warqa)}")

print(f"\nأمثلة:")
for code, items in sorted(multi_in_warqa.items())[:8]:
    print(f"  {code}: {items[0]['name'][:45]}")
    for it in items:
        print(f"    → {it['unit']} | عبوة={it['pack']} | تكلفة={it['cost']} | بيع={it['sell']} | bc={it['barcode'][:15]}")

# ملخص كامل
print(f"\n{'='*70}")
print("الملخص النهائي")
print("=" * 70)
all_codes = s1_codes | j_codes | w_codes
print(f"إجمالي الأكواد الفريدة: {len(all_codes)}")

# كم كود له أكثر من وحدة عبر كل الشيتات
multi_unit_codes = 0
for code in all_codes:
    all_units = set()
    for src in [sheet1_items, jahiz_items, warqa_items]:
        if code in src:
            for item in src[code]:
                all_units.add(item['unit'])
    if len(all_units) > 1:
        multi_unit_codes += 1

print(f"أصناف بوحدة واحدة فقط: {len(all_codes) - multi_unit_codes}")
print(f"أصناف بوحدات متعددة: {multi_unit_codes}")
