# -*- coding: utf-8 -*-
import os, glob, sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

data_dir = r"D:\supermarket-system\data"
files = sorted(glob.glob(os.path.join(data_dir, "*.xlsx")), key=lambda f: os.path.getsize(f), reverse=True)

# Main file
f = files[0]
wb = openpyxl.load_workbook(f, read_only=True)
ws = wb[wb.sheetnames[0]]

headers = [str(cell.value) if cell.value else '' for cell in next(ws.iter_rows(min_row=1, max_row=1))]
print("=== Main file headers ===")
for i, h in enumerate(headers):
    if h:
        print(f"  Col {chr(65+i)} ({i+1}): {h}")

# Analyze structure: check how many units per product name
print("\n=== Analyzing units per product ===")
product_units = {}  # name -> set of (barcode, pack_size)
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[3]:  # no arabic name
        continue
    name = str(row[3]).strip()
    barcode = str(row[1]).strip() if row[1] else ''
    pack_size = str(row[4]).strip() if row[4] else ''
    
    if name not in product_units:
        product_units[name] = []
    product_units[name].append((barcode, pack_size))

# Products with multiple barcodes
multi = {k: v for k, v in product_units.items() if len(v) > 1}
print(f"Total unique names: {len(product_units)}")
print(f"Names with multiple entries: {len(multi)}")

# Show examples
print("\n=== Examples of multi-entry products ===")
count = 0
for name, entries in sorted(multi.items(), key=lambda x: len(x[1]), reverse=True):
    if count >= 10:
        break
    print(f"\n  '{name}' ({len(entries)} entries):")
    for bc, ps in entries:
        print(f"    barcode={bc} pack={ps}")
    count += 1

wb.close()

# Now check export.xlsx more thoroughly
print("\n\n=== export.xlsx analysis ===")
export_files = [f for f in files if 'export' in os.path.basename(f).lower()]
if export_files:
    wb2 = openpyxl.load_workbook(export_files[0], read_only=True)
    ws2 = wb2[wb2.sheetnames[0]]
    
    # Count and check units
    units = {}
    total = 0
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if not row[1]:
            continue
        total += 1
        unit = str(row[2]).strip() if row[2] else 'حبه'
        units[unit] = units.get(unit, 0) + 1
    
    print(f"Total items: {total}")
    print(f"Units breakdown:")
    for u, c in sorted(units.items(), key=lambda x: x[1], reverse=True):
        print(f"  {u}: {c}")
    wb2.close()
