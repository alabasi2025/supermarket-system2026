# -*- coding: utf-8 -*-
import openpyxl, sys, psycopg2, psycopg2.extras
sys.stdout.reconfigure(encoding='utf-8')

# قراءة الملف
wb = openpyxl.load_workbook('D:/supermarket-system/web/new_items.xlsx', data_only=True)
ws = wb.active

print(f"Sheet: {ws.title}")
print(f"Rows: {ws.max_row}")
print(f"Cols: {ws.max_column}")
print()

# Headers
print("=== Headers ===")
headers = []
for col in range(1, ws.max_column + 1):
    val = ws.cell(row=1, column=col).value
    headers.append(val or f'col_{col}')
    print(f"  Col {col}: {val}")

print()
print("=== First 5 rows ===")
for row in range(2, min(7, ws.max_row + 1)):
    vals = []
    for col in range(1, ws.max_column + 1):
        vals.append(str(ws.cell(row=row, column=col).value or ''))
    print(f"  {' | '.join(vals[:10])}")

print()
print("=== Last 3 rows ===")
for row in range(max(2, ws.max_row - 2), ws.max_row + 1):
    vals = []
    for col in range(1, ws.max_column + 1):
        vals.append(str(ws.cell(row=row, column=col).value or ''))
    print(f"  {' | '.join(vals[:10])}")

# جمع الباركودات من الملف
print()
print("=== تحليل الباركودات ===")
barcode_col = None
name_col = None
for i, h in enumerate(headers):
    if h and 'بارك' in str(h).lower() or 'barcode' in str(h).lower():
        barcode_col = i + 1
    if h and ('اسم' in str(h) or 'صنف' in str(h) or 'name' in str(h).lower()):
        name_col = i + 1

if not barcode_col:
    # Try to detect barcode column by content
    for col in range(1, ws.max_column + 1):
        val = str(ws.cell(row=2, column=col).value or '')
        if val.isdigit() and len(val) >= 8:
            barcode_col = col
            break

print(f"  Barcode column: {barcode_col} ({headers[barcode_col-1] if barcode_col else 'NOT FOUND'})")
print(f"  Name column: {name_col} ({headers[name_col-1] if name_col else 'NOT FOUND'})")

if barcode_col:
    file_barcodes = set()
    file_items = []
    for row in range(2, ws.max_row + 1):
        bc = str(ws.cell(row=row, column=barcode_col).value or '').strip()
        name = str(ws.cell(row=row, column=name_col).value or '').strip() if name_col else ''
        if bc and bc != 'None':
            file_barcodes.add(bc)
            file_items.append((bc, name))
    
    print(f"  Total items in file: {len(file_items)}")
    print(f"  Unique barcodes: {len(file_barcodes)}")
    
    # مقارنة مع قاعدة البيانات
    conn = psycopg2.connect(host='localhost', database='supermarket', user='postgres', password='774424555')
    cur = conn.cursor()
    
    # باركودات النظام
    cur.execute("SELECT barcode FROM products WHERE is_active = TRUE AND barcode IS NOT NULL")
    db_barcodes = set(r[0] for r in cur.fetchall())
    cur.execute("SELECT barcode FROM product_barcodes WHERE barcode IS NOT NULL")
    db_barcodes.update(r[0] for r in cur.fetchall())
    
    # also check with leading zero
    db_barcodes_expanded = set()
    for bc in db_barcodes:
        db_barcodes_expanded.add(bc)
        if bc.startswith('0'):
            db_barcodes_expanded.add(bc.lstrip('0'))
        else:
            db_barcodes_expanded.add('0' + bc)
    
    found = file_barcodes & db_barcodes_expanded
    not_found = file_barcodes - db_barcodes_expanded
    
    print(f"\n=== المقارنة ===")
    print(f"  موجود في النظام: {len(found)}")
    print(f"  غير موجود في النظام: {len(not_found)}")
    
    if not_found:
        print(f"\n=== أصناف غير موجودة (أول 20) ===")
        count = 0
        for bc, name in file_items:
            if bc in not_found and count < 20:
                print(f"  {bc} | {name}")
                count += 1
    
    conn.close()
